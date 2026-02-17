import json
import numpy as np
import requests
from sklearn.metrics.pairwise import cosine_similarity
import os
import sys
from tqdm.auto import tqdm
import tkinter as tk
from tkinter import messagebox

# Add the current directory to Python path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from src.core.chromadb_manager import ChromaDBManager
import logging
from src.ai.hypothesis_tools import HypothesisGenerator, HypothesisCritic, MetaHypothesisGenerator, get_lab_goals, get_lab_config, is_ubr5_related
import time
import random
import threading
from collections import deque
import pandas as pd  # For Excel export
from openpyxl import load_workbook
from datetime import datetime
import re

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class HypothesisTimer:
    """Optimized timer for hypothesis critique process with reduced timeout."""

    def __init__(self, timeout_minutes=1):  # Reduced from 5 to 1 minute
        self.timeout_seconds = timeout_minutes * 60
        self.start_time = None
        self.is_expired = False
        self.lock = threading.RLock()  # Use RLock for better thread safety

    def start(self):
        """Start the timer."""
        try:
            with self.lock:
                self.start_time = time.time()
                self.is_expired = False
        except Exception as e:
            logger.warning(f"‚ö†Ô∏è Timer start error: {e}")

    def check_expired(self):
        """Check if the timer has expired."""
        if self.start_time is None:
            return False
        try:
            with self.lock:
                if not self.is_expired:
                    elapsed = time.time() - self.start_time
                    if elapsed >= self.timeout_seconds:
                        self.is_expired = True
                return self.is_expired
        except Exception as e:
            logger.warning(f"‚ö†Ô∏è Timer check error: {e}")
            return True  # Assume expired on error

    def get_remaining_time(self):
        """Get remaining time in seconds."""
        if self.start_time is None:
            return self.timeout_seconds
        try:
            elapsed = time.time() - self.start_time
            return max(0, self.timeout_seconds - elapsed)
        except Exception as e:
            logger.warning(f"‚ö†Ô∏è Timer remaining time error: {e}")
            return 0

    def reset(self):
        """Reset the timer."""
        try:
            with self.lock:
                self.start_time = None
                self.is_expired = False
        except Exception as e:
            logger.warning(f"‚ö†Ô∏è Timer reset error: {e}")

class TokenAwareRateLimiter:
    """Token-aware rate limiter for Gemini API calls with quota protection."""

    def __init__(self, max_requests_per_minute=1000, max_tokens_per_minute=1000000):
        self.max_requests = max_requests_per_minute
        self.max_tokens = max_tokens_per_minute
        self.request_times = deque()
        self.token_usage = deque()  # (tokens, timestamp) pairs
        self.lock = threading.RLock() if 'threading' in globals() else None
        self.max_wait_time = 10  # Maximum wait time in seconds

    def estimate_tokens(self, text):
        """Improved token estimation: ~3.2 characters per token for scientific text."""
        if not text:
            return 1
        # Use full text, not truncated sample
        char_count = len(text)
        # More accurate for scientific text
        base_tokens = char_count / 3.2
        # Add overhead for formatting
        overhead_factor = 1.15
        return int(base_tokens * overhead_factor)

    def wait_if_needed(self, estimated_tokens=0):
        """Wait if we're at the rate limit with intelligent pacing to prevent TPM spikes."""
        try:
            current_time = time.time()

            # Remove old requests (older than 1 minute)
            while self.request_times and current_time - self.request_times[0] > 60:
                self.request_times.popleft()
            
            # Remove old token usage (older than 1 minute)
            while self.token_usage and current_time - self.token_usage[0][1] > 60:
                self.token_usage.popleft()

            # Check token quota with intelligent pacing
            current_tokens = sum(tokens for tokens, _ in self.token_usage)
            token_usage_percentage = current_tokens / self.max_tokens
            
            # Implement progressive pacing based on token usage
            if token_usage_percentage > 0.9:  # 90%+ usage - aggressive pacing
                wait_time = 5.0 + random.uniform(1.0, 3.0)  # 5-8 seconds
                logger.warning(f"üö® High token usage ({token_usage_percentage:.1%}). Aggressive pacing: {wait_time:.1f}s")
            elif token_usage_percentage > 0.8:  # 80%+ usage - moderate pacing
                wait_time = 2.0 + random.uniform(0.5, 1.5)  # 2-3.5 seconds
                logger.info(f"‚ö†Ô∏è Moderate token usage ({token_usage_percentage:.1%}). Moderate pacing: {wait_time:.1f}s")
            elif token_usage_percentage > 0.7:  # 70%+ usage - light pacing
                wait_time = 1.0 + random.uniform(0.2, 0.8)  # 1-1.8 seconds
                logger.info(f"üìä Token usage at {token_usage_percentage:.1%}. Light pacing: {wait_time:.1f}s")
            elif current_tokens + estimated_tokens > self.max_tokens:  # Would exceed limit
                # Calculate wait time based on token usage - be more precise
                if self.token_usage:
                    oldest_token_time = self.token_usage[0][1]
                    time_elapsed = current_time - oldest_token_time
                    wait_time = min(60 - time_elapsed + 1, self.max_wait_time)
                else:
                    wait_time = 1
                
                if wait_time > 0:
                    logger.info(f"‚è≥ Token quota limit reached ({current_tokens:,}/{self.max_tokens:,} tokens). Waiting {wait_time:.1f}s...")
                    time.sleep(wait_time)
                    current_time = time.time()
                    # Clean up old data after waiting
                    while self.token_usage and current_time - self.token_usage[0][1] > 60:
                        self.token_usage.popleft()
                    return  # Skip the pacing logic below

            # Check request rate limit
            elif len(self.request_times) >= self.max_requests:
                if self.request_times:
                    oldest_request_time = self.request_times[0]
                    time_elapsed = current_time - oldest_request_time
                    wait_time = min(60 - time_elapsed + 1, self.max_wait_time)
                else:
                    wait_time = 1
                    
                if wait_time > 0:
                    logger.info(f"‚è≥ Request rate limit reached ({len(self.request_times)}/{self.max_requests}). Waiting {wait_time:.1f}s...")
                    time.sleep(wait_time)
                    current_time = time.time()
                    # Clean up old data after waiting
                    while self.request_times and current_time - self.request_times[0] > 60:
                        self.request_times.popleft()

            # Apply intelligent pacing if we're approaching limits
            if token_usage_percentage > 0.7:
                time.sleep(wait_time)
                current_time = time.time()
                # Clean up old data after pacing
                while self.token_usage and current_time - self.token_usage[0][1] > 60:
                    self.token_usage.popleft()

            # Add current request
            self.request_times.append(current_time)
            if estimated_tokens > 0:
                self.token_usage.append((estimated_tokens, current_time))
            
        except Exception as e:
            logger.warning(f"‚ö†Ô∏è Rate limiter error: {e}")
            # If rate limiter fails, add a conservative delay to prevent quota issues
            time.sleep(2.0)

    def get_remaining_requests(self):
        """Get number of remaining requests in current window."""
        try:
            current_time = time.time()
            while self.request_times and current_time - self.request_times[0] > 60:
                self.request_times.popleft()
            return max(0, self.max_requests - len(self.request_times))
        except Exception as e:
            logger.warning(f"‚ö†Ô∏è Rate limiter count error: {e}")
            return self.max_requests

    def get_remaining_tokens(self):
        """Get number of remaining tokens in current window."""
        try:
            current_time = time.time()
            while self.token_usage and current_time - self.token_usage[0][1] > 60:
                self.token_usage.popleft()
            current_tokens = sum(tokens for tokens, _ in self.token_usage)
            return max(0, self.max_tokens - current_tokens)
        except Exception as e:
            logger.warning(f"‚ö†Ô∏è Token limiter count error: {e}")
            return self.max_tokens

    def execute_with_retry(self, api_call_func, *args, max_retries=5, **kwargs):
        """
        Execute API call with exponential backoff retry logic for quota exceeded errors.
        
        Args:
            api_call_func: Function that makes the API call
            *args: Arguments to pass to the API call function
            max_retries: Maximum number of retry attempts
            **kwargs: Keyword arguments to pass to the API call function
            
        Returns:
            Result of the API call function
            
        Raises:
            Exception: If all retries are exhausted
        """
        last_exception = None
        
        for attempt in range(max_retries + 1):
            try:
                # Wait if needed before making the call
                self.wait_if_needed()
                
                # Make the API call
                result = api_call_func(*args, **kwargs)
                
                # If successful, return the result
                return result
                
            except Exception as e:
                last_exception = e
                error_str = str(e).lower()
                
                # Check if this is a quota exceeded error
                if any(keyword in error_str for keyword in ['quota', '429', 'rate limit', 'exceeded']):
                    if attempt < max_retries:
                        # Calculate exponential backoff delay with jitter
                        base_delay = min(2 ** attempt, 60)  # Cap at 60 seconds
                        jitter = random.uniform(0.1, 0.5)  # Add random jitter
                        delay = base_delay + jitter
                        
                        logger.warning(f"‚ö†Ô∏è Quota exceeded (attempt {attempt + 1}/{max_retries + 1}). Retrying in {delay:.1f}s...")
                        
                        # Wait with exponential backoff
                        time.sleep(delay)
                        
                        # Clean up old data to free up quota
                        current_time = time.time()
                        while self.request_times and current_time - self.request_times[0] > 60:
                            self.request_times.popleft()
                        while self.token_usage and current_time - self.token_usage[0][1] > 60:
                            self.token_usage.popleft()
                        
                        continue
                    else:
                        logger.error(f"‚ùå Quota exceeded after {max_retries} retries. Giving up.")
                        break
                else:
                    # Not a quota error, re-raise immediately
                    logger.error(f"‚ùå Non-quota error: {e}")
                    raise e
        
        # If we get here, all retries were exhausted
        raise last_exception

class EnhancedRAGQuery:
    """
    Enhanced RAG Query System with ChromaDB integration.

    Features:
    - Traditional similarity search (original functionality)
    - ChromaDB vector database for persistent storage
    - Metadata filtering
    - Hybrid search capabilities
    - Performance comparison between methods
    """

    def __init__(self, use_chromadb: bool = True, load_data_at_startup: bool = True):
        """Initialize the enhanced RAG query system."""
        self.use_chromadb = use_chromadb
        self.load_data_at_startup = load_data_at_startup
        self.embeddings_data = None
        self.chroma_manager = None
        self.hypothesis_generator = None
        self.hypothesis_critic = None
        self.meta_hypothesis_generator = None
        self.last_context_chunks = None
        self.hypothesis_records = []  # Track all hypotheses and results for export

        # Track the initial "add" quantity for dynamic chunk selection
        self.initial_add_quantity = None
        self.hypothesis_timer = HypothesisTimer(timeout_minutes=1)  # Reduced to 1-minute timer

        # Track used papers with deprioritization system
        self.used_papers = set()  # Set of paper identifiers (DOI or title) that have been used
        self.paper_usage_count = {}  # Track how many times each paper has been used
        self.deprioritization_factor = 0.3  # Reduce score by 30% for each previous use

        # Package system for collecting search results
        self.current_package = {
            "chunks": [],
            "metadata": [],
            "sources": set(),
            "total_chars": 0
        }

        # Initialize token-aware rate limiter with quota protection
        # Optimized limits for Gemini API - using actual API quotas for maximum performance
        self.gemini_rate_limiter = TokenAwareRateLimiter(max_requests_per_minute=900, max_tokens_per_minute=900000)
        
        # Performance monitoring
        self.query_count = 0
        self.total_query_time = 0
        self.cache = {}  # Simple in-memory cache
        self.cache_max_size = 1000  # Limit cache size
        self.performance_log = []

        print("üöÄ Initializing Enhanced RAG System with Performance Optimizations...")

        # Load API keys first
        try:
            with open("config/keys.json") as f:
                keys = json.load(f)
                self.api_key = keys["GOOGLE_API_KEY"]  # For embeddings
                self.gemini_api_key = keys["GEMINI_API_KEY"]  # For text generation
            print("‚úÖ API keys loaded successfully")
        except Exception as e:
            logger.error(f"‚ùå Failed to load API keys: {e}")
            self.api_key = None
            self.gemini_api_key = None

        # Initialize Gemini client
        try:
            import google.generativeai as genai
            genai.configure(api_key=self.gemini_api_key)
            self.gemini_client = genai
            print("‚úÖ Gemini client initialized successfully")
        except Exception as e:
            logger.error(f"‚ùå Failed to initialize Gemini client: {e}")
            self.gemini_client = None

        # Initialize hypothesis tools after Gemini client is ready
        self._initialize_hypothesis_tools()

        # Initialize ChromaDB if requested
        if self.use_chromadb:
            self._initialize_chromadb()

        # Load embeddings data (optional)
        if self.load_data_at_startup:
            print("üìö Loading embeddings data (this may take a while for large datasets)...")
            self._load_embeddings_data()
        else:
            print("‚è≠Ô∏è  Skipping data loading at startup (will load on-demand)")
            self.embeddings_data = None

        print("‚úÖ Enhanced RAG System initialized!")

    def _initialize_chromadb(self):
        """Initialize ChromaDB manager."""
        try:
            self.chroma_manager = ChromaDBManager()
            if not self.chroma_manager.create_collection():
                logger.error("‚ùå Failed to initialize ChromaDB")
                self.use_chromadb = False
            else:
                logger.info("‚úÖ ChromaDB initialized successfully")
        except Exception as e:
            logger.error(f"‚ùå ChromaDB initialization failed: {e}")
            self.use_chromadb = False

    def _load_embeddings_data(self):
        """Load all available embedding files."""
        # Check if ChromaDB already has data
        if self.use_chromadb and self.chroma_manager:
            stats = self.chroma_manager.get_collection_stats()
            if stats.get('total_documents', 0) > 0:
                logger.info(f"üìö ChromaDB already has {stats.get('total_documents', 0)} documents - skipping data loading")
                self.embeddings_data = None  # We'll use ChromaDB directly
                return

        # Only load embeddings if ChromaDB is empty
        self.embeddings_data = self.load_all_embeddings()

        # If ChromaDB is available, populate it
        if self.use_chromadb and self.embeddings_data:
            self._populate_chromadb()

    def _populate_chromadb(self):
        """Populate ChromaDB with embeddings data."""
        if not self.chroma_manager or not self.embeddings_data:
            return

        try:
            # Check if collection is empty
            stats = self.chroma_manager.get_collection_stats()
            if stats.get('total_documents', 0) > 0:
                logger.info(f"üìö ChromaDB collection already populated with {stats.get('total_documents', 0)} documents")
                return

            logger.info("üîÑ Populating ChromaDB with embeddings data...")

            # Add embeddings from each source
            for source_name, source_stats in self.embeddings_data["sources"].items():
                # Filter data for this source
                source_data = {
                    "chunks": [],
                    "embeddings": [],
                    "metadata": []
                }

                # Add progress bar for metadata processing
                with tqdm(total=len(self.embeddings_data["metadata"]), desc=f"Processing metadata for {source_name}", unit="entry") as meta_pbar:
                    for i, meta in enumerate(self.embeddings_data["metadata"]):
                        if meta.get("source_file") == source_name:
                            source_data["chunks"].append(self.embeddings_data["chunks"][i])
                            source_data["embeddings"].append(self.embeddings_data["embeddings"][i])
                            source_data["metadata"].append(meta)
                        meta_pbar.update(1)

                if source_data["chunks"]:
                    self.chroma_manager.add_embeddings_to_collection(source_data, source_name)
                    logger.info(f"‚úÖ Added {len(source_data['chunks'])} embeddings from {source_name}")

            # Show final stats
            final_stats = self.chroma_manager.get_collection_stats()
            logger.info(f"üéâ ChromaDB populated with {final_stats.get('total_documents', 0)} total documents")

        except Exception as e:
            logger.error(f"‚ùå Failed to populate ChromaDB: {e}")

    def get_google_embedding(self, text):
        """Get embedding for a query text using Google's API."""
        if not self.api_key:
            logger.error("‚ùå API key not available")
            return None

        url = f"https://generativelanguage.googleapis.com/v1beta/models/text-embedding-004:embedContent?key={self.api_key}"
        headers = {"Content-Type": "application/json"}
        data = {"model": "models/text-embedding-004", "content": {"parts": [{"text": text}]}}

        try:
            # Add timeout to prevent hanging
            response = requests.post(url, headers=headers, json=data, timeout=30)
            response.raise_for_status()
            embedding = response.json()["embedding"]["values"]
            return embedding
        except requests.exceptions.Timeout:
            logger.error("‚ùå Query embedding request timed out (30s)")
            return None
        except requests.exceptions.RequestException as e:
            logger.error(f"‚ùå Network error getting query embedding: {e}")
            return None
        except Exception as e:
            logger.error(f"‚ùå Error getting query embedding: {e}")
            return None

    def load_embeddings_from_json(self, filename):
        """Load embeddings from JSON file."""
        if not os.path.exists(filename):
            logger.warning(f"‚ö†Ô∏è Embeddings file '{filename}' not found!")
            return None

        try:
            with open(filename, 'r', encoding='utf-8') as f:
                data = json.load(f)

            logger.info(f"üìä Loaded {len(data['chunks'])} chunks from {filename}")
            return data
        except Exception as e:
            logger.error(f"‚ùå Failed to load {filename}: {e}")
            return None

    def load_all_embeddings(self):
        """Load all available embedding files (both single files and batch-based)."""
        all_data = {
            "chunks": [],
            "embeddings": [],
            "metadata": [],
            "sources": {}
        }

        total_chunks = 0

        # Load PubMed embeddings (single file)
        if os.path.exists("data/embeddings/xrvix_embeddings/pubmed_embeddings.json"):
            logger.info(f"üîÑ Loading pubmed_embeddings.json from data/embeddings/xrvix_embeddings folder...")
            data = self.load_embeddings_from_json("data/embeddings/xrvix_embeddings/pubmed_embeddings.json")
            if data:
                # Add source prefix to metadata with progress bar
                with tqdm(total=len(data["metadata"]), desc="Adding source prefix to PubMed metadata", unit="entry") as meta_pbar:
                    for meta in data["metadata"]:
                        meta["source_file"] = "pubmed"
                        meta_pbar.update(1)

                # Append to combined data
                all_data["chunks"].extend(data["chunks"])
                all_data["embeddings"].extend(data["embeddings"])
                all_data["metadata"].extend(data["metadata"])

                # Track source statistics
                all_data["sources"]["pubmed"] = {
                    "chunks": len(data["chunks"]),
                    "embeddings": len(data["embeddings"]),
                    "stats": data.get("stats", {})
                }

                total_chunks += len(data["chunks"])

        # Load xrvix embeddings (batch-based system)
        xrvix_dir = "data/embeddings/xrvix_embeddings"
        if os.path.exists(xrvix_dir):
            logger.info(f"üîÑ Loading xrvix embeddings from {xrvix_dir}...")

            # Load metadata first
            metadata_file = os.path.join(xrvix_dir, "metadata.json")
            if os.path.exists(metadata_file):
                try:
                    with open(metadata_file, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                    logger.info(f"üìä Found metadata: {metadata.get('total_embeddings', 0)} total embeddings")
                except Exception as e:
                    logger.error(f"‚ùå Failed to load metadata: {e}")
                    metadata = {}
            else:
                metadata = {}

            # Process each source directory (biorxiv, medrxiv, etc.)
            for source_dir in os.listdir(xrvix_dir):
                source_path = os.path.join(xrvix_dir, source_dir)
                if os.path.isdir(source_path) and source_dir in ["biorxiv", "medrxiv"]:
                    logger.info(f"üîÑ Processing {source_dir}...")

                    # Find all batch files
                    batch_files = [f for f in os.listdir(source_path) if f.startswith("batch_") and f.endswith(".json")]
                    batch_files.sort()  # Sort to process in order

                    source_chunks = 0
                    source_embeddings = 0

                    # Add progress bar for batch loading
                    for batch_file in batch_files:
                        batch_path = os.path.join(source_path, batch_file)
                        try:
                            with open(batch_path, 'r', encoding='utf-8') as f:
                                batch_data = json.load(f)

                            # Add source information to metadata with progress bar
                            batch_metadata = batch_data.get("metadata", [])
                            with tqdm(total=len(batch_metadata), desc=f"Adding source info to {source_dir} batch", unit="entry") as meta_pbar:
                                for meta in batch_metadata:
                                    meta["source_file"] = source_dir
                                    meta_pbar.update(1)

                            # Append to combined data
                            batch_chunks = len(batch_data.get("chunks", []))
                            batch_embeddings = len(batch_data.get("embeddings", []))

                            all_data["chunks"].extend(batch_data.get("chunks", []))
                            all_data["embeddings"].extend(batch_data.get("embeddings", []))
                            all_data["metadata"].extend(batch_data.get("metadata", []))

                            source_chunks += batch_chunks
                            source_embeddings += batch_embeddings

                        except Exception as e:
                            logger.error(f"‚ùå Failed to load batch {batch_file}: {e}")

                    # Track source statistics
                    if source_chunks > 0:
                        all_data["sources"][source_dir] = {
                            "chunks": source_chunks,
                            "embeddings": source_embeddings,
                            "batches": len(batch_files),
                            "stats": metadata.get("sources", {}).get(source_dir, {})
                        }
                        total_chunks += source_chunks
                        logger.info(f"‚úÖ Loaded {source_chunks} chunks from {source_dir} ({len(batch_files)} batches)")

        if total_chunks == 0:
            logger.warning("‚ö†Ô∏è No embedding files found!")
            return None

        logger.info(f"üéâ Combined {total_chunks} chunks from all sources")
        return all_data

    def search_chromadb(self, query, top_k=5, filter_dict=None, show_progress=True):
        """Search using ChromaDB."""
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available. Please initialize with ChromaDB enabled.")
            return []

        # Check if ChromaDB has data
        if not self.is_chromadb_ready():
            print("‚ùå ChromaDB is empty. Please load data first using the master processor (option 4).")
            return []

        # Get query embedding
        query_embedding = self.get_google_embedding(query)
        if not query_embedding:
            return []

        # Search in ChromaDB
        results = self.chroma_manager.search_similar(
            query_embedding=query_embedding,
            n_results=top_k,
            where_filter=filter_dict
        )

        return results

    def search_chromadb_with_deprioritization(self, query, top_k=5, filter_dict=None, show_progress=True):
        """Search using ChromaDB with deprioritization of previously used papers."""
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available. Please initialize with ChromaDB enabled.")
            return []

        # Check if ChromaDB has data
        if not self.is_chromadb_ready():
            print("‚ùå ChromaDB is empty. Please load data first using the master processor (option 4).")
            return []

        # Get query embedding
        query_embedding = self.get_google_embedding(query)
        if not query_embedding:
            return []

        # Search with larger result set to account for deprioritization
        search_k = min(top_k * 3, 1000)  # Search 3x more to account for deprioritization
        
        # Search in ChromaDB
        results = self.chroma_manager.search_similar(
            query_embedding=query_embedding,
            n_results=search_k,
            where_filter=filter_dict
        )

        if not results:
            return []

        # Apply deprioritization and re-rank
        deprioritized_results = []
        for i, result in enumerate(results):
            metadata = result.get('metadata', {})
            paper_id = self._get_paper_identifier(metadata)
            
            # Calculate deprioritization factor
            usage_count = self.paper_usage_count.get(paper_id, 0)
            deprioritization_multiplier = (1 - self.deprioritization_factor) ** usage_count
            
            # Apply deprioritization to similarity score
            original_similarity = result.get('similarity', 1.0)
            adjusted_similarity = original_similarity * deprioritization_multiplier
            
            # Create new result with deprioritization info
            deprioritized_result = result.copy()
            deprioritized_result.update({
                'adjusted_similarity': adjusted_similarity,
                'usage_count': usage_count,
                'deprioritization_multiplier': deprioritization_multiplier,
                'paper_id': paper_id,
                'original_rank': i + 1
            })
            
            deprioritized_results.append(deprioritized_result)

        # Sort by adjusted similarity (descending)
        deprioritized_results.sort(key=lambda x: x['adjusted_similarity'], reverse=True)
        
        # Take top_k results
        final_results = deprioritized_results[:top_k]
        
        # Update ranks
        for i, result in enumerate(final_results):
            result['rank'] = i + 1

        if show_progress and self.paper_usage_count:
            used_papers_in_results = sum(1 for r in final_results if r['usage_count'] > 0)
            print(f"üìä Deprioritization: {used_papers_in_results}/{len(final_results)} papers previously used")

        return final_results

    def search_hybrid(self, query, top_k=20, filter_dict=None):
        """Search using ChromaDB only (simplified from hybrid)."""
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available. Please initialize with ChromaDB enabled.")
            return []

        try:
            # Use only ChromaDB search
            results = self.search_chromadb(query, top_k=top_k, filter_dict=filter_dict)

            # Convert ChromaDB results to the expected format
            formatted_results = []
            for result in results:
                formatted_results.append({
                    "chunk": result.get("document", ""),
                    "metadata": result.get("metadata", {}),
                    "similarity": 1.0 - result.get("distance", 0) if result.get("distance") is not None else 0,
                    "method": "chromadb"
                })

            return formatted_results

        except Exception as e:
            print(f"‚ùå ChromaDB search failed: {e}")
            return []

    def display_results(self, results, show_method=True):
        """Display search results in a formatted way."""
        if not results:
            print("[display_results] No results found.")
            return
        print(f"[display_results] Found {len(results)} relevant chunks.")
        total_chars = sum(len(result.get("chunk", "")) for result in results)
        sources = set()
        methods = set()
        for result in results:
            sources.add(result.get("metadata", {}).get("source", "Unknown"))
            methods.add(result.get("method", "Unknown"))
        print(f"[display_results] Summary:")
        print(f"   üìö Total chunks: {len(results)}")
        print(f"   üìù Total characters: {total_chars:,}")
        if show_method:
            print(f"   üîç Search methods: {', '.join(methods)}")
        print(f"   üìñ Sources: {', '.join(sources)}")
        print(f"[display_results] Top 3 results (showing examples):")
        for i, result in enumerate(results[:3], 1):
            chunk = result.get("chunk", "")
            metadata = result.get("metadata", {})
            similarity = result.get("similarity", 0)
            title = metadata.get("title", chunk[:60])
            if len(title) > 60:
                title = title[:57] + "..."
            print(f"   {i}. {title} (similarity: {similarity:.3f})")
        if len(results) > 3:
            print(f"   ... and {len(results) - 3} more chunks")

    def display_statistics(self):
        """Display comprehensive statistics about the knowledge base."""
        print(f"[display_statistics] Knowledge Base Statistics:")
        print("=" * 50)

        # Show ChromaDB stats if available (primary data source)
        if self.use_chromadb and self.chroma_manager:
            print(f"[display_statistics] ChromaDB Statistics (Primary Data Source):")
            chroma_stats = self.chroma_manager.get_collection_stats()
            print(f"   üìä Total documents: {chroma_stats.get('total_documents', 0)}")
            print(f"   üìä Collection name: {chroma_stats.get('collection_name', 'N/A')}")
            print(f"   üìä Available collections: {self.chroma_manager.list_collections()}")

            # Show source breakdown if available
            batch_stats = self.chroma_manager.get_batch_statistics()
            if batch_stats and batch_stats.get('sources'):
                print(f"[display_statistics] Source Breakdown:")
                for source, stats in batch_stats['sources'].items():
                    print(f"   {source}: {stats['total_documents']} documents, {len(stats['batches'])} batches")

        # Show in-memory data stats if available (secondary/fallback)
        if self.embeddings_data:
            print(f"[display_statistics] In-Memory Data Statistics (Secondary):")
            total_chunks = len(self.embeddings_data["chunks"])
            total_embeddings = len(self.embeddings_data["embeddings"])
            print(f"   üìà Total chunks: {total_chunks}")
            print(f"   üìà Total embeddings: {total_embeddings}")
            if total_embeddings > 0:
                print(f"   üìà Embedding dimensions: {len(self.embeddings_data['embeddings'][0])}")
            print(f"[display_statistics] Source Breakdown:")
            for source, stats in self.embeddings_data["sources"].items():
                print(f"   {source}: {stats['chunks']} chunks")
                if "stats" in stats and stats["stats"]:
                    source_stats = stats["stats"]
                    print(f"     - Papers: {source_stats.get('total_papers', 'N/A')}")
                    print(f"     - Embeddings: {source_stats.get('total_embeddings', 'N/A')}")
        else:
            print(f"[display_statistics] In-Memory Data: Not loaded (using ChromaDB directly)")

    def extract_citations_from_chunks(self, context_chunks):
        """Extract citation information from context chunks used for hypothesis critique."""
        citations = []
        unique_sources = set()

        for chunk in context_chunks:
            if isinstance(chunk, dict):
                metadata = chunk.get('metadata', {})
                
                # Extract source information with better fallbacks
                source_name = metadata.get('source_name', metadata.get('source', 'Unknown'))
                title = metadata.get('title', 'No title')
                doi = metadata.get('doi', 'No DOI')
                
                # Try multiple field names for authors
                authors = (metadata.get('author') or 
                         metadata.get('authors') or 
                         metadata.get('author_list') or 
                         'Unknown authors')
                
                # Try multiple field names for journal
                journal = (metadata.get('journal') or 
                          metadata.get('journal_name') or 
                          metadata.get('publication_venue') or 
                          'Unknown journal')
                
                # Extract year from publication_date or other date fields
                publication_date = (metadata.get('publication_date') or 
                                   metadata.get('date') or 
                                   metadata.get('year') or 
                                   '')
                
                if publication_date and len(str(publication_date)) >= 4:
                    year = str(publication_date)[:4]
                else:
                    year = 'Unknown year'

                # Create citation entry with enhanced metadata
                citation = {
                    'source_name': source_name,
                    'title': title,
                    'doi': doi,
                    'authors': authors,
                    'journal': journal,
                    'year': year,
                    'paper_index': metadata.get('paper_index', ''),
                    'para_idx': metadata.get('para_idx', ''),
                    'chunk_length': metadata.get('chunk_length', ''),
                    'citation_count': metadata.get('citation_count', ''),
                    'impact_factor': metadata.get('impact_factor', ''),
                    'chunk_content': chunk.get('document', '') if isinstance(chunk, dict) else str(chunk),
                    'full_metadata': metadata  # Store complete metadata for reference
                }

                # Add to citations if not already present (based on DOI or title)
                citation_key = doi if doi != 'No DOI' else title
                if citation_key not in unique_sources:
                    citations.append(citation)
                    unique_sources.add(citation_key)
            else:
                # Fallback for chunks without metadata
                citations.append({
                    'source_name': 'Unknown',
                    'title': 'Unknown source',
                    'doi': 'No DOI',
                    'authors': 'Unknown authors',
                    'journal': 'Unknown journal',
                    'year': 'Unknown year',
                    'paper_index': '',
                    'para_idx': '',
                    'chunk_length': '',
                    'citation_count': '',
                    'impact_factor': '',
                    'chunk_content': str(chunk),
                    'full_metadata': {}
                })

        return citations

    def format_citations_for_export(self, citations):
        """Format citations for Excel export."""
        if not citations:
            return "No citations available"

        formatted_citations = []
        for citation in citations:
            # Format as academic citation
            if citation['doi'] != 'No DOI':
                formatted = f"{citation['authors']} ({citation['year']}). {citation['title']}. {citation['journal']}. DOI: {citation['doi']}"
            else:
                formatted = f"{citation['authors']} ({citation['year']}). {citation['title']}. {citation['journal']}."
            formatted_citations.append(formatted)

        return "; ".join(formatted_citations)

    def store_citation_chunks_for_hypothesis(self, hypothesis_text, citations, hypothesis_index, export_dir):
        """Store citation chunks in sources folder and create hypothesis-citation mapping."""
        import os
        import json
        from datetime import datetime
        
        # Create hypothesis-specific sources directory
        hypothesis_sources_dir = os.path.join(export_dir, "sources", f"hypothesis_{hypothesis_index}")
        os.makedirs(hypothesis_sources_dir, exist_ok=True)
        
        # Store individual citation files
        citation_files = []
        for i, citation in enumerate(citations):
            citation_filename = f"citation_{i+1}_{citation.get('doi', 'no_doi').replace('/', '_')}.json"
            citation_filepath = os.path.join(hypothesis_sources_dir, citation_filename)
            
            # Create citation data with chunk content
            citation_data = {
                "citation_info": {
                    "source_name": citation.get('source_name', ''),
                    "title": citation.get('title', ''),
                    "doi": citation.get('doi', ''),
                    "authors": citation.get('authors', ''),
                    "journal": citation.get('journal', ''),
                    "year": citation.get('year', ''),
                    "citation_count": citation.get('citation_count', ''),
                    "impact_factor": citation.get('impact_factor', ''),
                    "paper_index": citation.get('paper_index', ''),
                    "para_idx": citation.get('para_idx', ''),
                    "chunk_length": citation.get('chunk_length', '')
                },
                "chunk_content": citation.get('chunk_content', ''),
                "full_metadata": citation.get('full_metadata', {}),
                "hypothesis_association": {
                    "hypothesis_index": hypothesis_index,
                    "hypothesis_text": hypothesis_text,
                    "associated_at": datetime.now().isoformat()
                }
            }
            
            # Save citation file
            with open(citation_filepath, 'w', encoding='utf-8') as f:
                json.dump(citation_data, f, indent=2, ensure_ascii=False)
            
            citation_files.append({
                "filename": citation_filename,
                "filepath": citation_filepath,
                "citation_info": citation_data["citation_info"]
            })
        
        # Create hypothesis-citation mapping file
        mapping_data = {
            "hypothesis": {
                "index": hypothesis_index,
                "text": hypothesis_text,
                "total_citations": len(citations),
                "created_at": datetime.now().isoformat()
            },
            "citations": citation_files,
            "citation_cache_keys": [
                f"{citation.get('source_name', 'unknown')}_{citation.get('paper_index', 'unknown')}"
                for citation in citations
            ]
        }
        
        mapping_filepath = os.path.join(hypothesis_sources_dir, "hypothesis_citation_mapping.json")
        with open(mapping_filepath, 'w', encoding='utf-8') as f:
            json.dump(mapping_data, f, indent=2, ensure_ascii=False)
        
        return mapping_filepath, citation_files

    def _parse_citations_from_text(self, citations_text):
        """Parse citations from formatted text back to structured format."""
        if not citations_text or citations_text == 'No citations available':
            return []
        
        citations = []
        # Split by semicolon to get individual citations
        citation_parts = citations_text.split('; ')
        
        for citation_text in citation_parts:
            citation_text = citation_text.strip()
            if not citation_text:
                continue
                
            # Try to parse the formatted citation
            # Format: "Authors (Year). Title. Journal. DOI: doi"
            citation = {
                'source_name': 'Unknown',
                'title': 'Unknown title',
                'doi': 'No DOI',
                'authors': 'Unknown authors',
                'journal': 'Unknown journal',
                'year': 'Unknown year',
                'paper_index': '',
                'para_idx': '',
                'chunk_length': '',
                'citation_count': '',
                'impact_factor': '',
                'chunk_content': '',
                'full_metadata': {}
            }
            
            # Extract DOI if present
            if 'DOI:' in citation_text:
                doi_part = citation_text.split('DOI:')[-1].strip()
                citation['doi'] = doi_part
            
            # Extract year (first 4-digit number in parentheses)
            import re
            year_match = re.search(r'\((\d{4})\)', citation_text)
            if year_match:
                citation['year'] = year_match.group(1)
            
            # Extract authors (text before the year)
            if year_match:
                authors_part = citation_text[:year_match.start()].strip()
                citation['authors'] = authors_part
            
            # Extract title and journal (text between year and DOI)
            if year_match and 'DOI:' in citation_text:
                middle_part = citation_text[year_match.end():citation_text.find('DOI:')].strip()
                # Split by periods to separate title and journal
                parts = middle_part.split('.')
                if len(parts) >= 2:
                    citation['title'] = parts[0].strip()
                    citation['journal'] = parts[1].strip()
                elif len(parts) == 1:
                    citation['title'] = parts[0].strip()
            
            citations.append(citation)
        
        return citations

    def retrieve_relevant_chunks(self, query, top_k=1500, lab_paper_ratio=None, use_filtering=True):
        """Retrieve the most relevant chunks from all loaded batches using ChromaDB, with lab-authored papers included."""
        if lab_paper_ratio is None:
            lab_paper_ratio = getattr(self, 'lab_paper_ratio', 0.2)
        
        # Use filtering by default for better quality
        if use_filtering:
            return self.retrieve_relevant_chunks_with_filtering(query, top_k, lab_paper_ratio)
        else:
            return self.retrieve_relevant_chunks_with_lab_papers(query, top_k, lab_paper_ratio)

    def retrieve_relevant_chunks_with_deprioritization(self, query, top_k=1500, lab_paper_ratio=None, preferred_authors=None, use_filtering=True):
        """
        Retrieve relevant chunks with deprioritization of previously used papers.
        This ensures maximum citation diversity across hypotheses.
        """
        if lab_paper_ratio is None:
            lab_paper_ratio = getattr(self, 'lab_paper_ratio', 0.2)
        
        print(f"üéØ Retrieving chunks with deprioritization for maximum citation diversity...")
        print(f"üìä Deprioritization factor: {self.deprioritization_factor * 100}% reduction per use")
        
        # Use deprioritization method
        chunks = self.retrieve_relevant_chunks_with_lab_papers_deprioritized(
            query, top_k, lab_paper_ratio, preferred_authors
        )
        
        # Apply filtering if requested
        if use_filtering and chunks:
            print(f"üîç Applying quality filtering to deprioritized chunks...")
            query_keywords = query.lower().split()
            filtered_chunks = self._filter_context_chunks(chunks, query_keywords)
            return filtered_chunks
            
        return chunks

    def select_dynamic_chunks_for_generation(self, query, num_chunks=None, lab_paper_ratio=None, randomization_strategy='enhanced'):
        """
        Select new chunks from ChromaDB for hypothesis generation, ensuring diversity across sources and including lab-authored papers.
        
        Args:
            query: Search query
            num_chunks: Number of chunks to select (defaults to initial_add_quantity)
            lab_paper_ratio: Ratio of lab-authored papers to include
            randomization_strategy: Strategy for chunk randomization ('enhanced', 'shuffle', 'diversity', 'time_based')
        """
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available for dynamic chunk selection.")
            return None

        # Use configured lab paper ratio if not specified
        if lab_paper_ratio is None:
            lab_paper_ratio = getattr(self, 'lab_paper_ratio', None)

        # Use initial add quantity if not specified
        if num_chunks is None:
            if self.initial_add_quantity is None:
                print("‚ö†Ô∏è  No initial add quantity stored, using default of 1500 chunks")
                num_chunks = 1500
            else:
                num_chunks = self.initial_add_quantity
                print(f"üîÑ Selecting {num_chunks} new chunks from ChromaDB (based on initial add quantity)")

        try:
            # Get new chunks from ChromaDB using the original query
            if "prompt" in self.current_package and self.current_package["prompt"]:
                search_query = self.current_package["prompt"]
            else:
                search_query = query

            print(f"üîç Searching ChromaDB for '{search_query}' to select {num_chunks} new chunks...")
            print(f"üéØ Ensuring diversity across sources: pubmed, biorxiv, medrxiv")
            if lab_paper_ratio is None:
                print(f"üî¨ Lab papers: AUTO (automatically determined based on availability)")
            else:
                print(f"üî¨ Including {int(num_chunks * lab_paper_ratio)} lab-authored papers")

            # Use the sophisticated lab paper search method with filtering
            preferred_authors = getattr(self, 'preferred_authors', [])
            results = self.retrieve_relevant_chunks_with_filtering(
                search_query,
                top_k=min(num_chunks * 3, 5000),  # Search up to 3x the needed amount, max 5000
                lab_paper_ratio=lab_paper_ratio,
                preferred_authors=preferred_authors,
                query_keywords=search_query.lower().split()
            )

            if not results:
                print("‚ùå No new chunks found in ChromaDB.")
                return None

            # Apply randomization strategy
            results = self._apply_randomization_strategy(results, randomization_strategy)

            # Organize results by source for balanced selection
            source_chunks = {
                'pubmed': [],
                'biorxiv': [],
                'medrxiv': []
            }

            # Categorize results by source
            for result in results:
                metadata = result.get('metadata', {})
                source = metadata.get('source', '').lower()
                source_name = metadata.get('source_name', '').lower()

                # Determine the source
                if 'pubmed' in source or 'pubmed' in source_name:
                    source_key = 'pubmed'
                elif 'biorxiv' in source or 'biorxiv' in source_name:
                    source_key = 'biorxiv'
                elif 'medrxiv' in source or 'medrxiv' in source_name:
                    source_key = 'medrxiv'
                else:
                    # Default to pubmed if source is unclear
                    source_key = 'pubmed'

                source_chunks[source_key].append(result)

            # Calculate target chunks per source (balanced distribution)
            chunks_per_source = max(1, num_chunks // 3)  # At least 1 chunk per source
            remaining_chunks = num_chunks % 3

            print(f"üìä Source distribution target: ~{chunks_per_source} chunks per source")

            # Filter out chunks from already used papers and select balanced chunks
            new_chunks = []
            used_papers_in_this_selection = set()
            source_counts = {'pubmed': 0, 'biorxiv': 0, 'medrxiv': 0}

            # Select chunks from each source, avoiding used papers
            for source_key in ['pubmed', 'biorxiv', 'medrxiv']:
                source_target = chunks_per_source + (1 if remaining_chunks > 0 else 0)
                remaining_chunks = max(0, remaining_chunks - 1)

                source_results = source_chunks[source_key]
                source_selected = 0

                for result in source_results:
                    if source_selected >= source_target:
                        break

                    metadata = result.get('metadata', {})
                    paper_id = self._get_paper_identifier(metadata)

                    # Skip if this paper has been used before
                    if paper_id in self.used_papers:
                        continue

                    # Add to new chunks and mark as used
                    new_chunks.append(result['document'])
                    used_papers_in_this_selection.add(paper_id)
                    source_counts[source_key] += 1
                    source_selected += 1

            # If we don't have enough chunks from balanced selection, fill with any available chunks
            if len(new_chunks) < num_chunks:
                print(f"‚ö†Ô∏è  Balanced selection only found {len(new_chunks)} chunks, filling with additional chunks...")

                # Collect all remaining unused chunks
                remaining_chunks = []
                for result in results:
                    metadata = result.get('metadata', {})
                    paper_id = self._get_paper_identifier(metadata)

                    if paper_id not in self.used_papers and paper_id not in used_papers_in_this_selection:
                        remaining_chunks.append(result)

                # Add remaining chunks until we reach the target
                for result in remaining_chunks:
                    if len(new_chunks) >= num_chunks:
                        break

                    metadata = result.get('metadata', {})
                    paper_id = self._get_paper_identifier(metadata)

                    new_chunks.append(result['document'])
                    used_papers_in_this_selection.add(paper_id)

                    # Track source for remaining chunks
                    source = metadata.get('source', '').lower()
                    source_name = metadata.get('source_name', '').lower()
                    if 'biorxiv' in source or 'biorxiv' in source_name:
                        source_counts['biorxiv'] += 1
                    elif 'medrxiv' in source or 'medrxiv' in source_name:
                        source_counts['medrxiv'] += 1
                    else:
                        source_counts['pubmed'] += 1

            # Add the newly used papers to the tracking set
            self.used_papers.update(used_papers_in_this_selection)

            print(f"‚úÖ Selected {len(new_chunks)} new chunks from {len(used_papers_in_this_selection)} previously unused papers")
            print(f"üìä Source distribution: pubmed={source_counts['pubmed']}, biorxiv={source_counts['biorxiv']}, medrxiv={source_counts['medrxiv']}")
            print(f"üìä Total papers used so far: {len(self.used_papers)}")

            if len(new_chunks) < num_chunks:
                print(f"‚ö†Ô∏è  Only found {len(new_chunks)} new chunks (requested {num_chunks})")
                if len(self.used_papers) > 100:  # Arbitrary threshold
                    print("üí° Consider resetting used papers list with 'reset_papers' command")

            # Store the results for citation extraction
            self.last_dynamic_results = []
            for result in results:
                if result.get('chunk') in new_chunks:
                    self.last_dynamic_results.append(result)

            return new_chunks

        except Exception as e:
            print(f"‚ùå Error selecting dynamic chunks: {e}")
            return None

    def select_diverse_papers_for_hypothesis_generation(self, query, target_papers=1500, lab_paper_ratio=None, preferred_authors=None):
        """
        Select exactly 1500 different papers for hypothesis generation, ensuring maximum diversity
        while maintaining the existing prioritization system (lab author, date, citation count, impact factor).
        
        This function ensures that each hypothesis generation gets a completely different set of papers,
        maximizing the variety of source chunks as required by the new prompt format.
        
        Args:
            query: Search query
            target_papers: Number of different papers to select (default: 1500)
            lab_paper_ratio: Ratio of lab-authored papers to include
            preferred_authors: List of preferred author names to prioritize
            
        Returns:
            List of chunks from exactly target_papers different papers
        """
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available for diverse paper selection.")
            return None

        print(f"üéØ Selecting exactly {target_papers} different papers for hypothesis generation...")
        print(f"üîç Ensuring maximum diversity across all sources and papers")

        try:
            # Use configured lab paper ratio if not specified
            if lab_paper_ratio is None:
                lab_paper_ratio = getattr(self, 'lab_paper_ratio', None)

            # Search for a large pool of papers to ensure we have enough diversity
            # We need significantly more than target_papers to account for filtering and prioritization
            search_pool_size = max(target_papers * 5, 10000)  # At least 5x the target, minimum 10k
            
            print(f"üîç Searching ChromaDB for '{query}' to create pool of {search_pool_size} papers...")
            
            # Use the sophisticated lab paper search method to get a large pool
            preferred_authors = getattr(self, 'preferred_authors', [])
            search_results = self.retrieve_relevant_chunks_with_lab_papers(
                query,
                top_k=search_pool_size,
                lab_paper_ratio=lab_paper_ratio,
                preferred_authors=preferred_authors
            )

            if not search_results:
                print("‚ùå No papers found in ChromaDB.")
                return None

            print(f"üìä Found {len(search_results)} papers in search pool")

            # Step 1: Group results by paper (one chunk per paper)
            paper_chunks = {}  # paper_id -> best_chunk
            paper_metadata = {}  # paper_id -> metadata
            
            for result in search_results:
                metadata = result.get('metadata', {})
                paper_id = self._get_paper_identifier(metadata)
                
                if not paper_id:
                    continue
                    
                # Store the best chunk for each paper (first one found)
                if paper_id not in paper_chunks:
                    paper_chunks[paper_id] = result['document']
                    paper_metadata[paper_id] = metadata

            print(f"üìä Identified {len(paper_chunks)} unique papers")

            # Step 2: Apply existing prioritization system
            # Categorize papers by priority level
            lab_papers = []
            preferred_author_papers = []
            other_papers = []
            
            for paper_id, metadata in paper_metadata.items():
                if self.is_lab_authored_paper(metadata):
                    lab_papers.append((paper_id, metadata))
                elif preferred_authors and self._is_preferred_author_paper(metadata, preferred_authors):
                    preferred_author_papers.append((paper_id, metadata))
                else:
                    other_papers.append((paper_id, metadata))

            print(f"üìä Priority categorization: {len(lab_papers)} lab papers, {len(preferred_author_papers)} preferred author papers, {len(other_papers)} other papers")

            # Step 3: Calculate prioritization scores for other papers
            # Extract citation and journal data for scoring
            all_citations = []
            journal_impact_factors = {}
            
            for paper_id, metadata in paper_metadata.items():
                citation_count = metadata.get('citation_count', 'not found')
                journal = metadata.get('journal', 'Unknown journal')
                
                if citation_count != 'not found':
                    try:
                        all_citations.append(int(citation_count))
                    except (ValueError, TypeError):
                        pass
                
                if journal not in journal_impact_factors:
                    journal_impact_factors[journal] = []
                if citation_count != 'not found':
                    try:
                        journal_impact_factors[journal].append(int(citation_count))
                    except (ValueError, TypeError):
                        pass

            # Calculate statistics for scoring
            avg_citations = np.mean(all_citations) if all_citations else 0
            journal_impact_scores = {}
            for journal, citations in journal_impact_factors.items():
                if citations:
                    journal_impact_scores[journal] = np.mean(citations)

            # Step 4: Score and sort other papers
            def score_paper(paper_id, metadata):
                journal = metadata.get('journal', 'Unknown journal')
                citation_count = metadata.get('citation_count', 'not found')
                
                # Journal impact score (0-1 scale)
                journal_score = 0
                if journal in journal_impact_scores:
                    max_impact = max(journal_impact_scores.values()) if journal_impact_scores else 1
                    journal_score = journal_impact_scores[journal] / max_impact if max_impact > 0 else 0
                
                # Citation score (0-1 scale)
                citation_score = 0
                if citation_count != 'not found':
                    try:
                        citations = int(citation_count)
                        citation_score = min(citations / avg_citations, 2.0) if avg_citations > 0 else 0
                    except (ValueError, TypeError):
                        pass
                
                # Combined score (weighted average)
                combined_score = (journal_score * 0.4) + (citation_score * 0.6)
                return combined_score

            # Score and sort other papers
            other_papers_scored = [(paper_id, metadata, score_paper(paper_id, metadata)) for paper_id, metadata in other_papers]
            other_papers_scored.sort(key=lambda x: x[2], reverse=True)

            # Step 5: Select papers ensuring diversity with temporal balance
            selected_papers = set()
            selected_chunks = []
            
            # First, add all lab papers (highest priority)
            lab_papers_to_add = min(len(lab_papers), int(target_papers * 0.3))  # Up to 30% lab papers
            for paper_id, metadata in lab_papers[:lab_papers_to_add]:
                if paper_id not in selected_papers:
                    selected_papers.add(paper_id)
                    selected_chunks.append(paper_chunks[paper_id])
                    if len(selected_papers) >= target_papers:
                        break

            # Next, add preferred author papers
            preferred_author_papers_to_add = min(len(preferred_author_papers), int(target_papers * 0.2))  # Up to 20% preferred author papers
            for paper_id, metadata in preferred_author_papers[:preferred_author_papers_to_add]:
                if paper_id not in selected_papers:
                    selected_papers.add(paper_id)
                    selected_chunks.append(paper_chunks[paper_id])
                    if len(selected_papers) >= target_papers:
                        break

            # Calculate remaining slots for other papers
            remaining_needed = target_papers - len(selected_papers)
            
            # For other papers, ensure temporal balance: 50% from oldest 25% and newest 25%
            if remaining_needed > 0 and other_papers_scored:
                # Extract publication years for temporal analysis
                paper_years = []
                for paper_id, metadata, score in other_papers_scored:
                    publication_date = metadata.get('publication_date', '')
                    if publication_date and len(publication_date) >= 4:
                        try:
                            year = int(publication_date[:4])
                            paper_years.append((paper_id, metadata, score, year))
                        except (ValueError, TypeError):
                            # If we can't determine year, add with year 0 (will be treated as unknown)
                            paper_years.append((paper_id, metadata, score, 0))
                    else:
                        # No publication date, add with year 0
                        paper_years.append((paper_id, metadata, score, 0))
                
                if paper_years:
                    # Sort by year (oldest first)
                    paper_years.sort(key=lambda x: x[3])
                    
                    # Calculate temporal quartiles
                    total_other_papers = len(paper_years)
                    q1_size = max(1, total_other_papers // 4)  # Bottom 25% (oldest)
                    q4_size = max(1, total_other_papers // 4)  # Top 25% (newest)
                    
                    # Calculate how many papers to take from each temporal quartile
                    # We want 50% of remaining_needed from oldest+newest, 50% from middle
                    temporal_balance_count = min(remaining_needed // 2, q1_size + q4_size)
                    middle_count = remaining_needed - temporal_balance_count
                    
                    print(f"üìÖ Temporal balance: {temporal_balance_count} papers from oldest/newest quartiles, {middle_count} from middle")
                    
                    # Select from oldest papers (bottom 25%)
                    oldest_papers = paper_years[:q1_size]
                    oldest_selected = 0
                    for paper_id, metadata, score, year in oldest_papers:
                        if oldest_selected >= temporal_balance_count // 2:
                            break
                        if paper_id not in selected_papers:
                            selected_papers.add(paper_id)
                            selected_chunks.append(paper_chunks[paper_id])
                            oldest_selected += 1
                            if len(selected_papers) >= target_papers:
                                break
                    
                    # Select from newest papers (top 25%)
                    newest_papers = paper_years[-q4_size:]
                    newest_selected = 0
                    for paper_id, metadata, score, year in newest_papers:
                        if newest_selected >= temporal_balance_count // 2:
                            break
                        if paper_id not in selected_papers:
                            selected_papers.add(paper_id)
                            selected_chunks.append(paper_chunks[paper_id])
                            newest_selected += 1
                            if len(selected_papers) >= target_papers:
                                break
                    
                    # Fill remaining slots with middle-aged papers (prioritized by score)
                    middle_papers = paper_years[q1_size:-q4_size]  # Middle 50%
                    middle_papers.sort(key=lambda x: x[2], reverse=True)  # Sort by score (highest first)
                    
                    for paper_id, metadata, score, year in middle_papers:
                        if len(selected_papers) >= target_papers:
                            break
                        if paper_id not in selected_papers:
                            selected_papers.add(paper_id)
                            selected_chunks.append(paper_chunks[paper_id])
                    
                    print(f"üìä Temporal distribution: {oldest_selected} oldest, {newest_selected} newest, {len(selected_papers) - lab_papers_to_add - preferred_author_papers_to_add - oldest_selected - newest_selected} middle-aged")
                else:
                    # Fallback: no temporal data available, use original scoring
                    for paper_id, metadata, score in other_papers_scored[:remaining_needed]:
                        if paper_id not in selected_papers:
                            selected_papers.add(paper_id)
                            selected_chunks.append(paper_chunks[paper_id])
                            if len(selected_papers) >= target_papers:
                                break

            # Step 6: Ensure we have exactly target_papers
            if len(selected_papers) < target_papers:
                print(f"‚ö†Ô∏è  Only found {len(selected_papers)} unique papers (requested {target_papers})")
                print("üîÑ Attempting to fill remaining slots with any available papers...")
                
                # Fill remaining slots with any available papers
                remaining_papers = [p for p in other_papers_scored if p[0] not in selected_papers]
                for paper_id, metadata, score in remaining_papers:
                    if len(selected_papers) >= target_papers:
                        break
                    selected_papers.add(paper_id)
                    selected_chunks.append(paper_chunks[paper_id])

            # Step 7: Update used papers tracking
            self.used_papers.update(selected_papers)

            print(f"‚úÖ Successfully selected exactly {len(selected_papers)} different papers")
            print(f"üìä Lab papers: {len([p for p in selected_papers if p in [pid for pid, _ in lab_papers[:lab_papers_to_add]]])}")
            print(f"üìä Preferred author papers: {len([p for p in selected_papers if p in [pid for pid, _ in preferred_author_papers[:preferred_author_papers_to_add]]])}")
            
            # Calculate temporal distribution for other papers
            other_papers_count = len(selected_papers) - len([p for p in selected_papers if p in [pid for pid, _ in lab_papers[:lab_papers_to_add]]]) - len([p for p in selected_papers if p in [pid for pid, _ in preferred_author_papers[:preferred_author_papers_to_add]]])
            print(f"üìä Other papers: {other_papers_count}")
            
            if other_papers_count > 0:
                # Re-analyze the temporal distribution for better reporting
                other_paper_years = []
                for paper_id in selected_papers:
                    if (paper_id not in [pid for pid, _ in lab_papers[:lab_papers_to_add]] and 
                        paper_id not in [pid for pid, _ in preferred_author_papers[:preferred_author_papers_to_add]]):
                        metadata = paper_metadata[paper_id]
                        publication_date = metadata.get('publication_date', '')
                        if publication_date and len(publication_date) >= 4:
                            try:
                                year = int(publication_date[:4])
                                other_paper_years.append(year)
                            except (ValueError, TypeError):
                                pass
                
                if other_paper_years:
                    other_paper_years.sort()
                    total_other = len(other_paper_years)
                    q1_idx = max(0, total_other // 4 - 1)
                    q3_idx = min(total_other - 1, 3 * total_other // 4)
                    
                    oldest_papers = len([y for y in other_paper_years if y <= other_paper_years[q1_idx]])
                    newest_papers = len([y for y in other_paper_years if y >= other_paper_years[q3_idx]])
                    middle_papers = total_other - oldest_papers - newest_papers
                    
                    print(f"   üìÖ Temporal distribution of other papers:")
                    print(f"      ‚Ä¢ Oldest 25%: {oldest_papers} papers (‚â§{other_paper_years[q1_idx] if oldest_papers > 0 else 'N/A'})")
                    print(f"      ‚Ä¢ Middle 50%: {middle_papers} papers")
                    print(f"      ‚Ä¢ Newest 25%: {newest_papers} papers (‚â•{other_paper_years[q3_idx] if newest_papers > 0 else 'N/A'})")
                    print(f"      ‚Ä¢ Year range: {min(other_paper_years)} - {max(other_paper_years)}")
            
            print(f"üìä Total papers used so far: {len(self.used_papers)}")

            # Store the results for citation extraction
            self.last_diverse_results = []
            for paper_id in selected_papers:
                metadata = paper_metadata[paper_id]
                chunk = paper_chunks[paper_id]
                self.last_diverse_results.append({
                    'chunk': chunk,
                    'metadata': metadata
                })

            return selected_chunks

        except Exception as e:
            print(f"‚ùå Error selecting diverse papers: {e}")
            return None

    def _filter_context_chunks(self, context_chunks, query_keywords=None):
        """
        Filter and clean context chunks to improve quality for hypothesis generation.
        Removes low-quality chunks and ensures diverse, relevant content.
        Preserves metadata structure for citation extraction.
        """
        filtered_chunks = []
        
        for chunk in context_chunks:
            if isinstance(chunk, dict):
                text = chunk.get("document", "")
                metadata = chunk.get("metadata", {})
            else:
                text = str(chunk)
                metadata = {}
            
            # Skip empty or very short chunks
            if not text or len(text.strip()) < 50:
                continue
                
            # Skip chunks that are mostly metadata or formatting
            if text.count('\n') > len(text) / 20:  # Too many line breaks
                continue
                
            # Skip chunks that are mostly numbers or special characters
            alpha_ratio = sum(c.isalpha() for c in text) / len(text) if text else 0
            if alpha_ratio < 0.6:  # Less than 60% alphabetic characters
                continue
            
            # Create filtered chunk preserving metadata structure
            filtered_chunk = {
                "document": text,
                "metadata": metadata
            }
                
            # Prioritize chunks with query-relevant content if keywords provided
            if query_keywords:
                text_lower = text.lower()
                relevance_score = sum(1 for keyword in query_keywords if keyword.lower() in text_lower)
                if relevance_score > 0:
                    filtered_chunks.insert(0, filtered_chunk)  # Add to beginning for priority
                else:
                    filtered_chunks.append(filtered_chunk)
            else:
                # General quality filtering without specific keyword prioritization
                filtered_chunks.append(filtered_chunk)
        
        # Limit to reasonable number of chunks to avoid token limits
        max_chunks = 1000  # Reduced from 1500 to improve quality
        if len(filtered_chunks) > max_chunks:
            # Keep the first max_chunks (prioritizing relevant content)
            filtered_chunks = filtered_chunks[:max_chunks]
            
        print(f"üìä Filtered context: {len(filtered_chunks)} high-quality chunks (from {len(context_chunks)} original)")
        return filtered_chunks

    def _is_chunk_high_quality(self, chunk_text, query_keywords=None):
        """
        Check if a chunk meets quality standards for hypothesis generation.
        Returns True if chunk is high quality, False otherwise.
        """
        if not chunk_text or len(chunk_text.strip()) < 50:
            return False
            
        # Skip chunks that are mostly metadata or formatting
        if chunk_text.count('\n') > len(chunk_text) / 20:  # Too many line breaks
            return False
            
        # Skip chunks that are mostly numbers or special characters
        alpha_ratio = sum(c.isalpha() for c in chunk_text) / len(chunk_text) if chunk_text else 0
        if alpha_ratio < 0.6:  # Less than 60% alphabetic characters
            return False
            
        # Check for query relevance if keywords provided
        if query_keywords:
            text_lower = chunk_text.lower()
            relevance_score = sum(1 for keyword in query_keywords if keyword.lower() in text_lower)
            return relevance_score > 0
            
        return True

    def _get_replacement_chunks(self, query, num_needed, excluded_chunks=None, query_keywords=None, lab_paper_ratio=None, preferred_authors=None):
        """
        Get replacement chunks from ChromaDB using sophisticated search algorithm.
        Applies lab paper prioritization, citation analysis, and temporal balance.
        """
        if not self.use_chromadb or not self.chroma_manager:
            return []
            
        if excluded_chunks is None:
            excluded_chunks = set()
            
        try:
            print(f"üîç Searching for {num_needed} replacement chunks with sophisticated algorithm...")
            
            # Use sophisticated search to get high-quality replacement chunks
            # Search for more chunks to allow for filtering and exclusion
            search_results = self.sophisticated_lab_paper_search(
                query, 
                top_k=num_needed * 4,  # Get 4x more to account for filtering and exclusions
                lab_paper_ratio=lab_paper_ratio,
                preferred_authors=preferred_authors
            )
            
            if not search_results:
                print("‚ùå No replacement chunks found")
                return []
                
            print(f"üìö Found {len(search_results)} potential replacement chunks")
            
            # Process results and apply sophisticated filtering
            replacement_chunks = []
            lab_papers = []
            preferred_author_papers = []
            other_papers = []
            
            # Track citation statistics for analysis
            all_citations = []
            journal_impact_factors = {}
            
            for chunk in search_results:
                if isinstance(chunk, dict):
                    chunk_text = chunk.get("document", "")
                    metadata = chunk.get("metadata", {})
                else:
                    chunk_text = str(chunk)
                    metadata = {}
                
                # Create robust exclusion ID using content hash
                chunk_id = hash(chunk_text.strip())
                
                # Skip if already excluded
                if chunk_id in excluded_chunks:
                    continue
                    
                # Check quality first
                if not self._is_chunk_high_quality(chunk_text, query_keywords):
                    continue
                
                # Extract citation count for analysis
                citation_count = metadata.get('citation_count', 'not found')
                if citation_count != 'not found':
                    try:
                        citations = int(citation_count)
                        all_citations.append(citations)
                    except (ValueError, TypeError):
                        pass
                
                # Track journal for impact factor analysis
                journal = metadata.get('journal', 'Unknown journal')
                if journal not in journal_impact_factors:
                    journal_impact_factors[journal] = []
                if citation_count != 'not found':
                    try:
                        journal_impact_factors[journal].append(int(citation_count))
                    except (ValueError, TypeError):
                        pass
                
                # Categorize papers
                if self.is_lab_authored_paper(metadata):
                    lab_papers.append((chunk_text, metadata, citation_count))
                elif preferred_authors and self._is_preferred_author_paper(metadata, preferred_authors):
                    preferred_author_papers.append((chunk_text, metadata, citation_count))
                else:
                    other_papers.append((chunk_text, metadata, citation_count))
            
            print(f"üìä Replacement candidates: {len(lab_papers)} lab papers, {len(preferred_author_papers)} preferred author papers, {len(other_papers)} other papers")
            
            # Calculate statistics for prioritization
            avg_citations = np.mean(all_citations) if all_citations else 0
            median_citations = np.median(all_citations) if all_citations else 0
            
            # Calculate journal impact factors
            journal_impact_scores = {}
            for journal, citations in journal_impact_factors.items():
                if citations:
                    journal_impact_scores[journal] = np.mean(citations)
            
            print(f"üìà Replacement citation stats: avg={avg_citations:.1f}, median={median_citations:.1f}")
            
            # Score and rank all candidates
            all_candidates = []
            
            # Score lab papers (highest priority)
            for chunk_text, metadata, citation_count in lab_papers:
                score = self._calculate_replacement_score(metadata, citation_count, avg_citations, journal_impact_scores, is_lab=True)
                all_candidates.append((score, chunk_text, metadata, 'lab'))
            
            # Score preferred author papers
            for chunk_text, metadata, citation_count in preferred_author_papers:
                score = self._calculate_replacement_score(metadata, citation_count, avg_citations, journal_impact_scores, is_preferred=True)
                all_candidates.append((score, chunk_text, metadata, 'preferred'))
            
            # Score other papers
            for chunk_text, metadata, citation_count in other_papers:
                score = self._calculate_replacement_score(metadata, citation_count, avg_citations, journal_impact_scores, is_lab=False)
                all_candidates.append((score, chunk_text, metadata, 'other'))
            
            # Sort by score (highest first)
            all_candidates.sort(key=lambda x: x[0], reverse=True)
            
            # Apply temporal balance (50/50 old/new)
            old_candidates = []
            new_candidates = []
            
            for score, chunk_text, metadata, paper_type in all_candidates:
                year = metadata.get('year', 'Unknown')
                try:
                    year_int = int(year) if year != 'Unknown' else 2020
                    if year_int <= 2020:  # Same temporal dividing line as main search
                        old_candidates.append((score, chunk_text, metadata, paper_type))
                    else:
                        new_candidates.append((score, chunk_text, metadata, paper_type))
                except (ValueError, TypeError):
                    # Default to old if year parsing fails
                    old_candidates.append((score, chunk_text, metadata, paper_type))
            
            # Select balanced replacements
            old_target = num_needed // 2
            new_target = num_needed - old_target
            
            # Select old papers
            for score, chunk_text, metadata, paper_type in old_candidates:
                if len(replacement_chunks) >= num_needed:
                    break
                # Return full chunk object with metadata instead of just text
                chunk_object = {"document": chunk_text, "metadata": metadata}
                replacement_chunks.append(chunk_object)
                excluded_chunks.add(hash(chunk_text.strip()))
                print(f"‚úÖ Added {paper_type} replacement (old, score: {score:.1f})")
            
            # Select new papers
            for score, chunk_text, metadata, paper_type in new_candidates:
                if len(replacement_chunks) >= num_needed:
                    break
                # Return full chunk object with metadata instead of just text
                chunk_object = {"document": chunk_text, "metadata": metadata}
                replacement_chunks.append(chunk_object)
                excluded_chunks.add(hash(chunk_text.strip()))
                print(f"‚úÖ Added {paper_type} replacement (new, score: {score:.1f})")
            
            print(f"üìä Selected {len(replacement_chunks)} replacement chunks")
            return replacement_chunks
            
        except Exception as e:
            print(f"‚ùå Error getting replacement chunks: {e}")
            return []

    def _calculate_replacement_score(self, metadata, citation_count, avg_citations, journal_impact_scores, is_lab=False, is_preferred=False):
        """
        Calculate a comprehensive score for replacement chunk prioritization.
        """
        score = 0.0
        
        # Base score
        score += 1.0
        
        # Lab paper bonus (highest priority)
        if is_lab:
            score += 10.0
        
        # Preferred author bonus
        if is_preferred:
            score += 5.0
        
        # Citation analysis
        if citation_count != 'not found':
            try:
                citations = int(citation_count)
                # Bonus for high citations
                if citations > avg_citations * 1.5:
                    score += 3.0
                elif citations > avg_citations:
                    score += 1.5
                # Penalty for very low citations
                elif citations < avg_citations * 0.3:
                    score -= 1.0
            except (ValueError, TypeError):
                pass
        
        # Journal impact factor
        journal = metadata.get('journal', 'Unknown journal')
        if journal in journal_impact_scores:
            journal_score = journal_impact_scores[journal]
            if journal_score > avg_citations * 1.2:
                score += 2.0
            elif journal_score > avg_citations:
                score += 1.0
        
        # Preprint penalty
        if self._is_preprint(metadata):
            score -= 2.0
        
        # Year bonus (prefer recent papers)
        year = metadata.get('year', 'Unknown')
        try:
            year_int = int(year) if year != 'Unknown' else 2020
            if year_int >= 2020:
                score += 1.0
            elif year_int >= 2015:
                score += 0.5
        except (ValueError, TypeError):
            pass
        
        return score

    def _is_preprint(self, metadata):
        """
        Check if a paper is a preprint based on metadata.
        """
        source = metadata.get('source', '').lower()
        source_name = metadata.get('source_name', '').lower()
        journal = metadata.get('journal', '').lower()
        
        # Check for preprint indicators
        preprint_indicators = ['biorxiv', 'medrxiv', 'arxiv', 'preprint', 'pre-print']
        
        for indicator in preprint_indicators:
            if indicator in source or indicator in source_name or indicator in journal:
                return True
                
        return False

    def retrieve_relevant_chunks_with_filtering(self, query, top_k=1500, lab_paper_ratio=None, preferred_authors=None, query_keywords=None):
        """
        Retrieve relevant chunks with real-time filtering and replacement.
        This ensures high-quality context by filtering during retrieval and replacing filtered chunks.
        """
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available.")
            return []

        print(f"üîç Retrieving chunks with real-time filtering for: {query}")
        print(f"üìä Target: {top_k} chunks with quality filtering")
        
        # Extract keywords from query for relevance scoring
        if query_keywords is None:
            query_keywords = query.lower().split()
            
        # Get initial results using sophisticated search
        initial_chunks = self.sophisticated_lab_paper_search(query, top_k * 2, lab_paper_ratio, preferred_authors)
        
        if not initial_chunks:
            print("‚ùå No initial chunks found")
            return []
            
        print(f"üìö Retrieved {len(initial_chunks)} initial chunks")
        
        # Filter chunks and track filtered ones
        high_quality_chunks = []
        filtered_count = 0
        excluded_chunks = set()
        
        for chunk in initial_chunks:
            if isinstance(chunk, dict):
                chunk_text = chunk.get("document", "")
            else:
                chunk_text = str(chunk)
                
            # Use robust hash-based exclusion ID
            chunk_id = hash(chunk_text.strip())
            
            if self._is_chunk_high_quality(chunk_text, query_keywords):
                # Preserve the full chunk object with metadata instead of just text
                high_quality_chunks.append(chunk)
            else:
                filtered_count += 1
                excluded_chunks.add(chunk_id)
                
        print(f"üìä Filtered out {filtered_count} low-quality chunks")
        
        # Replace filtered chunks with better alternatives
        if filtered_count > 0:
            print(f"üîÑ Replacing {filtered_count} filtered chunks with better alternatives...")
            replacement_chunks = self._get_replacement_chunks(
                query, 
                filtered_count, 
                excluded_chunks, 
                query_keywords,
                lab_paper_ratio=lab_paper_ratio,
                preferred_authors=preferred_authors
            )
            
            if replacement_chunks:
                high_quality_chunks.extend(replacement_chunks)
                print(f"‚úÖ Added {len(replacement_chunks)} replacement chunks")
            else:
                print(f"‚ö†Ô∏è  Could not find replacement chunks")
                
        # Limit to target number
        if len(high_quality_chunks) > top_k:
            high_quality_chunks = high_quality_chunks[:top_k]
            
        print(f"üìä Final result: {len(high_quality_chunks)} high-quality chunks")
        return high_quality_chunks

    def _apply_randomization_strategy(self, results, strategy='enhanced'):
        """
        Apply different randomization strategies to chunk selection results.
        
        Args:
            results: List of chunk results from ChromaDB
            strategy: Randomization strategy ('enhanced', 'shuffle', 'diversity', 'time_based', 'none')
        
        Returns:
            Randomized list of results
        """
        import random
        import time
        
        if not results:
            return results
            
        print(f"üé≤ Applying {strategy} randomization strategy to {len(results)} chunks...")
        
        if strategy == 'none':
            print("   No additional randomization applied")
            return results
            
        elif strategy == 'shuffle':
            # Simple random shuffle
            randomized_results = results.copy()
            random.shuffle(randomized_results)
            print("   Applied simple random shuffle")
            
        elif strategy == 'diversity':
            # Prioritize diversity by publication date and citations
            randomized_results = results.copy()
            
            # Sort by publication date (newer first) and citation count (higher first)
            def diversity_score(result):
                metadata = result.get('metadata', {})
                publication_date = metadata.get('publication_date', '')
                citation_count = metadata.get('citation_count', 'not found')
                
                # Extract year from publication date
                year = 0
                if publication_date and len(publication_date) >= 4:
                    try:
                        year = int(publication_date[:4])
                    except ValueError:
                        year = 0
                
                # Convert citation count to number
                citations = 0
                if citation_count != 'not found':
                    try:
                        citations = int(citation_count)
                    except (ValueError, TypeError):
                        citations = 0
                
                # Higher score for newer papers and more citations
                return (year * 1000) + citations
            
            # Sort by diversity score (descending)
            randomized_results.sort(key=diversity_score, reverse=True)
            
            # Then shuffle within similar score ranges
            if len(randomized_results) > 10:
                # Shuffle in groups of 10 to maintain some order while adding randomness
                for i in range(0, len(randomized_results), 10):
                    end_idx = min(i + 10, len(randomized_results))
                    group = randomized_results[i:end_idx]
                    random.shuffle(group)
                    randomized_results[i:end_idx] = group
            
            print("   Applied diversity-based randomization (newer, more cited papers first)")
            
        elif strategy == 'time_based':
            # Use current time as seed for reproducible but time-varying randomization
            current_minute = int(time.time() // 60)  # Changes every minute
            random.seed(current_minute)
            randomized_results = results.copy()
            random.shuffle(randomized_results)
            random.seed()  # Reset seed
            print(f"   Applied time-based randomization (seed: {current_minute})")
            
        elif strategy == 'enhanced':
            # Enhanced randomization combining relevance and diversity
            randomized_results = results.copy()
            
            # Tier-based approach: divide into tiers and shuffle within each
            num_tiers = min(5, len(randomized_results) // 10)  # 5 tiers or fewer if not enough results
            if num_tiers < 2:
                # If too few results, just shuffle
                random.shuffle(randomized_results)
                print("   Applied enhanced randomization (simple shuffle due to small dataset)")
            else:
                # Divide into tiers
                tier_size = len(randomized_results) // num_tiers
                for i in range(num_tiers):
                    start_idx = i * tier_size
                    end_idx = start_idx + tier_size if i < num_tiers - 1 else len(randomized_results)
                    
                    # Shuffle within each tier
                    tier = randomized_results[start_idx:end_idx]
                    random.shuffle(tier)
                    randomized_results[start_idx:end_idx] = tier
                
                # Then shuffle the entire list slightly
                random.shuffle(randomized_results)
                print(f"   Applied enhanced randomization (tier-based with {num_tiers} tiers)")
        
        else:
            print(f"   Unknown strategy '{strategy}', using default shuffle")
            randomized_results = results.copy()
            random.shuffle(randomized_results)
        
        return randomized_results

    def _get_paper_identifier(self, metadata):
        """Extract a unique identifier for a paper from its metadata."""
        # Try DOI first, then title, then source + title
        doi = metadata.get('doi', '')
        title = metadata.get('title', '')
        source = metadata.get('source', '')

        if doi and doi != 'No DOI':
            return f"doi:{doi}"
        elif title:
            # Use a normalized version of the title (lowercase, no extra spaces)
            normalized_title = ' '.join(title.lower().split())
            return f"title:{normalized_title}"
        else:
            # Fallback to source + title if available
            if source and title:
                normalized_title = ' '.join(title.lower().split())
                return f"source_title:{source}_{normalized_title}"
            else:
                # Last resort: use a hash of the metadata
                import hashlib
                metadata_str = str(sorted(metadata.items()))
                return f"hash:{hashlib.md5(metadata_str.encode()).hexdigest()[:8]}"

    def reset_used_papers(self):
        """Reset the list of used papers to allow reusing them."""
        previous_count = len(self.used_papers)
        previous_usage_count = len(self.paper_usage_count)
        self.used_papers.clear()
        self.paper_usage_count.clear()
        print(f"üîÑ Reset used papers list. Previously used {previous_count} papers.")
        print(f"üîÑ Reset paper usage counts. Previously tracked {previous_usage_count} papers.")
        print("üí° You can now reuse papers that were previously selected.")
    
    def reset_deprioritization(self):
        """Reset the deprioritization system to start fresh."""
        self.reset_used_papers()
        print("üéØ Deprioritization system reset. All papers will be treated equally.")
    
    def show_deprioritization_status(self):
        """Show the current status of the deprioritization system."""
        print(f"üìä Deprioritization System Status:")
        print(f"   Total unique papers used: {len(self.paper_usage_count)}")
        print(f"   Deprioritization factor: {self.deprioritization_factor * 100}% reduction per use")
        
        if self.paper_usage_count:
            # Show usage distribution
            usage_counts = list(self.paper_usage_count.values())
            max_usage = max(usage_counts)
            avg_usage = sum(usage_counts) / len(usage_counts)
            
            print(f"   Usage statistics:")
            print(f"     Maximum uses per paper: {max_usage}")
            print(f"     Average uses per paper: {avg_usage:.1f}")
            
            # Show papers with highest usage
            most_used = sorted(self.paper_usage_count.items(), key=lambda x: x[1], reverse=True)[:5]
            print(f"   Most frequently used papers:")
            for paper_id, count in most_used:
                if paper_id.startswith('doi:'):
                    print(f"     - DOI: {paper_id[4:]} (used {count} times)")
                elif paper_id.startswith('title:'):
                    print(f"     - Title: {paper_id[6:][:50]}... (used {count} times)")
                else:
                    print(f"     - {paper_id[:50]}... (used {count} times)")
        else:
            print("   No papers have been used yet.")
        
        print(f"   üí° Use 'reset_deprioritization' to clear all usage tracking")

    def show_used_papers_status(self):
        """Show the current status of used papers tracking with source breakdown."""
        print(f"üìä Used Papers Status:")
        print(f"   Total papers used: {len(self.used_papers)}")

        # Analyze source distribution of used papers
        source_counts = {'pubmed': 0, 'biorxiv': 0, 'medrxiv': 0, 'unknown': 0}

        for paper_id in self.used_papers:
            if paper_id.startswith('doi:'):
                # For DOI-based IDs, we can't easily determine source without metadata
                source_counts['unknown'] += 1
            elif paper_id.startswith('title:'):
                # For title-based IDs, we can't easily determine source without metadata
                source_counts['unknown'] += 1
            elif paper_id.startswith('source_title:'):
                # Extract source from source_title format
                if 'biorxiv_' in paper_id:
                    source_counts['biorxiv'] += 1
                elif 'medrxiv_' in paper_id:
                    source_counts['medrxiv'] += 1
                elif 'pubmed_' in paper_id:
                    source_counts['pubmed'] += 1
                else:
                    source_counts['unknown'] += 1
            else:
                source_counts['unknown'] += 1

        print(f"   Source breakdown:")
        print(f"     PubMed: {source_counts['pubmed']} papers")
        print(f"     BioRxiv: {source_counts['biorxiv']} papers")
        print(f"     MedRxiv: {source_counts['medrxiv']} papers")
        if source_counts['unknown'] > 0:
            print(f"     Unknown: {source_counts['unknown']} papers")

        if self.used_papers:
            # Show a sample of used papers
            sample_size = min(5, len(self.used_papers))
            sample_papers = list(self.used_papers)[:sample_size]
            print(f"   Sample of used papers:")
            for paper_id in sample_papers:
                if paper_id.startswith('doi:'):
                    print(f"     - DOI: {paper_id[4:]}")
                elif paper_id.startswith('title:'):
                    print(f"     - Title: {paper_id[6:][:50]}...")
                else:
                    print(f"     - {paper_id[:50]}...")

            if len(self.used_papers) > sample_size:
                print(f"     ... and {len(self.used_papers) - sample_size} more")
        else:
            print("   No papers have been used yet.")

        print(f"   üí° Use 'reset_papers' to clear the used papers list and allow reusing papers")

    @staticmethod
    def automated_verdict(accuracy, novelty, relevancy, accuracy_threshold=4, novelty_threshold=4, relevancy_threshold=3):
        if accuracy is not None and novelty is not None and relevancy is not None:
            return "ACCEPTED" if (accuracy >= accuracy_threshold and novelty >= novelty_threshold and relevancy >= relevancy_threshold) else "REJECTED"
        return "REJECTED"

    def iterative_hypothesis_generation(self, user_prompt, max_rounds=5, n=3):
        """Run generator-critic feedback loop for each hypothesis until accepted with iterative refinement."""
        novelty_threshold = 5  # Increased threshold for true novelty
        accuracy_threshold = 4
        relevancy_threshold = 3
        if not self.hypothesis_generator or not self.hypothesis_critic:
            print("[iterative_hypothesis_generation] ERROR: Hypothesis tools not initialized. Check if Gemini API key is available.")
            return
        # Generate initial hypotheses with fresh chunks for each
        print(f"[iterative_hypothesis_generation] Generating initial hypotheses with unique chunks...")
        hypotheses = []

        for i in range(n):
            print(f"[iterative_hypothesis_generation] Getting fresh chunks for hypothesis {i+1}/{n}...")
            # Get fresh chunks for each hypothesis to ensure maximum diversity
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                # Use dynamic chunk selection for maximum diversity
                randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                context_chunks = self.select_dynamic_chunks_for_generation(user_prompt, randomization_strategy=randomization_strategy)
                if not context_chunks:
                    print(f"[iterative_hypothesis_generation] Falling back to direct retrieval for hypothesis {i+1}...")
                    context_chunks = self.retrieve_relevant_chunks(user_prompt, top_k=1500)
            else:
                # Use direct retrieval
                context_chunks = self.retrieve_relevant_chunks(user_prompt, top_k=1500)

            if not context_chunks:
                print(f"[iterative_hypothesis_generation] ERROR: No relevant context found for hypothesis {i+1}.")
                return

            try:
                # Context chunks are already filtered during retrieval
                context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                hypothesis_list = self.hypothesis_generator.generate(context_texts, n=1, meta_hypothesis=user_prompt)
                if hypothesis_list:
                    hypotheses.append(hypothesis_list[0])
                    print(f"[iterative_hypothesis_generation] Generated hypothesis {i+1} with {len(context_chunks)} unique chunks")
                else:
                    print(f"[iterative_hypothesis_generation] ERROR: Failed to generate hypothesis {i+1}")
                    return
            except Exception as e:
                print(f"[iterative_hypothesis_generation] ERROR: Failed to generate hypothesis {i+1}: {e}")
                return
        accepted = [False] * n
        rounds = [0] * n
        critiques = [{} for _ in range(n)]
        from tqdm.auto import tqdm
        total_iterations = max_rounds * n
        current_iteration = 0
        print(f"[iterative_hypothesis_generation] Starting iterative refinement (max {max_rounds} rounds per hypothesis, 5-minute time limit per critique, acceptance: Novelty >= {novelty_threshold}, Accuracy >= {accuracy_threshold}, Relevancy >= {relevancy_threshold})...")
        with tqdm(total=total_iterations, desc="Hypothesis refinement", unit="iteration") as pbar:
            while not all(accepted) and max(rounds) < max_rounds:
                for i in range(n):
                    if accepted[i]:
                        continue
                    rounds[i] += 1
                    current_iteration += 1
                    print(f"[iterative_hypothesis_generation] Critique Round {rounds[i]} for Hypothesis {i+1} (5-minute timer started)...")
                    self.hypothesis_timer.start()
                    try:
                        # Context chunks are already filtered during retrieval
                        context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                        if self.hypothesis_timer.check_expired():
                            print(f"[iterative_hypothesis_generation] TIMEOUT: Hypothesis {i+1} critique timed out.")
                            citations = self.extract_citations_from_chunks(context_chunks)
                            formatted_citations = self.format_citations_for_export(citations)
                            self.hypothesis_records.append({
                                'Hypothesis': hypotheses[i],
                                'Accuracy': None,
                                'Novelty': None,
                                'Verdict': 'TIMEOUT',
                                'Critique': 'Critique process timed out after 5 minutes',
                                'Citations': formatted_citations
                            })
                            pbar.update(1)
                            continue
                        critique_result = self.hypothesis_critic.critique(hypotheses[i], context_texts, prompt=user_prompt, lab_goals=get_lab_goals(), meta_hypothesis=user_prompt)
                        if self.hypothesis_timer.check_expired():
                            print(f"[iterative_hypothesis_generation] TIMEOUT: Hypothesis {i+1} critique timed out after critique.")
                            citations = self.extract_citations_from_chunks(context_chunks)
                            formatted_citations = self.format_citations_for_export(citations)
                            self.hypothesis_records.append({
                                'Hypothesis': hypotheses[i],
                                'Accuracy': critique_result.get('accuracy', None),
                                'Novelty': critique_result.get('novelty', None),
                                'Verdict': 'TIMEOUT',
                                'Critique': critique_result.get('critique', '') + ' [Process timed out]',
                                'Citations': formatted_citations
                            })
                            pbar.update(1)
                            continue
                        critiques[i] = critique_result
                        novelty = critique_result.get('novelty', 0)
                        accuracy = critique_result.get('accuracy', 0)
                        relevancy = critique_result.get('relevancy', 0)
                        verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold, novelty_threshold, relevancy_threshold)
                        remaining_time = self.hypothesis_timer.get_remaining_time()
                        print(f"\033[1mHypothesis {i+1}:\033[0m \033[96m{hypotheses[i]}\033[0m")
                        print(f"[iterative_hypothesis_generation] Critique: {critique_result['critique']}")
                        print(f"[iterative_hypothesis_generation] Scores: Novelty={novelty}/{novelty_threshold}, Accuracy={accuracy}/{accuracy_threshold}, Relevancy={relevancy}/{relevancy_threshold}, Automated Verdict={verdict}, Time left={remaining_time:.1f}s")
                        citations = self.extract_citations_from_chunks(context_chunks)
                        formatted_citations = self.format_citations_for_export(citations)
                        self.hypothesis_records.append({
                            'Hypothesis': hypotheses[i],
                            'Accuracy': accuracy,
                            'Novelty': novelty,
                            'Relevancy': relevancy,
                            'Verdict': verdict,
                            'Critique': critique_result.get('critique', ''),
                            'Citations': formatted_citations
                        })
                        if verdict == 'ACCEPTED':
                            print(f"\033[92mACCEPTED\033[0m: Hypothesis {i+1} (Novelty: {novelty}/{novelty_threshold}, Accuracy: {accuracy}/{accuracy_threshold}, Relevancy: {relevancy}/{relevancy_threshold})")
                            accepted[i] = True
                        else:
                            print(f"\033[91mREJECTED\033[0m: Hypothesis {i+1} (Novelty: {novelty}/{novelty_threshold}, Accuracy: {accuracy}/{accuracy_threshold}, Relevancy: {relevancy}/{relevancy_threshold}) - Refining...")
                            if self.hypothesis_timer.check_expired():
                                print(f"[iterative_hypothesis_generation] TIMEOUT: Hypothesis {i+1} refinement timed out.")
                                pbar.update(1)
                                continue
                            try:
                                # Context chunks are already filtered during retrieval
                                context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                                # Try to refine the hypothesis based on critique feedback
                                refined_hypothesis = self.hypothesis_generator.refine_hypothesis(hypotheses[i], critique_result, context_texts, user_prompt)
                                if refined_hypothesis and refined_hypothesis != hypotheses[i]:
                                    hypotheses[i] = refined_hypothesis
                                    print(f"[iterative_hypothesis_generation] ‚ú® Hypothesis {i+1} refined based on critique feedback")
                                else:
                                    # Fallback to regeneration if refinement fails
                                    print(f"[iterative_hypothesis_generation] ‚ö†Ô∏è  Refinement failed, regenerating hypothesis {i+1}...")
                                    new_hypothesis = self.hypothesis_generator.generate(context_texts, n=1, meta_hypothesis=user_prompt)
                                    if new_hypothesis:
                                        hypotheses[i] = new_hypothesis[0]
                            except Exception as e:
                                print(f"[iterative_hypothesis_generation] ERROR: Failed to refine/regenerate hypothesis {i+1}: {e}")
                                continue
                        avg_rating = (novelty + accuracy + relevancy) / 3 if novelty is not None and accuracy is not None and relevancy is not None else 0
                        pbar.set_postfix({
                            "accepted": sum(accepted),
                            "round": max(rounds),
                            "current": f"H{i+1}",
                            "avg_rating": f"{avg_rating:.1f}",
                            "time_left": f"{remaining_time:.0f}s"
                        })
                        pbar.update(1)
                    except Exception as e:
                        print(f"[iterative_hypothesis_generation] ERROR: Failed to critique hypothesis {i+1}: {e}")
                        citations = self.extract_citations_from_chunks(context_chunks)
                        formatted_citations = self.format_citations_for_export(citations)
                        self.hypothesis_records.append({
                            'Hypothesis': hypotheses[i],
                            'Accuracy': None,
                            'Novelty': None,
                            'Relevancy': None,
                            'Verdict': 'ERROR',
                            'Critique': f'Error during critique: {str(e)}',
                            'Citations': formatted_citations
                        })
                        pbar.update(1)
                        continue
                    if all(accepted):
                        break
        print("[iterative_hypothesis_generation] Final Hypotheses:")
        for i, hyp in enumerate(hypotheses, 1):
            status = "ACCEPTED" if accepted[i-1] else f"FAILED (max {max_rounds} rounds)"
            print(f"   {i}. {hyp} [{status}]")
            if isinstance(critiques[i-1], dict) and critiques[i-1]:
                print(f"      Critique: {critiques[i-1].get('critique','')}")
                print(f"      Novelty: {critiques[i-1].get('novelty','')}/{novelty_threshold}, Accuracy: {critiques[i-1].get('accuracy','')}/{accuracy_threshold}, Relevancy: {critiques[i-1].get('relevancy','')}/{relevancy_threshold}, Automated Verdict: {self.automated_verdict(critiques[i-1].get('accuracy',0), critiques[i-1].get('novelty',0), critiques[i-1].get('relevancy',0), accuracy_threshold, novelty_threshold, relevancy_threshold)}")
        return hypotheses

    def add_to_package(self, search_results):
        """Add search results to the current package."""
        current_size = self.current_package["total_chars"]
        new_size = current_size + sum(len(result.get("chunk", "")) for result in search_results)
        if new_size > 2000000:
            print(f"[add_to_package] WARNING: Adding {len(search_results)} chunks would make package very large ({new_size:,} characters)")
            print("[add_to_package] This may cause API errors. Consider using 'clear' first.")
            response = input("Continue anyway? (y/n): ").lower()
            if response != 'y':
                return

        # Store the initial add quantity for dynamic chunk selection
        if self.initial_add_quantity is None:
            self.initial_add_quantity = len(search_results)
            print(f"[add_to_package] Stored initial add quantity: {self.initial_add_quantity} chunks for dynamic selection")

        for result in search_results:
            chunk = result.get("chunk", "")
            metadata = result.get("metadata", {})
            self.current_package["chunks"].append(chunk)
            self.current_package["metadata"].append(metadata)
            self.current_package["sources"].add(metadata.get("source", "Unknown"))
            self.current_package["total_chars"] += len(chunk)
        print(f"[add_to_package] Added {len(search_results)} chunks to package.")
        print(f"[add_to_package] Package now contains {len(self.current_package['chunks'])} chunks from {len(self.current_package['sources'])} sources.")
        if self.current_package["total_chars"] > 1000000:
            print(f"[add_to_package] WARNING: Package size: {self.current_package['total_chars']:,} characters (consider using 'clear' if you get API errors)")

    def clear_package(self):
        """Clear the current package."""
        self.current_package = {
            "chunks": [],
            "metadata": [],
            "sources": set(),
            "total_chars": 0
        }
        # Reset the initial add quantity when clearing package
        self.initial_add_quantity = None
        # Reset used papers tracking when clearing package
        previous_papers_count = len(self.used_papers)
        self.used_papers.clear()
        print("[clear_package] Package cleared.")
        print("[clear_package] Initial add quantity reset.")
        print(f"[clear_package] Used papers list reset ({previous_papers_count} papers cleared).")

    def show_package(self):
        """Display information about the current package with source breakdown."""
        if not self.current_package["chunks"]:
            print("[show_package] Package is empty.")
            return

        # Calculate source distribution
        source_counts = {'pubmed': 0, 'biorxiv': 0, 'medrxiv': 0, 'unknown': 0}

        for metadata in self.current_package["metadata"]:
            source = metadata.get('source', '').lower()
            source_name = metadata.get('source_name', '').lower()

            if 'pubmed' in source or 'pubmed' in source_name:
                source_counts['pubmed'] += 1
            elif 'biorxiv' in source or 'biorxiv' in source_name:
                source_counts['biorxiv'] += 1
            elif 'medrxiv' in source or 'medrxiv' in source_name:
                source_counts['medrxiv'] += 1
            else:
                source_counts['unknown'] += 1

        print(f"[show_package] Current Package:")
        print(f"   üìö Total chunks: {len(self.current_package['chunks'])}")
        print(f"   üìù Total characters: {self.current_package['total_chars']:,}")
        print(f"   üìñ Sources: {', '.join(self.current_package['sources'])}")
        print(f"   üìä Used papers tracked: {len(self.used_papers)}")

        print(f"   üìä Source distribution:")
        print(f"     PubMed: {source_counts['pubmed']} chunks")
        print(f"     BioRxiv: {source_counts['biorxiv']} chunks")
        print(f"     MedRxiv: {source_counts['medrxiv']} chunks")
        if source_counts['unknown'] > 0:
            print(f"     Unknown: {source_counts['unknown']} chunks")

        print(f"\n[show_package] Sample chunks:")
        for i, chunk in enumerate(self.current_package["chunks"][:3], 1):
            metadata = self.current_package["metadata"][i-1]
            title = metadata.get("title", "No title")[:60]
            source = metadata.get("source", "Unknown")
            print(f"   {i}. {title}... (Source: {source})")
        if len(self.current_package["chunks"]) > 3:
            print(f"   ... and {len(self.current_package['chunks']) - 3} more chunks")

    def generate_hypotheses_from_package(self, n=5):
        novelty_threshold = 4
        accuracy_threshold = 4
        relevancy_threshold = 3
        if not self.current_package["chunks"]:
            print("‚ùå Package is empty. Add some chunks first.")
            return
        if not self.hypothesis_generator or not self.hypothesis_critic:
            print("‚ùå Hypothesis tools not initialized. Check if Gemini API key is available.")
            return
        # Prompt for the query if not already stored
        if "prompt" not in self.current_package or not self.current_package["prompt"]:
            self.current_package["prompt"] = input("Enter the original query for this package: ").strip()
        package_prompt = self.current_package["prompt"]
        print(f"\nüß† Generating {n} hypotheses...")
        print(f"üì¶ Using {len(self.current_package['chunks'])} package chunks for critique")
        print(f"[INFO] Acceptance criteria: Novelty >= {novelty_threshold}, Accuracy >= {accuracy_threshold}, Relevancy >= {relevancy_threshold}")

        # Check package size and warn if too large
        package_size = sum(len(chunk) for chunk in self.current_package["chunks"])
        if package_size > 1000000:  # 1MB limit for package
            print(f"‚ö†Ô∏è  Package is very large ({package_size:,} characters). This may cause API errors.")
            print("üí° Consider using 'clear' and adding fewer chunks.")
            response = input("Continue anyway? (y/n): ").lower()
            if response != 'y':
                return

        # For generation, use a sample of the database instead of the entire thing
        # This prevents the 300MB payload limit error
        if not self.embeddings_data:
            # Check if ChromaDB has data first
            if self.use_chromadb and self.chroma_manager and self.is_chromadb_ready():
                print("üìö ChromaDB has data - using package chunks for generation (no need to load embeddings)")
                # Use only the package chunks for generation when ChromaDB is available
                all_chunks = self.current_package["chunks"]
            else:
                print("üìö Loading embeddings data...")
                self.embeddings_data = self.load_all_embeddings()
                if not self.embeddings_data:
                    print("‚ùå No embeddings data available")
                    return
                all_chunks = self.embeddings_data["chunks"]
        else:
            all_chunks = self.embeddings_data["chunks"]
        package_chunks = self.current_package["chunks"]
        package_metadata = self.current_package["metadata"]

        # Calculate safe sample size (aim for ~50MB payload)
        # Each chunk is roughly 1000 characters, so 50MB = ~50,000 chunks
        max_chunks_for_generation = 50000

        # Check if we're using package chunks (from ChromaDB) or all chunks (from embeddings)
        if all_chunks is self.current_package["chunks"]:
            # Using package chunks from ChromaDB
            generation_chunks = all_chunks
            print(f"üìö Using {len(generation_chunks)} package chunks from ChromaDB for generation")
        else:
            # Using all chunks from embeddings data
            if len(all_chunks) > max_chunks_for_generation:
                print(f"üìö Database contains {len(all_chunks)} total chunks")
                print(f"üìö Using sample of {max_chunks_for_generation} chunks for generation (to avoid API limits)")
                # Take a representative sample from different parts of the database
                step = len(all_chunks) // max_chunks_for_generation
                generation_chunks = all_chunks[::step][:max_chunks_for_generation]
            else:
                generation_chunks = all_chunks
                print(f"üìö Database contains {len(all_chunks)} total chunks")

        print(f"üì¶ Package contains {len(package_chunks)} chunks")

        # Sequential, real-time generator-critic loop for 5 accepted hypotheses
        print(f"\nüß† Generating hypotheses one at a time until 5 are accepted (novelty >= {novelty_threshold}, accuracy >= {accuracy_threshold}, 5-minute time limit per critique)...")

        # Initialize Excel file for incremental saving
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Ensure hypothesis_export directory exists
        export_dir = "hypothesis_export"
        os.makedirs(export_dir, exist_ok=True)
        excel_filename = os.path.join(export_dir, f"hypothesis_export_{timestamp}.xlsx")
        print(f"üíæ Will save hypotheses incrementally to: {excel_filename}")

        accepted_hypotheses = []
        attempts = 0
        max_attempts = 100  # Prevent infinite loops
        while len(accepted_hypotheses) < n and attempts < max_attempts:
            attempts += 1
            print(f"\nüß† Generating hypothesis attempt {attempts}...")

            # ALWAYS select new chunks for EVERY hypothesis attempt to ensure diversity
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                print(f"üîÑ Selecting new chunks for hypothesis attempt {attempts}...")
                randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                dynamic_chunks = self.select_dynamic_chunks_for_generation(package_prompt, randomization_strategy=randomization_strategy)
                if dynamic_chunks:
                    generation_chunks = dynamic_chunks
                    print(f"üìö Using {len(generation_chunks)} dynamically selected chunks for generation")
                else:
                    print(f"‚ö†Ô∏è  Failed to select dynamic chunks, using original chunks")
            else:
                print(f"üìö Using original chunks for generation (no dynamic selection available)")

            # Start timer for this hypothesis
            self.hypothesis_timer.start()
            print(f"‚è±Ô∏è  Starting 5-minute timer for hypothesis attempt {attempts}...")

            hypothesis_list = self._generate_hypotheses_with_retry(generation_chunks, n=1)
            if not hypothesis_list:
                print("‚ùå Failed to generate hypothesis. Skipping...")
                # Record failed generation
                failed_record = {
                    'Hypothesis': f'Failed generation attempt {attempts}',
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'GENERATION_FAILED',
                    'Critique': 'Failed to generate hypothesis',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(failed_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(failed_record, excel_filename)
                continue

            hypothesis = hypothesis_list[0]
            
            # Validate hypothesis format before proceeding
            from src.ai.hypothesis_tools import validate_hypothesis_format
            is_valid_format, format_reason = validate_hypothesis_format(hypothesis)
            if not is_valid_format:
                print(f"‚ùå Hypothesis rejected due to format: {format_reason}")
                print(f"‚è≠Ô∏è  Skipping critique for format-rejected hypothesis")
                # Record format rejection without critique
                format_rejected_record = {
                    'Hypothesis': hypothesis,
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'FORMAT_REJECTED',
                    'Critique': f'Hypothesis rejected: {format_reason}',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(format_rejected_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(format_rejected_record, excel_filename)
                continue

            hypothesis = hypothesis_list[0]
            print(f"Generated Hypothesis: {hypothesis}")
            print(f"\nüîÑ Critiquing Hypothesis: {hypothesis}")

            # Check timer before critique
            if self.hypothesis_timer.check_expired():
                print(f"‚è∞ Time limit reached for hypothesis attempt {attempts}. Moving to next attempt.")
                # Record timeout result
                timeout_record = {
                    'Hypothesis': hypothesis,
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'TIMEOUT',
                    'Critique': 'Critique process timed out after 5 minutes',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(timeout_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(timeout_record, excel_filename)
                continue

            critique_result = self._critique_hypothesis_with_retry(hypothesis, package_chunks, package_prompt)
            if not critique_result:
                print(f"‚ùå Failed to critique hypothesis after retries. Skipping...")
                # Record failed critique
                critique_failed_record = {
                    'Hypothesis': hypothesis,
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'CRITIQUE_FAILED',
                    'Critique': 'Failed to critique hypothesis after retries',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(critique_failed_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(critique_failed_record, excel_filename)
                continue

            # Check timer after critique
            if self.hypothesis_timer.check_expired():
                print(f"‚è∞ Time limit reached for hypothesis attempt {attempts} after critique. Moving to next attempt.")
                # Record timeout result with partial critique
                partial_timeout_record = {
                    'Hypothesis': hypothesis,
                    'Accuracy': critique_result.get('accuracy', None),
                    'Novelty': critique_result.get('novelty', None),
                    'Verdict': 'TIMEOUT',
                    'Critique': critique_result.get('critique', '') + ' [Process timed out]',
                    'Citations': 'No citations available'
                }
                self.hypothesis_records.append(partial_timeout_record)
                # Save incrementally
                self.save_hypothesis_record_incrementally(partial_timeout_record, excel_filename)
                continue

            novelty = critique_result.get('novelty', 0)
            accuracy = critique_result.get('accuracy', 0)
            relevancy = critique_result.get('relevancy', 0)
            
            # Handle None values by converting to 0
            if novelty is None:
                novelty = 0
            if accuracy is None:
                accuracy = 0
            if relevancy is None:
                relevancy = 0
                
            # Use the same thresholds as the acceptance logic
            verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold, novelty_threshold, relevancy_threshold)
            remaining_time = self.hypothesis_timer.get_remaining_time()
            print(f"   Novelty: {novelty}/{novelty_threshold}, Accuracy: {accuracy}/{accuracy_threshold}, Relevancy: {relevancy}/{relevancy_threshold}, Automated Verdict: {verdict}")
            print(f"   ‚è±Ô∏è  Time remaining: {remaining_time:.1f}s")
            
            # Print rejection reason if hypothesis is rejected
            if verdict != 'ACCEPTED':
                if accuracy < accuracy_threshold:
                    print(f"‚ùå Hypothesis rejected: Low accuracy score ({accuracy}/{accuracy_threshold})")
                elif novelty < novelty_threshold:
                    print(f"‚ùå Hypothesis rejected: Low novelty score ({novelty}/{novelty_threshold})")
                elif relevancy < relevancy_threshold:
                    print(f"‚ùå Hypothesis rejected: Low relevancy score ({relevancy}/{relevancy_threshold})")
                else:
                    print(f"‚ùå Hypothesis rejected: Unknown reason")

            # Extract citations from the dynamically selected chunks used for this specific hypothesis
            citations = []
            unique_sources = set()
            
            # Get metadata from the dynamically selected chunks
            dynamic_metadata = []
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                # If we used dynamic chunk selection, get metadata from those chunks
                if hasattr(self, 'last_dynamic_results') and self.last_dynamic_results:
                    dynamic_metadata = [result.get('metadata', {}) for result in self.last_dynamic_results]
                else:
                    # Fallback: extract from generation_chunks if they have metadata
                    dynamic_metadata = []
                    for chunk in generation_chunks:
                        if isinstance(chunk, dict) and 'metadata' in chunk:
                            dynamic_metadata.append(chunk['metadata'])
                        else:
                            # If chunks don't have metadata, use package metadata as fallback
                            dynamic_metadata = package_metadata
                            break
            else:
                # If not using dynamic selection, use package metadata
                dynamic_metadata = package_metadata
            
            # Extract citations from the metadata
            for metadata in dynamic_metadata:
                if metadata:
                    source_name = metadata.get('source_name', metadata.get('source', 'Unknown'))
                    title = metadata.get('title', 'No title')
                    doi = metadata.get('doi', 'No DOI')
                    
                    # Try multiple field names for authors
                    authors = (metadata.get('author') or 
                             metadata.get('authors') or 
                             metadata.get('author_list') or 
                             'Unknown authors')
                    
                    # Try multiple field names for journal
                    journal = (metadata.get('journal') or 
                              metadata.get('journal_name') or 
                              metadata.get('publication_venue') or 
                              'Unknown journal')
                    
                    # Extract year from publication_date or other date fields
                    publication_date = (metadata.get('publication_date') or 
                                       metadata.get('date') or 
                                       metadata.get('year') or 
                                       '')
                    
                    if publication_date and len(str(publication_date)) >= 4:
                        year = str(publication_date)[:4]
                    else:
                        year = 'Unknown year'

                    citation = {
                        'source_name': source_name,
                        'title': title,
                        'doi': doi,
                        'authors': authors,
                        'journal': journal,
                        'year': year
                    }

                    # Add to citations if not already present
                    citation_key = doi if doi != 'No DOI' else title
                    if citation_key not in unique_sources:
                        citations.append(citation)
                        unique_sources.add(citation_key)

            formatted_citations = self.format_citations_for_export(citations)

            # Track record for export
            hypothesis_record = {
                'Hypothesis': hypothesis,
                'Accuracy': accuracy,
                'Novelty': novelty,
                'Relevancy': relevancy,
                'Verdict': verdict,
                'Critique': critique_result.get('critique', ''),
                'Citations': formatted_citations
            }
            self.hypothesis_records.append(hypothesis_record)

            # Save incrementally to Excel
            self.save_hypothesis_record_incrementally(hypothesis_record, excel_filename)

            if verdict == 'ACCEPTED':
                accepted_hypotheses.append({
                    "hypothesis": hypothesis,
                    "critique": critique_result,
                    "score": (novelty + accuracy + relevancy) / 3
                })
                print(f"‚úÖ Hypothesis accepted! ({len(accepted_hypotheses)}/{n})-----------------------------------\n")
            else:
                print(f"‚ùå Hypothesis rejected (verdict: {verdict})-----------------------------------\n")

        if not accepted_hypotheses:
            print("‚ùå No hypotheses were successfully critiqued.")
            return

        print(f"\nüèÜ Top {len(accepted_hypotheses)} Hypotheses:")
        print("=" * 80)
        for i, result in enumerate(accepted_hypotheses, 1):
            hypothesis = result["hypothesis"]
            score = result["score"]
            critique = result["critique"]
            print(f"\n{i}. {hypothesis}")
            print(f"   Score: {score:.1f}")
            print(f"   Novelty: {critique.get('novelty', 'N/A')}")
            print(f"   Accuracy: {critique.get('accuracy', 'N/A')}")
            print(f"   Relevancy: {critique.get('relevancy', 'N/A')}")
            print(f"   Automated Verdict: {self.automated_verdict(critique.get('accuracy', 0), critique.get('novelty', 0), critique.get('relevancy', 0), accuracy_threshold, novelty_threshold, relevancy_threshold)}")
            print(f"   Critique: {critique.get('critique', 'N/A')}")
            print("-" * 80)
        return accepted_hypotheses[:n]

    def _generate_hypotheses_with_retry(self, generation_chunks, n, max_retries=3):
        """Generate hypotheses with retry logic for rate limiting."""
        if not self.hypothesis_generator:
            print("‚ùå Hypothesis generator not available")
            return None

        for attempt in range(max_retries):
            try:
                # Check rate limit before making API call
                remaining_requests = self.gemini_rate_limiter.get_remaining_requests()
                remaining_tokens = self.gemini_rate_limiter.get_remaining_tokens()
                
                # Estimate tokens for hypothesis generation (prompt + context + response)
                # Use only a small sample of context for token estimation to avoid false quota limits
                context_sample = " ".join(generation_chunks[:1])[:2000]  # Limit to first chunk, max 2000 chars
                estimated_tokens = self.gemini_rate_limiter.estimate_tokens(context_sample) + 1500  # Buffer for response
                
                if remaining_requests <= 0 or remaining_tokens < estimated_tokens:
                    print(f"‚è≥ Gemini API quota limit reached. Requests: {remaining_requests}, Tokens: {remaining_tokens:,}/{estimated_tokens:,}")
                    self.gemini_rate_limiter.wait_if_needed(estimated_tokens)

                # Extract chunk text from the sample
                generation_chunk_texts = [chunk for chunk in generation_chunks]
                hypotheses = self.hypothesis_generator.generate(generation_chunk_texts, n=n)
                return hypotheses
            except Exception as e:
                error_str = str(e)
                if "429" in error_str or "RESOURCE_EXHAUSTED" in error_str or "quota" in error_str.lower():
                    # Extract retry delay from error message if available
                    import re
                    retry_match = re.search(r'Please retry in (\d+\.?\d*)s', error_str)
                    if retry_match:
                        wait_time = float(retry_match.group(1)) + 2  # Add 2s buffer
                    else:
                        wait_time = (2 ** attempt) * 30 + random.randint(0, 10)  # Exponential backoff
                    
                    print(f"‚ö†Ô∏è  API quota exceeded (attempt {attempt + 1}/{max_retries}). Waiting {wait_time:.1f}s...")
                    print(f"   Error: {error_str[:200]}...")
                    time.sleep(wait_time)
                    if attempt == max_retries - 1:
                        print("‚ùå Max retries reached. API quota exceeded.")
                        return None
                elif "payload size exceeds the limit" in error_str:
                    print("üí° Payload too large. Try reducing package size with 'clear' and adding fewer chunks")
                    return None
                else:
                    print(f"‚ùå Unexpected error: {e}")
                    return None
        return None

    def _critique_hypothesis_with_retry(self, hypothesis, package_chunks, prompt, max_retries=3):
        """Critique hypothesis with retry logic for rate limiting."""
        if not self.hypothesis_critic:
            print("‚ùå Hypothesis critic not available")
            return None

        for attempt in range(max_retries):
            try:
                # Check rate limit before making API call
                remaining_requests = self.gemini_rate_limiter.get_remaining_requests()
                remaining_tokens = self.gemini_rate_limiter.get_remaining_tokens()
                
                # Estimate tokens for critique generation (hypothesis + context + prompt + response)
                # Use only a small sample of context for token estimation to avoid false quota limits
                context_sample = " ".join(package_chunk_texts[:1])[:2000]  # Limit to first chunk, max 2000 chars
                estimated_tokens = self.gemini_rate_limiter.estimate_tokens(hypothesis + context_sample + prompt) + 2000  # Buffer for response
                
                if remaining_requests <= 0 or remaining_tokens < estimated_tokens:
                    print(f"‚è≥ Gemini API quota limit reached. Requests: {remaining_requests}, Tokens: {remaining_tokens:,}/{estimated_tokens:,}")
                    self.gemini_rate_limiter.wait_if_needed(estimated_tokens)

                # Extract chunk text from package chunks
                package_chunk_texts = [chunk for chunk in package_chunks]
                critique_result = self.hypothesis_critic.critique(hypothesis, package_chunk_texts, prompt=prompt, lab_goals=get_lab_goals(), meta_hypothesis=prompt)
                return critique_result
            except Exception as e:
                error_str = str(e)
                if "429" in error_str or "RESOURCE_EXHAUSTED" in error_str or "quota" in error_str.lower():
                    # Extract retry delay from error message if available
                    import re
                    retry_match = re.search(r'Please retry in (\d+\.?\d*)s', error_str)
                    if retry_match:
                        wait_time = float(retry_match.group(1)) + 2  # Add 2s buffer
                    else:
                        wait_time = (2 ** attempt) * 30 + random.randint(0, 10)  # Exponential backoff
                    
                    print(f"‚ö†Ô∏è  API quota exceeded (attempt {attempt + 1}/{max_retries}). Waiting {wait_time:.1f}s...")
                    print(f"   Error: {error_str[:200]}...")
                    time.sleep(wait_time)
                    if attempt == max_retries - 1:
                        print("‚ùå Max retries reached. API quota exceeded.")
                        return None
                elif "payload size exceeds the limit" in error_str:
                    print("üí° Package is too large for critique. Try 'clear' and add fewer chunks.")
                    return None
                else:
                    print(f"‚ùå Unexpected error: {e}")
                    return None
        return None

    def generate_hypotheses_offline(self, n=5):
        """Generate hypotheses without using the API (fallback mode)."""
        if not self.current_package["chunks"]:
            print("‚ùå Package is empty. Add some chunks first.")
            return

        print(f"\nüß† Generating {n} hypotheses (offline mode)...")
        print(f"üì¶ Using {len(self.current_package['chunks'])} package chunks")

        # Extract key terms and patterns from the package
        package_text = " ".join(self.current_package["chunks"])

        # Simple keyword extraction
        import re
        from collections import Counter

        # Extract UBR-5 related terms
        ubr5_terms = re.findall(r'\b(?:UBR5?|ubiquitin|ligase|protein|E3|ubiquitination|degradation|regulation|pathway|signaling|cancer|immunology|immune|response|activation|inhibition|expression|function|mechanism|target|therapeutic|drug|treatment|therapy)\b', package_text, re.IGNORECASE)
        term_counts = Counter(ubr5_terms)

        # Generate simple hypotheses based on common patterns
        hypotheses = []

        # Hypothesis 1: UBR5 regulation
        if any(term in term_counts for term in ['regulation', 'expression', 'function']):
            hypotheses.append("UBR5 expression and function are regulated by specific signaling pathways in immune cells, affecting immune response and cancer progression.")

        # Hypothesis 2: Therapeutic targeting
        if any(term in term_counts for term in ['therapeutic', 'drug', 'treatment', 'target']):
            hypotheses.append("UBR5 represents a novel therapeutic target for cancer immunotherapy, with potential for drug development and clinical applications.")

        # Hypothesis 3: Immune response
        if any(term in term_counts for term in ['immune', 'response', 'activation', 'immunology']):
            hypotheses.append("UBR5 plays a critical role in regulating immune cell activation and response, influencing cancer immunosurveillance and immunotherapy efficacy.")

        # Hypothesis 4: Ubiquitination pathway
        if any(term in term_counts for term in ['ubiquitin', 'ligase', 'ubiquitination', 'degradation']):
            hypotheses.append("UBR5-mediated ubiquitination regulates key proteins in immune signaling pathways, affecting cell fate decisions and immune function.")

        # Hypothesis 5: Cancer mechanism
        if any(term in term_counts for term in ['cancer', 'mechanism', 'pathway', 'signaling']):
            hypotheses.append("UBR5 functions as a key regulator in cancer cell signaling pathways, with implications for tumor progression and therapeutic resistance.")

        # Fill remaining slots with generic hypotheses if needed
        while len(hypotheses) < n:
            hypotheses.append(f"UBR5 may have additional roles in cellular processes related to {list(term_counts.keys())[:3] if term_counts else 'protein regulation'}.")

        print(f"\nüèÜ Generated {len(hypotheses)} Hypotheses (Offline Mode):")
        print("=" * 80)
        for i, hypothesis in enumerate(hypotheses[:n], 1):
            print(f"\n{i}. {hypothesis}")
            print(f"   Score: N/A (offline mode)")
            print(f"   Novelty: N/A (offline mode)")
            print(f"   Accuracy: N/A (offline mode)")
            print(f"   Verdict: N/A (offline mode)")
            print(f"   Note: This is a fallback hypothesis based on keyword analysis")
            print("-" * 80)

        print(f"\nüí° Note: These are simplified hypotheses generated without API access.")
        print(f"üí° For more sophisticated analysis, try 'generate' when API is available.")

        return hypotheses[:n]

    def generate_hypotheses_with_meta_generator(self, user_prompt: str, n_per_meta: int = 5, chunks_per_meta: int = 1500, progress_callback=None):
        """
        Generate hypotheses using the meta-hypothesis generator approach.

        This method:
        1. Takes the user's prompt and generates 5 meta-hypotheses
        2. For each meta-hypothesis, generates hypotheses until exactly 5 are accepted
        3. Only proceeds to the next meta-hypothesis after 5 hypotheses are accepted for the current one
        4. Returns all generated hypotheses with their critiques

        Args:
            user_prompt: The original user query
            n_per_meta: Number of accepted hypotheses required per meta-hypothesis (default: 5)
            chunks_per_meta: Number of context chunks to use per meta-hypothesis (default: 1500)
        """
        if not self.meta_hypothesis_generator or not self.hypothesis_generator or not self.hypothesis_critic:
            print("‚ùå Meta-hypothesis tools not initialized. Check if Gemini API key is available.")
            return None

        print(f"\nüß† META-HYPOTHESIS GENERATION SESSION")
        print("=" * 80)
        print(f"Original Query: {user_prompt}")
        print(f"Requiring {n_per_meta} ACCEPTED hypotheses per meta-hypothesis")
        print("=" * 80)

        # Step 1: Generate meta-hypotheses
        print(f"\nüîç Step 1: Generating 5 meta-hypotheses from user query...")
        if progress_callback:
            progress_callback("Generating 5 meta-hypotheses from user query...", 50)
        try:
            meta_hypotheses = self.meta_hypothesis_generator.generate_meta_hypotheses(user_prompt)
            if not meta_hypotheses or len(meta_hypotheses) < 5:
                print("‚ùå Failed to generate sufficient meta-hypotheses")
                if progress_callback:
                    progress_callback("Failed to generate sufficient meta-hypotheses", 0)
                return None

            print(f"‚úÖ Generated {len(meta_hypotheses)} meta-hypotheses:")
            for i, meta_hyp in enumerate(meta_hypotheses, 1):
                print(f"  {i}. {meta_hyp}")
        except Exception as e:
            print(f"‚ùå Error generating meta-hypotheses: {e}")
            if progress_callback:
                progress_callback("Error generating meta-hypotheses", 0)
            return None

        # Step 2: For each meta-hypothesis, generate hypotheses until exactly n_per_meta are accepted
        all_hypotheses = []
        total_meta_hypotheses = len(meta_hypotheses)

        # Initialize Excel file for incremental saving
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Ensure hypothesis_export directory exists
        export_dir = "hypothesis_export"
        os.makedirs(export_dir, exist_ok=True)
        excel_filename = os.path.join(export_dir, f"meta_hypothesis_export_{timestamp}.xlsx")
        print(f"üíæ Will save meta-hypotheses incrementally to: {excel_filename}")

        # Initialize hypothesis records for this session
        self.hypothesis_records = []
        
        # Reset deprioritization system for fresh citation diversity
        self.reset_deprioritization()
        print(f"üéØ Reset deprioritization system for maximum citation diversity across meta-hypotheses")

        for meta_idx, meta_hypothesis in enumerate(meta_hypotheses, 1):
            print(f"\nüîç Step 2.{meta_idx}: Generating hypotheses for meta-hypothesis {meta_idx}/{total_meta_hypotheses}")
            print(f"Meta-hypothesis: {meta_hypothesis}")
            print("-" * 60)
            
            # Update progress for meta-hypothesis processing
            if progress_callback:
                progress_percentage = 50 + (meta_idx / total_meta_hypotheses) * 40  # 50-90% range
                progress_callback(f"Processing meta-hypothesis {meta_idx}/{total_meta_hypotheses}: {meta_hypothesis[:50]}...", progress_percentage)

            # Generate hypotheses for this meta-hypothesis until exactly n_per_meta are accepted
            meta_hypotheses_generated = []
            attempts = 0
            max_attempts = n_per_meta * 10  # Allow more attempts to get enough accepted hypotheses

            while len(meta_hypotheses_generated) < n_per_meta and attempts < max_attempts:
                attempts += 1
                print(f"\nüß† Generating hypothesis attempt {attempts} for meta-hypothesis {meta_idx}...")
                print(f"üìä Progress: {len(meta_hypotheses_generated)}/{n_per_meta} accepted hypotheses")

                # ALWAYS select new chunks for EVERY hypothesis attempt to ensure maximum diversity
                if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                    print(f"üîÑ Selecting new chunks for hypothesis attempt {attempts}...")
                    randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                    dynamic_chunks = self.select_dynamic_chunks_for_generation(meta_hypothesis, randomization_strategy=randomization_strategy)
                    if dynamic_chunks:
                        context_chunks = dynamic_chunks
                        print(f"üìö Using {len(context_chunks)} dynamically selected chunks for generation")
                    else:
                        print(f"‚ö†Ô∏è  Failed to select dynamic chunks, falling back to direct retrieval...")
                        # Fallback: get fresh chunks with deprioritization for citation diversity
                        context_chunks = self.retrieve_relevant_chunks_with_deprioritization(meta_hypothesis, top_k=chunks_per_meta)
                        if context_chunks:
                            print(f"üìö Using {len(context_chunks)} fresh chunks for generation")
                        else:
                            print(f"‚ùå No chunks available for meta-hypothesis {meta_idx}. Skipping...")
                            break
                else:
                    print(f"üìö Using direct chunk retrieval (no dynamic selection available)...")
                    # Get fresh chunks with deprioritization for each hypothesis attempt to ensure citation diversity
                    context_chunks = self.retrieve_relevant_chunks_with_deprioritization(meta_hypothesis, top_k=chunks_per_meta)
                    if not context_chunks:
                        print(f"‚ùå No relevant context found for meta-hypothesis {meta_idx}. Skipping...")
                        break
                    print(f"üìö Using {len(context_chunks)} fresh chunks for generation")

                # Generate hypothesis without time limit

                # Generate hypothesis
                try:
                    # Context chunks are already filtered during retrieval
                    context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                    hypothesis_list = self.hypothesis_generator.generate(context_texts, n=1, meta_hypothesis=meta_hypothesis)

                    if not hypothesis_list:
                        print("‚ùå Failed to generate hypothesis. Skipping...")
                        continue

                    hypothesis = hypothesis_list[0]
                    
                    # Validate hypothesis format before proceeding
                    from src.ai.hypothesis_tools import validate_hypothesis_format
                    is_valid_format, format_reason = validate_hypothesis_format(hypothesis)
                    if not is_valid_format:
                        print(f"‚ùå Hypothesis rejected due to format: {format_reason}")
                        print(f"‚è≠Ô∏è  Skipping critique for format-rejected hypothesis")
                        # Record format rejection without critique
                        format_rejected_record = {
                            'Meta_Hypothesis': meta_hypothesis,
                            'Meta_Hypothesis_Index': meta_idx,
                            'Hypothesis': hypothesis,
                            'Accuracy': None,
                            'Novelty': None,
                            'Relevancy': None,
                            'Verdict': 'FORMAT_REJECTED',
                            'Critique': f'Hypothesis rejected: {format_reason}',
                            'Citations': 'No citations available',
                            'Original_Query': user_prompt
                        }
                        self.hypothesis_records.append(format_rejected_record)
                        # Save incrementally
                        self.save_hypothesis_record_incrementally(format_rejected_record, excel_filename)
                        continue
                    
                    print(f"üìù Generated: {hypothesis}")

                    # Proceed with critique

                    # Critique hypothesis
                    print(f"üîç Critiquing hypothesis...")
                    critique_result = self.hypothesis_critic.critique(hypothesis, context_texts, prompt=meta_hypothesis, lab_goals=get_lab_goals(), meta_hypothesis=meta_hypothesis)

                    if critique_result:
                        novelty = critique_result.get('novelty', 0)
                        accuracy = critique_result.get('accuracy', 0)
                        relevancy = critique_result.get('relevancy', 0)
                        critique = critique_result.get('critique', 'No critique available')

                        # Handle None values by converting to 0
                        if novelty is None:
                            novelty = 0
                        if accuracy is None:
                            accuracy = 0
                        if relevancy is None:
                            relevancy = 0

                        # Determine verdict
                        verdict = self.automated_verdict(accuracy, novelty, relevancy)

                        # Extract citations from the dynamically selected chunks used for this specific hypothesis
                        citations = []
                        unique_sources = set()
                        
                        # Get metadata from the dynamically selected chunks
                        dynamic_metadata = []
                        if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                            # If we used dynamic chunk selection, get metadata from those chunks
                            if hasattr(self, 'last_dynamic_results') and self.last_dynamic_results:
                                dynamic_metadata = [result.get('metadata', {}) for result in self.last_dynamic_results]
                            else:
                                # Fallback: extract from context_chunks if they have metadata
                                dynamic_metadata = []
                                for chunk in context_chunks:
                                    if isinstance(chunk, dict) and 'metadata' in chunk:
                                        dynamic_metadata.append(chunk['metadata'])
                                    else:
                                        # If chunks don't have metadata, use extract_citations_from_chunks as fallback
                                        citations = self.extract_citations_from_chunks(context_chunks)
                                        break
                        else:
                            # If not using dynamic selection, use extract_citations_from_chunks
                            citations = self.extract_citations_from_chunks(context_chunks)
                        
                        # Extract citations from the metadata if we have it
                        if not citations and dynamic_metadata:
                            for metadata in dynamic_metadata:
                                if metadata:
                                    source_name = metadata.get('source_name', metadata.get('source', 'Unknown'))
                                    title = metadata.get('title', 'No title')
                                    doi = metadata.get('doi', 'No DOI')
                                    
                                    # Try multiple field names for authors
                                    authors = (metadata.get('author') or 
                                             metadata.get('authors') or 
                                             metadata.get('author_list') or 
                                             'Unknown authors')
                                    
                                    # Try multiple field names for journal
                                    journal = (metadata.get('journal') or 
                                              metadata.get('journal_name') or 
                                              metadata.get('publication_venue') or 
                                              'Unknown journal')
                                    
                                    # Extract year from publication_date or other date fields
                                    publication_date = (metadata.get('publication_date') or 
                                                       metadata.get('date') or 
                                                       metadata.get('year') or 
                                                       '')
                                    
                                    if publication_date and len(str(publication_date)) >= 4:
                                        year = str(publication_date)[:4]
                                    else:
                                        year = 'Unknown year'

                                    citation = {
                                        'source_name': source_name,
                                        'title': title,
                                        'doi': doi,
                                        'authors': authors,
                                        'journal': journal,
                                        'year': year
                                    }

                                    # Add to citations if not already present
                                    citation_key = doi if doi != 'No DOI' else title
                                    if citation_key not in unique_sources:
                                        citations.append(citation)
                                        unique_sources.add(citation_key)
                        
                        formatted_citations = self.format_citations_for_export(citations)

                        # Track papers used for deprioritization to ensure citation diversity
                        for chunk in context_chunks:
                            if isinstance(chunk, dict) and 'metadata' in chunk:
                                metadata = chunk['metadata']
                                paper_id = self._get_paper_identifier(metadata)
                                if paper_id:
                                    self.paper_usage_count[paper_id] = self.paper_usage_count.get(paper_id, 0) + 1
                                    self.used_papers.add(paper_id)

                        # Create record
                        record = {
                            'Meta_Hypothesis': meta_hypothesis,
                            'Meta_Hypothesis_Index': meta_idx,
                            'Hypothesis': hypothesis,
                            'Accuracy': accuracy,
                            'Novelty': novelty,
                            'Relevancy': relevancy,
                            'Verdict': verdict,
                            'Critique': critique,
                            'Citations': formatted_citations,
                            'Original_Query': user_prompt
                        }

                        # Add to records
                        self.hypothesis_records.append(record)

                        # Save incrementally to Excel
                        self.save_hypothesis_record_incrementally(record, excel_filename)

                        # Display result
                        print(f"\nüìä Hypothesis Result:")
                        print(f"   Novelty: {novelty}/4")
                        print(f"   Accuracy: {accuracy}/4")
                        print(f"   Relevancy: {relevancy}/3")
                        print(f"   Verdict: {verdict}")

                        # Print rejection reason if hypothesis is rejected
                        if verdict != 'ACCEPTED':
                            if accuracy < 4:
                                print(f"‚ùå Hypothesis rejected: Low accuracy score ({accuracy}/4)")
                            elif novelty < 4:
                                print(f"‚ùå Hypothesis rejected: Low novelty score ({novelty}/4)")
                            elif relevancy < 3:
                                print(f"‚ùå Hypothesis rejected: Low relevancy score ({relevancy}/3)")
                            else:
                                print(f"‚ùå Hypothesis rejected: Unknown reason")

                        # Accept hypothesis if it meets criteria
                        if verdict == "ACCEPTED":
                            meta_hypotheses_generated.append(record)
                            print(f"‚úÖ Hypothesis accepted for meta-hypothesis {meta_idx}! ({len(meta_hypotheses_generated)}/{n_per_meta})")
                            
                            # Check if we have enough accepted hypotheses for this meta-hypothesis
                            if len(meta_hypotheses_generated) >= n_per_meta:
                                print(f"üéØ Successfully generated {n_per_meta} accepted hypotheses for meta-hypothesis {meta_idx}. Moving to next meta-hypothesis.")
                                break
                        else:
                            print(f"‚ùå Hypothesis rejected for meta-hypothesis {meta_idx}")

                        # Continue generating hypotheses until we have enough accepted ones

                except Exception as e:
                    print(f"‚ùå Error generating/critiquing hypothesis: {e}")
                    continue

            # Check if we successfully generated enough accepted hypotheses for this meta-hypothesis
            if len(meta_hypotheses_generated) < n_per_meta:
                print(f"‚ö†Ô∏è  Warning: Only generated {len(meta_hypotheses_generated)}/{n_per_meta} accepted hypotheses for meta-hypothesis {meta_idx}")
                print(f"   This may indicate that the meta-hypothesis is too specific or the acceptance criteria are too strict.")
            else:
                print(f"‚úÖ Successfully generated {len(meta_hypotheses_generated)}/{n_per_meta} accepted hypotheses for meta-hypothesis {meta_idx}")
            
            all_hypotheses.extend(meta_hypotheses_generated)

        # Summary
        print(f"\nüéâ META-HYPOTHESIS GENERATION COMPLETE")
        print("=" * 80)
        print(f"Total meta-hypotheses processed: {total_meta_hypotheses}")
        print(f"Total accepted hypotheses generated: {len(all_hypotheses)}")
        print(f"Average accepted hypotheses per meta-hypothesis: {len(all_hypotheses)/total_meta_hypotheses:.1f}")

        # Final export summary
        if all_hypotheses:
            print(f"\nüíæ Final export summary:")
            print(f"üìä All results have been incrementally saved to: {excel_filename}")
            print(f"üìÅ File location: {os.path.dirname(os.path.abspath(excel_filename))}")

            # Create comprehensive export with popup notification
            print(f"\nüì¶ Creating comprehensive export...")
            comprehensive_export_path = self.create_comprehensive_export()
            if comprehensive_export_path:
                print(f"üìã Comprehensive results exported to: {comprehensive_export_path}")
            else:
                print(f"‚ö†Ô∏è  Comprehensive export failed, but incremental saves are complete")
            
            # Final progress callback
            if progress_callback:
                progress_callback(f"‚úÖ Meta-hypothesis generation completed! Generated {len(all_hypotheses)} hypotheses", 100)
        else:
            print(f"‚ùå No hypotheses were successfully generated")
            if progress_callback:
                progress_callback("‚ùå No hypotheses were successfully generated", 0)

        return all_hypotheses

    def check_api_status(self):
        """Check Gemini API status and provide guidance."""
        print("\nüîç Checking Gemini API Status...")

        if not self.gemini_client:
            print("‚ùå Gemini client not initialized")
            print("üí° Check your API key in keys.json")
            return

        # Show current rate limit status
        remaining_requests = self.gemini_rate_limiter.get_remaining_requests()
        print(f"üìä Gemini API Rate Limit Status:")
        print(f"   - Limit: 1000 requests per minute")
        print(f"   - Remaining requests: {remaining_requests}")
        print(f"   - Used requests: {1000 - remaining_requests}")

        try:
            # Try a simple test request
            test_response = self.gemini_client.models.generate_content(
                model="gemini-2.5-flash",
                contents="Hello"
            )
            print("‚úÖ Gemini API is working")
            print("üí° You can use 'generate' command for hypothesis generation")
        except Exception as e:
            error_str = str(e)
            if "429" in error_str or "RESOURCE_EXHAUSTED" in error_str or "quota" in error_str.lower():
                print("‚ùå API quota exceeded (rate limited)")
                print("üí° Solutions:")
                print("   - Wait 30-60 minutes for quota reset")
                print("   - Use 'generate_offline' for fallback hypotheses")
                print("   - Check your billing/plan at https://ai.google.dev/")
                print("   - Consider upgrading your plan for higher limits")
            elif "401" in error_str or "403" in error_str:
                print("‚ùå API authentication failed")
                print("üí° Check your API key in keys.json")
            else:
                print(f"‚ùå API error: {e}")
                print("üí° Try again later or use 'generate_offline'")

    def is_chromadb_ready(self):
        """Check if ChromaDB is ready and has data for searching."""
        if not self.use_chromadb or not self.chroma_manager:
            return False

        try:
            stats = self.chroma_manager.get_collection_stats()
            return stats.get('total_documents', 0) > 0
        except Exception as e:
            logger.error(f"Error checking ChromaDB readiness: {e}")
            return False

    def _initialize_hypothesis_tools(self):
        """Initialize the HypothesisGenerator, HypothesisCritic, and MetaHypothesisGenerator."""
        if self.gemini_client:
            self.hypothesis_generator = HypothesisGenerator(model=self.gemini_client)
            def embedding_fn(text):
                return np.array(self.get_google_embedding(text))
            self.hypothesis_critic = HypothesisCritic(model=self.gemini_client, embedding_fn=embedding_fn)
            self.meta_hypothesis_generator = MetaHypothesisGenerator(model=self.gemini_client)
            
            # Connect rate limiter to hypothesis tools
            self.hypothesis_generator.rate_limiter = self.gemini_rate_limiter
            self.hypothesis_critic.rate_limiter = self.gemini_rate_limiter
            self.meta_hypothesis_generator.rate_limiter = self.gemini_rate_limiter
            
            print("‚úÖ Hypothesis tools initialized successfully with rate limiting.")
        else:
            print("‚ö†Ô∏è Gemini client not initialized, skipping hypothesis tools.")

    def get_performance_stats(self):
        """Get performance statistics for the system."""
        avg_query_time = self.total_query_time / max(self.query_count, 1)
        cache_hit_rate = len(self.cache) / max(self.query_count, 1) * 100
        
        return {
            "total_queries": self.query_count,
            "total_query_time": self.total_query_time,
            "average_query_time": avg_query_time,
            "cache_size": len(self.cache),
            "cache_hit_rate": cache_hit_rate,
            "performance_log_entries": len(self.performance_log)
        }

    def clear_cache(self):
        """Clear the performance cache."""
        self.cache.clear()
        logger.info("üßπ Performance cache cleared")

    def log_performance(self, operation: str, duration: float, details: dict = None):
        """Log performance metrics for an operation."""
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "operation": operation,
            "duration": duration,
            "details": details or {}
        }
        
        self.performance_log.append(log_entry)
        
        # Keep only last 1000 entries to prevent memory issues
        if len(self.performance_log) > 1000:
            self.performance_log = self.performance_log[-1000:]
        
        logger.info(f"üìä Performance: {operation} took {duration:.2f}s")

    def optimize_memory_usage(self):
        """Optimize memory usage by clearing unused data."""
        # Clear cache if it's too large
        if len(self.cache) > self.cache_max_size:
            # Keep only most recent entries
            cache_items = list(self.cache.items())
            self.cache = dict(cache_items[-self.cache_max_size//2:])
            logger.info(f"üßπ Reduced cache size to {len(self.cache)} entries")
        
        # Clear used papers set if it's too large
        if len(self.used_papers) > 10000:
            self.used_papers.clear()
            logger.info("üßπ Cleared used papers set")
        
        # Clear performance log if it's too large
        if len(self.performance_log) > 1000:
            self.performance_log = self.performance_log[-500:]
            logger.info("üßπ Reduced performance log size")

    def export_hypotheses_to_excel(self, filename=None):
        """Export hypothesis records to Excel file with grouping by meta-hypothesis."""
        if not self.hypothesis_records:
            print("‚ùå No hypothesis records to export.")
            return None

        try:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if filename is None:
                # Ensure hypothesis_export directory exists
                export_dir = "hypothesis_export"
                os.makedirs(export_dir, exist_ok=True)
                filename = os.path.join(export_dir, f"hypothesis_export_{timestamp}.xlsx")

            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Check if we have meta-hypothesis data
            has_meta_data = any('Meta_Hypothesis' in record for record in self.hypothesis_records)

            if has_meta_data:
                # Group by meta-hypothesis
                meta_groups = {}
                for record in self.hypothesis_records:
                    meta_hypothesis = record.get('Meta_Hypothesis', 'Unknown')
                    if meta_hypothesis not in meta_groups:
                        meta_groups[meta_hypothesis] = []
                    meta_groups[meta_hypothesis].append(record)

                # Create workbook with multiple sheets
                workbook = Workbook()
                workbook.remove(workbook.active)  # Remove default sheet

                # Define styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                meta_header_font = Font(bold=True, color="FFFFFF", size=12)
                meta_header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Create summary sheet
                summary_sheet = workbook.create_sheet("Summary")
                summary_sheet['A1'] = "Meta-Hypothesis Summary"
                summary_sheet['A1'].font = meta_header_font
                summary_sheet['A1'].fill = meta_header_fill

                summary_sheet['A3'] = "Meta-Hypothesis"
                summary_sheet['B3'] = "Hypotheses Generated"
                summary_sheet['C3'] = "Accepted"
                summary_sheet['D3'] = "Rejected"
                summary_sheet['E3'] = "Average Novelty"
                summary_sheet['F3'] = "Average Accuracy"

                for col in ['A3', 'B3', 'C3', 'D3', 'E3', 'F3']:
                    summary_sheet[col].font = header_font
                    summary_sheet[col].fill = header_fill
                    summary_sheet[col].border = border

                row = 4
                for meta_hypothesis, records in meta_groups.items():
                    summary_sheet[f'A{row}'] = meta_hypothesis[:100] + "..." if len(meta_hypothesis) > 100 else meta_hypothesis
                    summary_sheet[f'B{row}'] = len(records)

                    accepted_count = sum(1 for r in records if r.get('Verdict') == 'ACCEPTED')
                    rejected_count = len(records) - accepted_count
                    summary_sheet[f'C{row}'] = accepted_count
                    summary_sheet[f'D{row}'] = rejected_count

                    avg_novelty = sum(r.get('Novelty', 0) for r in records if r.get('Novelty') is not None) / len(records)
                    avg_accuracy = sum(r.get('Accuracy', 0) for r in records if r.get('Accuracy') is not None) / len(records)
                    summary_sheet[f'E{row}'] = round(avg_novelty, 1)
                    summary_sheet[f'F{row}'] = round(avg_accuracy, 1)

                    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                        summary_sheet[f'{col}{row}'].border = border

                    row += 1

                # Auto-adjust column widths for summary
                for column in summary_sheet.columns:
                    max_length = 0
                    try:
                        column_letter = column[0].column_letter
                    except AttributeError:
                        # Skip merged cells that don't have column_letter attribute
                        continue
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    summary_sheet.column_dimensions[column_letter].width = adjusted_width

                # Create detailed sheets for each meta-hypothesis
                for i, (meta_hypothesis, records) in enumerate(meta_groups.items(), 1):
                    # Create sheet name (Excel has 31 character limit)
                    sheet_name = f"Meta_{i}"
                    if len(meta_hypothesis) > 20:
                        sheet_name += f"_{meta_hypothesis[:20].replace(' ', '_')}"
                    else:
                        sheet_name += f"_{meta_hypothesis.replace(' ', '_')}"

                    # Clean sheet name for Excel compatibility
                    sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in ('_', '-'))[:31]

                    sheet = workbook.create_sheet(sheet_name)

                    # Add meta-hypothesis header
                    sheet['A1'] = f"Meta-Hypothesis {i}: {meta_hypothesis}"
                    sheet['A1'].font = meta_header_font
                    sheet['A1'].fill = meta_header_fill
                    sheet.merge_cells('A1:H1')

                    # Add hypothesis data
                    df_group = pd.DataFrame(records)
                    if not df_group.empty:
                        # Write headers
                        headers = list(df_group.columns)
                        for col, header in enumerate(headers, 1):
                            cell = sheet.cell(row=3, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = border

                        # Write data
                        for row_idx, record in enumerate(records, 4):
                            for col_idx, (key, value) in enumerate(record.items(), 1):
                                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                                cell.border = border

                        # Auto-adjust column widths
                        for column in sheet.columns:
                            max_length = 0
                            try:
                                column_letter = column[0].column_letter
                            except AttributeError:
                                # Skip merged cells that don't have column_letter attribute
                                continue
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            sheet.column_dimensions[column_letter].width = adjusted_width

                # Save workbook
                workbook.save(filename)

            else:
                # Fallback to simple export for non-meta-hypothesis data
                df = pd.DataFrame(self.hypothesis_records)
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Hypotheses', index=False)
                    worksheet = writer.sheets['Hypotheses']
                    for column in worksheet.columns:
                        max_length = 0
                        try:
                            column_letter = column[0].column_letter
                        except AttributeError:
                            # Skip merged cells that don't have column_letter attribute
                            continue
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"‚úÖ Successfully exported {len(self.hypothesis_records)} hypothesis records to: {filename}")
            print(f"üìÅ File saved in: {os.path.dirname(os.path.abspath(filename))}")
            if has_meta_data:
                print(f"üìä Organized into {len(meta_groups)} meta-hypothesis groups with summary sheet")
            return filename
        except Exception as e:
            print(f"‚ùå Failed to export hypotheses to Excel: {e}")
            return None

    def save_hypothesis_record_incrementally(self, record, filename=None):
        """Save a single hypothesis record to Excel file incrementally."""
        if filename is None:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Ensure hypothesis_export directory exists
            export_dir = "hypothesis_export"
            os.makedirs(export_dir, exist_ok=True)
            filename = os.path.join(export_dir, f"hypothesis_export_{timestamp}.xlsx")

        try:
            import pandas as pd
            from openpyxl import load_workbook

            # Check if file exists
            if os.path.exists(filename):
                # Load existing workbook
                workbook = load_workbook(filename)

                # Get or create the Hypotheses sheet
                if 'Hypotheses' in workbook.sheetnames:
                    worksheet = workbook['Hypotheses']
                    # Find the next empty row
                    next_row = worksheet.max_row + 1
                else:
                    worksheet = workbook.create_sheet('Hypotheses')
                    next_row = 1
                    # Add headers for new sheet
                    headers = list(record.keys())
                    for col, header in enumerate(headers, 1):
                        worksheet.cell(row=1, column=col, value=header)
                    next_row = 2

                # Add the record data
                for col, value in enumerate(record.values(), 1):
                    worksheet.cell(row=next_row, column=col, value=value)

                # Save the workbook
                workbook.save(filename)

            else:
                # Create new file
                df_row = pd.DataFrame([record])
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df_row.to_excel(writer, sheet_name='Hypotheses', index=False)
                    worksheet = writer.sheets['Hypotheses']
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        try:
                            column_letter = column[0].column_letter
                        except AttributeError:
                            # Skip merged cells that don't have column_letter attribute
                            continue
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"üíæ Saved hypothesis record to: {filename}")
            print(f"üìÅ File saved in: {os.path.dirname(os.path.abspath(filename))}")
            return filename

        except Exception as e:
            print(f"‚ùå Failed to save hypothesis record incrementally: {e}")
            return None

    def clear_hypothesis_records(self):
        """Clear all stored hypothesis records."""
        self.hypothesis_records = []
        print("üóëÔ∏è Hypothesis records cleared.")

    def show_hypothesis_records(self):
        """Display current hypothesis records."""
        if not self.hypothesis_records:
            print("üìù No hypothesis records available.")
            return

        print(f"\nüìù Current Hypothesis Records ({len(self.hypothesis_records)} total):")
        print("=" * 80)

        for i, record in enumerate(self.hypothesis_records, 1):
            print(f"\n{i}. Hypothesis: {record['Hypothesis'][:100]}{'...' if len(record['Hypothesis']) > 100 else ''}")
            print(f"   Accuracy: {record['Accuracy']}, Novelty: {record['Novelty']}, Relevancy: {record['Relevancy']}, Verdict: {record['Verdict']}")
            print(f"   Citations: {len(record['Citations'].split(';')) if record['Citations'] != 'No citations available' else 0} sources")
            print("-" * 80)

    def _display_commands(self):
        """Display available commands."""
        print(f"\nüí° Available commands:")
        print(f"   - 'add <query>': Search, batch, and generate hypotheses with automatic diverse paper selection (e.g., 'add cancer')")
        print(f"   - 'add <N> <query>': Search, batch, and generate hypotheses from up to N results (e.g., 'add 2000 cancer')")
        print(f"   - 'add all <query>': Search, batch, and generate hypotheses from ALL found results (e.g., 'add all cancer')")
        print(f"   - 'meta <query>': Use meta-hypothesis generator to create 5 diverse research directions, then generate hypotheses for each (e.g., 'meta UBR-5 in cancer')")
        print(f"   - 'meta <chunks> <query>': Advanced usage - specify number of chunks per meta-hypothesis (e.g., 'meta 2000 UBR-5 in cancer')")
        print(f"   - 'clear': Clear current package")
        print(f"   - 'export': Export hypothesis records to Excel")
        print(f"   - 'records': Show current hypothesis records")
        print(f"   - 'clear_records': Clear hypothesis records")
        print(f"   - 'reset_papers': Reset used papers list (allow reusing papers)")
        print(f"   - 'reset_diverse': Reset diverse paper selection (get new set of 1500 papers)")
        print(f"   - 'papers_status': Show used papers tracking status")
        print(f"   - 'lab_ratio': Configure ratio of lab-authored papers (auto/fixed ratio)")
        print(f"   - 'preferred_authors': Configure preferred authors for search prioritization")
        print(f"   - 'randomization': Configure chunk randomization strategy")
        print(f"   - 'test_lab': Test lab paper detection with current query")
        print(f"   - 'diverse <query>': Manually select exactly 1500 different papers for hypothesis generation (diverse search is now DEFAULT)")
        print(f"   üí° Note: Large queries may take longer and use more API calls. Use 'clear' if you get errors.")
        print(f"   üìö Default: Diverse search with 1500 chunks from different papers for maximum variety!")
        print()

    def interactive_search(self):
        novelty_threshold = 4
        accuracy_threshold = 4
        relevancy_threshold = 3
        print("=== Enhanced RAG Query System ===")
        print("üîç Search across all your knowledge bases!")
        self._display_commands()
        while True:
            query = input("‚ùì Your question (or command): ").strip()
            if query.lower() == 'clear':
                self.clear_package()
            elif query.lower() == 'export':
                self.export_hypotheses_to_excel()
            elif query.lower() == 'records':
                self.show_hypothesis_records()
            elif query.lower() == 'clear_records':
                self.clear_hypothesis_records()
            elif query.lower() == 'reset_papers':
                self.reset_used_papers()
            elif query.lower() == 'papers_status':
                self.show_used_papers_status()
            elif query.lower() == 'reset_diverse':
                if (hasattr(self, 'current_package') and 
                    self.current_package.get('method') == 'diverse_paper_selection'):
                    print("üîÑ Resetting diverse paper selection...")
                    self.current_package = {}
                    print("‚úÖ Diverse paper selection reset. Use 'diverse <query>' to select a new set of 1500 papers.")
                else:
                    print("‚ÑπÔ∏è  No diverse paper selection to reset.")
            elif query.lower() == 'lab_ratio':
                self.configure_lab_paper_ratio()
            elif query.lower() == 'preferred_authors':
                self.configure_preferred_authors()
            elif query.lower() == 'randomization':
                self.configure_randomization_strategy()
            elif query.lower() == 'test_lab':
                # Get the current query from package or use default
                if "prompt" in self.current_package and self.current_package["prompt"]:
                    test_query = self.current_package["prompt"]
                else:
                    test_query = "UBR5 cancer"
                self.test_lab_paper_detection(test_query)
            elif query.lower().startswith('diverse '):
                try:
                    parts = query.split(' ', 1)
                    if len(parts) < 2:
                        print("‚ùå Usage: diverse <query> (e.g., 'diverse UBR-5 in cancer')")
                        continue
                    
                    diverse_query = parts[1]
                    print(f"\nüéØ Selecting exactly 1500 different papers for hypothesis generation...")
                    print(f"üîç Query: '{diverse_query}'")
                    print(f"üìö This ensures maximum diversity across all sources and papers")
                    
                    response = input("Continue? (y/n): ").lower()
                    if response == 'y':
                        try:
                            # Use the new diversity function
                            diverse_chunks = self.select_diverse_papers_for_hypothesis_generation(
                                diverse_query, 
                                target_papers=1500
                            )
                            
                            if diverse_chunks:
                                print(f"\n‚úÖ Successfully selected {len(diverse_chunks)} chunks from exactly 1500 different papers!")
                                print(f"üéØ Each hypothesis generation will now use a completely different set of source chunks")
                                print(f"üìä Papers are prioritized by: lab authorship, preferred authors, citation count, and impact factor")
                                
                                # Store the diverse chunks for future hypothesis generation
                                self.current_package = {
                                    "chunks": diverse_chunks,
                                    "prompt": diverse_query,
                                    "method": "diverse_paper_selection"
                                }
                                
                                print(f"üì¶ Diverse paper package stored. Use 'add' command to generate hypotheses with these diverse chunks.")
                            else:
                                print("‚ùå Failed to select diverse papers.")
                        except Exception as e:
                            print(f"‚ùå Error during diverse paper selection: {e}")
                    else:
                        print("Diverse paper selection cancelled.")
                except Exception as e:
                    print(f"‚ùå Error processing diverse command: {e}")
            elif query.lower().startswith('meta '):
                try:
                    parts = query.split(' ', 1)
                    if len(parts) < 2:
                        print("‚ùå Usage: meta <query> (e.g., 'meta UBR-5 in cancer')")
                        print("   Advanced: meta <chunks> <query> (e.g., 'meta 2000 UBR-5 in cancer' for 2000 chunks per meta-hypothesis)")
                        continue

                    # Check if first part is a number (chunks specification)
                    query_parts = parts[1].split(' ', 1)
                    if len(query_parts) > 1 and query_parts[0].isdigit():
                        chunks_per_meta = int(query_parts[0])
                        # Validate chunks number
                        if chunks_per_meta < 5:
                            print(f"‚ùå Too few chunks ({chunks_per_meta}). Minimum is 5 chunks per meta-hypothesis.")
                            continue
                        elif chunks_per_meta > 5000:
                            print(f"‚ö†Ô∏è  Very large number of chunks ({chunks_per_meta}). This may cause API errors or timeouts.")
                            response = input("Continue anyway? (y/n): ").lower()
                            if response != 'y':
                                continue
                        meta_query = query_parts[1]
                        print(f"\nüß† Meta-hypothesis generation for: '{meta_query}'")
                        print(f"üìö Using {chunks_per_meta} chunks per meta-hypothesis")
                    else:
                        chunks_per_meta = 1500  # default
                        meta_query = parts[1]
                        print(f"\nüß† Meta-hypothesis generation for: '{meta_query}'")
                        print(f"üìö Using default {chunks_per_meta} chunks per meta-hypothesis")

                    print("This will generate 5 diverse research directions, then create hypotheses for each.")
                    response = input("Continue? (y/n): ").lower()
                    if response == 'y':
                        try:
                            results = self.generate_hypotheses_with_meta_generator(meta_query, n_per_meta=5, chunks_per_meta=chunks_per_meta)
                            if results:
                                print(f"\n‚úÖ Meta-hypothesis generation complete! Generated {len(results)} accepted hypotheses across 5 research directions.")
                            else:
                                print("‚ùå Meta-hypothesis generation failed.")
                        except Exception as e:
                            print(f"‚ùå Error during meta-hypothesis generation: {e}")
                    else:
                        print("Meta-hypothesis generation cancelled.")
                except Exception as e:
                    print(f"‚ùå Error processing meta command: {e}")
            elif query.lower().startswith('add '):
                try:
                    parts = query.split(' ', 2)
                    if len(parts) < 2:
                        print("‚ùå Usage: add <query> (e.g., 'add cancer')")
                        continue
                    # Check for 'all' keyword
                    if len(parts) > 2 and parts[1].lower() == 'all':
                        num_results = None  # None means all
                        search_query = parts[2]
                    elif len(parts) > 2:
                        try:
                            num_results = int(parts[1])
                            if num_results <= 0:
                                print("‚ùå Number of results must be positive")
                                continue
                            if num_results > 100000:
                                print("‚ö†Ô∏è  Number of results is very large (>100,000). This may use a lot of memory and cause API errors.")
                            search_query = parts[2]
                        except ValueError:
                            num_results = 100000
                            search_query = parts[1] + " " + parts[2]
                    else:
                        num_results = 100000
                        search_query = parts[1]
                    print(f"\nüîç Searching for: '{search_query}'" + (f" (requesting ALL results)" if num_results is None else f" (requesting up to {num_results} results)"))
                    try:
                        # Decide which method to use
                        if num_results is None or num_results > 5000:
                            # Use get_all_documents and compute similarity in Python
                            print("‚ö†Ô∏è  Using full-document scan and in-memory similarity search. This may use a lot of RAM and be slow for very large collections.")
                            # Get all documents (limit to 100,000 for safety)
                            max_docs = num_results if (num_results is not None and num_results < 100000) else 100000
                            all_docs = self.chroma_manager.get_all_documents(limit=max_docs)
                            if not all_docs:
                                print("‚ùå No documents found in ChromaDB.")
                                continue
                            print(f"üì¶ Retrieved {len(all_docs)} documents from ChromaDB. Computing similarities...")
                            # Get query embedding
                            query_embedding = self.get_google_embedding(search_query)
                            if not query_embedding:
                                print("‚ùå Failed to get query embedding.")
                                continue
                            import numpy as np
                            from tqdm.auto import tqdm
                            doc_embeddings = []
                            for doc in all_docs:
                                emb = doc['metadata'].get('embedding', None)
                                if emb is not None:
                                    doc_embeddings.append(emb)
                                else:
                                    doc_embeddings.append([0.0]*len(query_embedding))  # fallback if missing
                            doc_embeddings = np.array(doc_embeddings)
                            from sklearn.metrics.pairwise import cosine_similarity
                            similarities = np.zeros(len(doc_embeddings))
                            print(f"üîç Computing similarities across {len(doc_embeddings):,} documents...")
                            with tqdm(total=len(doc_embeddings), desc="Similarity", unit="doc") as pbar:
                                batch_size = 1000
                                for start in range(0, len(doc_embeddings), batch_size):
                                    end = min(start + batch_size, len(doc_embeddings))
                                    similarities[start:end] = cosine_similarity([query_embedding], doc_embeddings[start:end])[0]
                                    pbar.update(end - start)
                            top_indices = np.argsort(similarities)[::-1][:num_results if num_results is not None else len(all_docs)]
                            results = []
                            for idx in top_indices:
                                results.append({
                                    'chunk': all_docs[idx]['document'],
                                    'metadata': all_docs[idx]['metadata'],
                                    'similarity': similarities[idx],
                                    'method': 'manual_cosine'
                                })
                        else:
                            # Use fast ChromaDB search
                            results = self.search_hybrid(search_query, top_k=num_results, filter_dict=None)
                        if not results:
                            print("‚ùå No results found. Try a different search term or check if data is loaded.")
                            continue
                        print(f"üì¶ Found {len(results)} relevant chunks. Beginning batch processing...")
                        # Batch process for hypothesis generation
                        batch_size = 500
                        
                        # AUTOMATICALLY use diverse paper selection as default for hypothesis generation
                        print(f"\nüéØ Automatically selecting diverse papers for hypothesis generation...")
                        print(f"üìö This ensures maximum diversity across all sources and papers")
                        
                        diverse_chunks = self.select_diverse_papers_for_hypothesis_generation(
                            search_query, 
                            target_papers=1500
                        )
                        
                        if diverse_chunks:
                            print(f"‚úÖ Successfully selected {len(diverse_chunks)} chunks from exactly 1500 different papers!")
                            print(f"üéØ Each hypothesis generation will now use a completely different set of source chunks")
                            print(f"üìä Papers are prioritized by: lab authorship, preferred authors, citation count, and impact factor")
                            all_chunks = diverse_chunks
                        else:
                            print(f"‚ö†Ô∏è  Failed to select diverse papers, falling back to original chunks")
                            all_chunks = [r['chunk'] for r in results]
                        # Sequential, real-time generator-critic loop for 5 accepted hypotheses
                        print(f"\nüß† Generating hypotheses one at a time until 5 are accepted (novelty >= {novelty_threshold}, accuracy >= {accuracy_threshold}, 5-minute time limit per critique)...")
                        accepted_hypotheses = []
                        attempts = 0
                        max_attempts = 100  # Prevent infinite loops
                        while len(accepted_hypotheses) < 5 and attempts < max_attempts:
                            attempts += 1
                            print(f"\nüß† Generating hypothesis attempt {attempts}...")

                            # Use the diverse chunks for all attempts (no need to re-select)
                            print(f"üéØ Using diverse paper chunks for hypothesis attempt {attempts}...")
                            print(f"üìö Using {len(all_chunks)} chunks from exactly 1500 different papers")

                            # Start timer for this hypothesis
                            self.hypothesis_timer.start()
                            print(f"‚è±Ô∏è  Starting 5-minute timer for hypothesis attempt {attempts}...")

                            hypothesis_list = self._generate_hypotheses_with_retry(all_chunks, n=1)
                            if not hypothesis_list:
                                print("‚ùå Failed to generate hypothesis. Skipping...")
                                # Record failed generation
                                self.hypothesis_records.append({
                                    'Hypothesis': f'Failed generation attempt {attempts}',
                                    'Accuracy': None,
                                    'Novelty': None,
                                    'Verdict': 'GENERATION_FAILED',
                                    'Critique': 'Failed to generate hypothesis',
                                    'Citations': 'No citations available'
                                })
                                continue

                            hypothesis = hypothesis_list[0]
                            print(f"Generated Hypothesis: {hypothesis}")
                            print(f"\nüîÑ Critiquing Hypothesis: {hypothesis}")

                            # Check timer before critique
                            if self.hypothesis_timer.check_expired():
                                print(f"‚è∞ Time limit reached for hypothesis attempt {attempts}. Moving to next attempt.")
                                # Record timeout result
                                self.hypothesis_records.append({
                                    'Hypothesis': hypothesis,
                                    'Accuracy': None,
                                    'Novelty': None,
                                    'Verdict': 'TIMEOUT',
                                    'Critique': 'Critique process timed out after 5 minutes',
                                    'Citations': 'No citations available'
                                })
                                continue

                            critique_result = self._critique_hypothesis_with_retry(hypothesis, all_chunks, search_query)
                            if not critique_result:
                                print(f"‚ùå Failed to critique hypothesis after retries. Skipping...")
                                # Record failed critique
                                self.hypothesis_records.append({
                                    'Hypothesis': hypothesis,
                                    'Accuracy': None,
                                    'Novelty': None,
                                    'Verdict': 'CRITIQUE_FAILED',
                                    'Critique': 'Failed to critique hypothesis after retries',
                                    'Citations': 'No citations available'
                                })
                                continue

                            # Check timer after critique
                            if self.hypothesis_timer.check_expired():
                                print(f"‚è∞ Time limit reached for hypothesis attempt {attempts} after critique. Moving to next attempt.")
                                # Record timeout result with partial critique
                                self.hypothesis_records.append({
                                    'Hypothesis': hypothesis,
                                    'Accuracy': critique_result.get('accuracy', None),
                                    'Novelty': critique_result.get('novelty', None),
                                    'Verdict': 'TIMEOUT',
                                    'Critique': critique_result.get('critique', '') + ' [Process timed out]',
                                    'Citations': 'No citations available'
                                })
                                continue

                            novelty = critique_result.get('novelty', 0)
                            accuracy = critique_result.get('accuracy', 0)
                            relevancy = critique_result.get('relevancy', 0)
                            
                            # Handle None values by converting to 0
                            if novelty is None:
                                novelty = 0
                            if accuracy is None:
                                accuracy = 0
                            if relevancy is None:
                                relevancy = 0
                                
                            verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold, novelty_threshold, relevancy_threshold)
                            remaining_time = self.hypothesis_timer.get_remaining_time()
                            print(f"   Novelty: {novelty}/{novelty_threshold}, Accuracy: {accuracy}/{accuracy_threshold}, Relevancy: {relevancy}/{relevancy_threshold}, Automated Verdict: {verdict}")
                            print(f"   ‚è±Ô∏è  Time remaining: {remaining_time:.1f}s")

                            # Extract citations from results metadata
                            citations = []
                            unique_sources = set()
                            for result in results:
                                metadata = result.get('metadata', {})
                                if metadata:
                                    source_name = metadata.get('source_name', metadata.get('source', 'Unknown'))
                                    title = metadata.get('title', 'No title')
                                    doi = metadata.get('doi', 'No DOI')
                                    
                                    # Try multiple field names for authors
                                    authors = (metadata.get('author') or 
                                             metadata.get('authors') or 
                                             metadata.get('author_list') or 
                                             'Unknown authors')
                                    
                                    # Try multiple field names for journal
                                    journal = (metadata.get('journal') or 
                                              metadata.get('journal_name') or 
                                              metadata.get('publication_venue') or 
                                              'Unknown journal')
                                    
                                    # Extract year from publication_date or other date fields
                                    publication_date = (metadata.get('publication_date') or 
                                                       metadata.get('date') or 
                                                       metadata.get('year') or 
                                                       '')
                                    
                                    if publication_date and len(str(publication_date)) >= 4:
                                        year = str(publication_date)[:4]
                                    else:
                                        year = 'Unknown year'

                                    citation = {
                                        'source_name': source_name,
                                        'title': title,
                                        'doi': doi,
                                        'authors': authors,
                                        'journal': journal,
                                        'year': year
                                    }

                                    # Add to citations if not already present
                                    citation_key = doi if doi != 'No DOI' else title
                                    if citation_key not in unique_sources:
                                        citations.append(citation)
                                        unique_sources.add(citation_key)

                            formatted_citations = self.format_citations_for_export(citations)

                            # Track record for export
                            self.hypothesis_records.append({
                                'Hypothesis': hypothesis,
                                'Accuracy': accuracy,
                                'Novelty': novelty,
                                'Relevancy': relevancy,
                                'Verdict': verdict,
                                'Critique': critique_result.get('critique', ''),
                                'Citations': formatted_citations
                            })

                            if verdict == 'ACCEPTED':
                                accepted_hypotheses.append({
                                    "hypothesis": hypothesis,
                                    "critique": critique_result,
                                    "score": (novelty + accuracy + relevancy) / 3
                                })
                                print(f"‚úÖ Hypothesis accepted! ({len(accepted_hypotheses)}/5)")
                            else:
                                print(f"‚ùå Hypothesis rejected (verdict: {verdict})")

                        if not accepted_hypotheses:
                            print("‚ùå No hypotheses were successfully critiqued.")
                            continue

                        print(f"\nüèÜ Top {len(accepted_hypotheses)} Hypotheses:")
                        print("=" * 80)
                        for i, result in enumerate(accepted_hypotheses, 1):
                            hypothesis = result["hypothesis"]
                            score = result["score"]
                            critique = result["critique"]
                            print(f"\n{i}. {hypothesis}")
                            print(f"   Score: {score:.1f}")
                            print(f"   Novelty: {critique.get('novelty', 'N/A')}")
                            print(f"   Accuracy: {critique.get('accuracy', 'N/A')}")
                            print(f"   Relevancy: {critique.get('relevancy', 'N/A')}")
                            print(f"   Automated Verdict: {self.automated_verdict(critique.get('accuracy', 0), critique.get('novelty', 0), critique.get('relevancy', 0), accuracy_threshold, novelty_threshold, relevancy_threshold)}")
                            print(f"   Critique: {critique.get('critique', 'N/A')}")
                            print("-" * 80)

                        # Auto-export results
                        print(f"\nüíæ Auto-exporting hypothesis records...")
                        export_filename = self.export_hypotheses_to_excel()
                        if export_filename:
                            print(f"üìä Results exported to: {export_filename}")
                            print(f"üìÅ File saved in: {os.path.dirname(os.path.abspath(export_filename))}")

                        # Enter discussion loop
                        print("\nüí¨ You can now discuss these hypotheses with the AI. Type the number of a hypothesis (1-5) to select it, or 'exit' to leave discussion mode.")
                        selected_idx = 0
                        while True:
                            user_input = input("[Discussion] > ").strip()
                            if user_input.lower() in ['exit', 'quit']:
                                print("Exiting discussion mode.")
                                break
                            if user_input.isdigit() and 1 <= int(user_input) <= len(accepted_hypotheses):
                                selected_idx = int(user_input) - 1
                                print(f"Selected Hypothesis {user_input}:\n{accepted_hypotheses[selected_idx]['hypothesis']}")
                                print("You can now ask questions about this hypothesis. Type 'back' to select another, or 'exit' to leave.")
                                while True:
                                    q = input(f"[H{selected_idx+1} Q&A] > ").strip()
                                    if q.lower() in ['exit', 'quit']:
                                        print("Exiting discussion mode.")
                                        return
                                    if q.lower() == 'back':
                                        print("Returning to hypothesis selection.")
                                        break
                                    # Use the critic to answer the question about the selected hypothesis
                                    hypothesis = accepted_hypotheses[selected_idx]['hypothesis']
                                    context_chunks = all_chunks if 'all_chunks' in locals() else []
                                    if self.hypothesis_critic and self.hypothesis_critic.model:
                                        # Build a prompt for Q&A
                                        prompt = f"You are an expert scientific reviewer. The user has a question about the following hypothesis.\n\nHypothesis:\n{hypothesis}\n\nUser Question:\n{q}\n\nPlease answer in detail, using the literature context if relevant."
                                        response = self.hypothesis_critic.model.models.generate_content(
                                            model="gemini-2.5-flash",
                                            contents=prompt
                                        )
                                        print(f"AI: {response.text.strip()}")
                                    else:
                                        print(f"AI: (No LLM available) This is a placeholder answer about '{hypothesis}'. User asked: {q}")
                            else:
                                print(f"Please enter a number between 1 and {len(accepted_hypotheses)}, or 'exit'.")
                    except Exception as e:
                        print(f"‚ùå Search or generation failed: {e}")
                        print("üí° Try again with a different query.")
                    self._display_commands()
                except Exception as e:
                    print(f"‚ùå Command parsing error: {e}")
                    print("‚ùå Usage: add <query> (e.g., 'add cancer')")
            elif query == '':
                continue
            else:
                print("‚ùå Unknown command. Available commands: 'add', 'clear', 'export', 'records', 'clear_records'")
                self._display_commands()

    def run_comprehensive_hypothesis_session(self, query, max_hypotheses=5, hypotheses_per_meta=None, progress_callback=None):
        """
        Run a comprehensive hypothesis generation session using META HYPOTHESIS GENERATION by default.
        This method now uses the meta-hypothesis generator approach which:
        - Generates 5 meta-hypotheses from the user query
        - For each meta-hypothesis, generates hypotheses until exactly 5 are accepted
        - Provides maximum diversity and comprehensive coverage
        
        Features:
        - Meta-hypothesis generation for diverse research directions
        - Timer-controlled critique process
        - Automated verdict determination
        - Citation tracking
        - Excel export
        - Progress callbacks for GUI integration
        """
        print("=" * 80)
        print("üß† COMPREHENSIVE META-HYPOTHESIS GENERATION SESSION")
        print("=" * 80)
        print(f"Query: {query}")
        print(f"Features: Meta-hypothesis generation, automated verdicts, citation tracking, Excel export")
        print("=" * 80)

        # Clear previous records
        self.clear_hypothesis_records()

        # Use meta hypothesis generation by default
        print(f"\nüß† Using META-HYPOTHESIS GENERATION for maximum diversity and comprehensive coverage...")
        if progress_callback:
            progress_callback("Generating 5 meta-hypotheses from user query...", 50)
        
        # Use user-specified hypotheses per meta-hypothesis, or calculate from max_hypotheses
        if hypotheses_per_meta is not None:
            n_per_meta = hypotheses_per_meta
            print(f"üìä Using user-specified value: {n_per_meta} hypotheses per meta-hypothesis")
        else:
            # Calculate how many hypotheses per meta-hypothesis we need
            # We want to generate 5 meta-hypotheses, and get max_hypotheses total accepted hypotheses
            n_per_meta = max(1, max_hypotheses // 5)  # Distribute hypotheses across meta-hypotheses
            print(f"üìä Calculated value: {n_per_meta} hypotheses per meta-hypothesis (from max_hypotheses={max_hypotheses})")
        
        # Use the meta hypothesis generation method
        all_hypotheses = self.generate_hypotheses_with_meta_generator(
            user_prompt=query,
            n_per_meta=n_per_meta,
            chunks_per_meta=1500,
            progress_callback=progress_callback
        )
        
        if all_hypotheses:
            print(f"\n‚úÖ Meta-hypothesis generation completed!")
            print(f"üìä Total accepted hypotheses generated: {len(all_hypotheses)}")
            if progress_callback:
                progress_callback("Meta-hypothesis generation completed successfully!", 90)
            
            # Return the hypotheses in the expected format
            accepted_hypotheses = []
            for record in all_hypotheses:
                if isinstance(record, dict):
                    hypothesis_text = record.get('Hypothesis', '')
                    if hypothesis_text:
                        accepted_hypotheses.append(record)
            
            return accepted_hypotheses[:max_hypotheses]  # Limit to requested number
        else:
            print("‚ùå Meta-hypothesis generation failed. Falling back to regular generation...")
            if progress_callback:
                progress_callback("Meta-hypothesis generation failed, falling back to regular generation...", 70)
            
            # Fallback to regular generation if meta generation fails
            return self._run_regular_hypothesis_generation(query, max_hypotheses)

    def _run_regular_hypothesis_generation(self, query, max_hypotheses=5):
        """
        Fallback method for regular hypothesis generation (original implementation).
        """
        print("=" * 80)
        print("üß† REGULAR HYPOTHESIS GENERATION (FALLBACK)")
        print("=" * 80)
        print(f"Query: {query}")
        print(f"Features: 5-minute timer per hypothesis, automated verdicts, citation tracking, Excel export")
        print("=" * 80)

        # Run hypothesis generation with deprioritization for maximum citation diversity
        print(f"\nüîç Searching for relevant context with deprioritization...")
        context_chunks = self.retrieve_relevant_chunks_with_deprioritization(query, top_k=1500)
        if not context_chunks:
            print("‚ùå No relevant context found.")
            return None

        print(f"üìö Found {len(context_chunks)} relevant context chunks")

        # Generate hypotheses with all features
        print(f"\nüß† Generating {max_hypotheses} hypotheses with comprehensive evaluation...")

        accepted_hypotheses = []
        attempts = 0
        max_attempts = 50  # Prevent infinite loops

        while len(accepted_hypotheses) < max_hypotheses and attempts < max_attempts:
            attempts += 1
            print(f"\nüîÑ Hypothesis Attempt {attempts}/{max_attempts}")
            print("-" * 60)

            # ALWAYS select new chunks for EVERY hypothesis attempt to ensure diversity
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                print(f"üîÑ Selecting new chunks with deprioritization for hypothesis attempt {attempts}...")
                randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                dynamic_chunks = self.select_dynamic_chunks_for_generation(query, randomization_strategy=randomization_strategy)
                if dynamic_chunks:
                    context_chunks = dynamic_chunks
                    print(f"üìö Using {len(context_chunks)} dynamically selected chunks for generation")
                else:
                    print(f"‚ö†Ô∏è  Failed to select dynamic chunks, using deprioritized retrieval...")
                    # Fallback to deprioritized retrieval
                    context_chunks = self.retrieve_relevant_chunks_with_deprioritization(query, top_k=1500)
                    print(f"üìö Using {len(context_chunks)} deprioritized chunks for generation")
            else:
                print(f"üìö Using deprioritized retrieval for generation...")
                context_chunks = self.retrieve_relevant_chunks_with_deprioritization(query, top_k=1500)
                print(f"üìö Using {len(context_chunks)} deprioritized chunks for generation")

            # Start timer
            self.hypothesis_timer.start()
            print(f"‚è±Ô∏è  Timer started (5-minute limit)")

            # Generate hypothesis
            try:
                # Context chunks are already filtered during retrieval
                context_texts = [chunk.get("document", "") if isinstance(chunk, dict) else chunk for chunk in context_chunks]
                hypothesis_list = self.hypothesis_generator.generate(context_texts, n=1, meta_hypothesis=query)
                # Add rate limiting delay to prevent hitting quota limits
                import time
                import random
                time.sleep(2.0 + random.uniform(0.2, 0.8))  # Increased delay for token quota

                if not hypothesis_list:
                    print("‚ùå Failed to generate hypothesis (all hypotheses rejected due to format issues)")
                    self.hypothesis_records.append({
                        'Hypothesis': f'Failed generation attempt {attempts} - all hypotheses rejected',
                        'Accuracy': None,
                        'Novelty': None,
                        'Verdict': 'GENERATION_FAILED',
                        'Critique': 'All generated hypotheses were rejected due to format validation failures',
                        'Citations': 'No citations available'
                    })
                    continue

                hypothesis = hypothesis_list[0]
                
                # Validate hypothesis format before proceeding
                from src.ai.hypothesis_tools import validate_hypothesis_format
                is_valid_format, format_reason = validate_hypothesis_format(hypothesis)
                if not is_valid_format:
                    print(f"‚ùå Hypothesis rejected due to format: {format_reason}")
                    print(f"‚è≠Ô∏è  Skipping critique for format-rejected hypothesis")
                    # Record format rejection without critique
                    format_rejected_record = {
                        'Hypothesis': hypothesis,
                        'Accuracy': None,
                        'Novelty': None,
                        'Verdict': 'FORMAT_REJECTED',
                        'Critique': f'Hypothesis rejected: {format_reason}',
                        'Citations': 'No citations available'
                    }
                    self.hypothesis_records.append(format_rejected_record)
                    continue
                
                print(f"üìù Generated: {hypothesis}")

                # Check timer before critique
                if self.hypothesis_timer.check_expired():
                    print(f"‚è∞ Time limit reached before critique")
                    self.hypothesis_records.append({
                        'Hypothesis': hypothesis,
                        'Accuracy': None,
                        'Novelty': None,
                        'Verdict': 'TIMEOUT',
                        'Critique': 'Process timed out before critique',
                        'Citations': 'No citations available'
                    })
                    continue

                # Critique hypothesis
                print(f"üîç Critiquing hypothesis...")
                critique_result = self.hypothesis_critic.critique(hypothesis, context_texts, prompt=query, lab_goals=get_lab_goals(), meta_hypothesis=query)
                # Add rate limiting delay to prevent hitting quota limits
                import time
                import random
                time.sleep(2.0 + random.uniform(0.2, 0.8))  # Increased delay for token quota

                # Check timer after critique
                if self.hypothesis_timer.check_expired():
                    print(f"‚è∞ Time limit reached after critique")
                    self.hypothesis_records.append({
                        'Hypothesis': hypothesis,
                        'Accuracy': critique_result.get('accuracy', None),
                        'Novelty': critique_result.get('novelty', None),
                        'Verdict': 'TIMEOUT',
                        'Critique': critique_result.get('critique', '') + ' [Process timed out]',
                        'Citations': 'No citations available'
                    })
                    continue

                # Extract scores and determine verdict
                novelty = critique_result.get('novelty', 0)
                accuracy = critique_result.get('accuracy', 0)
                relevancy = critique_result.get('relevancy', 0)
                
                # Handle None values by converting to 0
                if novelty is None:
                    novelty = 0
                if accuracy is None:
                    accuracy = 0
                if relevancy is None:
                    relevancy = 0
                    
                verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold=4, novelty_threshold=4, relevancy_threshold=3)
                remaining_time = self.hypothesis_timer.get_remaining_time()

                print(f"üìä Scores: Accuracy={accuracy}, Novelty={novelty}, Relevancy={relevancy}")
                print(f"‚öñÔ∏è  Automated Verdict: {verdict}")
                print(f"‚è±Ô∏è  Time remaining: {remaining_time:.1f}s")

                # Extract citations
                citations = self.extract_citations_from_chunks(context_chunks)
                formatted_citations = self.format_citations_for_export(citations)
                citation_count = len(citations)

                print(f"üìö Citations: {citation_count} unique sources")

                # Record hypothesis with enhanced citation data
                hypothesis_record = {
                    'Hypothesis': hypothesis,
                    'Accuracy': accuracy,
                    'Novelty': novelty,
                    'Relevancy': relevancy,
                    'Verdict': verdict,
                    'Critique': critique_result.get('critique', ''),
                    'Citations': formatted_citations,
                    'Citation_Count': citation_count,
                    'Raw_Citations': citations,  # Store structured citation data
                    'Context_Chunks': context_chunks  # Store context chunks for later processing
                }
                self.hypothesis_records.append(hypothesis_record)

                # Check if accepted
                if verdict == 'ACCEPTED':
                    accepted_hypotheses.append({
                        "hypothesis": hypothesis,
                        "critique": critique_result,
                        "score": (novelty + accuracy + relevancy) / 3
                    })
                    print(f"‚úÖ ACCEPTED! ({len(accepted_hypotheses)}/{max_hypotheses})")
                else:
                    print(f"‚ùå REJECTED (below thresholds)")

            except Exception as e:
                print(f"‚ùå Error during hypothesis processing: {e}")
                self.hypothesis_records.append({
                    'Hypothesis': f'Error in attempt {attempts}',
                    'Accuracy': None,
                    'Novelty': None,
                    'Verdict': 'ERROR',
                    'Critique': f'Error: {str(e)}',
                    'Citations': 'No citations available'
                })
                continue

        # Final summary
        print(f"\n" + "=" * 80)
        print(f"üèÅ REGULAR GENERATION SESSION COMPLETE")
        print(f"=" * 80)
        print(f"Total attempts: {attempts}")
        print(f"Accepted hypotheses: {len(accepted_hypotheses)}")
        print(f"Total records: {len(self.hypothesis_records)}")

        # Export results
        print(f"\nüíæ Exporting results to Excel...")
        export_filename = self.export_hypotheses_to_excel()

        if export_filename:
            print(f"‚úÖ Results exported to: {export_filename}")
            print(f"üìÅ File saved in: {os.path.dirname(os.path.abspath(export_filename))}")
            
            # Create comprehensive export with popup notification
            print(f"\nüì¶ Creating comprehensive export...")
            comprehensive_export_path = self.create_comprehensive_export()
            if comprehensive_export_path:
                print(f"üìã Comprehensive results exported to: {comprehensive_export_path}")
            else:
                print(f"‚ö†Ô∏è  Comprehensive export failed, but Excel export is complete")
        else:
            print(f"‚ùå Export failed")

        return accepted_hypotheses

    def generate_hypotheses_with_per_hypothesis_timer(self, n=5, max_rounds=10, filename=None):
        """
        Generate n hypotheses with optimized timer (1-minute timeout) and performance monitoring.
        After each iteration, print the result, time left, scores, and hypothesis text.
        Save every iteration in self.hypothesis_records and append to Excel after each iteration.
        Adds empty verifier columns separated by an empty column.
        """
        import pandas as pd
        from openpyxl import load_workbook
        from datetime import datetime
        if not self.current_package["chunks"]:
            print("‚ùå Package is empty. Add some chunks first.")
            return
        if not self.hypothesis_generator or not self.hypothesis_critic:
            print("‚ùå Hypothesis tools not initialized. Check if Gemini API key is available.")
            return
        novelty_threshold = 4
        accuracy_threshold = 4
        relevancy_threshold = 3
        print(f"\nüß† Generating {n} hypotheses (1-minute timer per hypothesis with performance monitoring)...")
        print(f"üì¶ Using {len(self.current_package['chunks'])} package chunks for critique")
        print(f"[INFO] Acceptance criteria: Novelty >= {novelty_threshold}, Accuracy >= {accuracy_threshold}, Relevancy >= {relevancy_threshold}")
        
        # Start performance monitoring
        start_time = time.time()
        self.log_performance("hypothesis_generation_start", 0, {"n_hypotheses": n, "max_rounds": max_rounds})
        package_chunks = self.current_package["chunks"]
        self.hypothesis_records = []  # Clear previous records
        # Prepare Excel file
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Ensure hypothesis_export directory exists
            export_dir = "hypothesis_export"
            os.makedirs(export_dir, exist_ok=True)
            filename = os.path.join(export_dir, f"hypothesis_export_{timestamp}.xlsx")
        # Write header row at the start, with empty separator and verifier columns
        columns = [
            'HypothesisNumber', 'Iteration', 'Status', 'Hypothesis', 'Novelty', 'Accuracy', 'Critique', 'TimeLeft',
            '',  # Empty separator column
            'Verifier Novelty Score (0-100)', 'Verifier Accuracy Score (0-100)', 'Verifier Verdict (accept, refuse)'
        ]
        pd.DataFrame(columns=columns).to_excel(filename, index=False)
        for hyp_idx in range(n):
            print(f"\nGenerating hypothesis {hyp_idx+1}/{n}\n")

            # ALWAYS select new chunks for EVERY hypothesis to ensure diversity
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                print(f"üîÑ Selecting new chunks for hypothesis {hyp_idx+1}...")
                randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                dynamic_chunks = self.select_dynamic_chunks_for_generation("package query", randomization_strategy=randomization_strategy)
                if dynamic_chunks:
                    context_texts = dynamic_chunks
                    print(f"üìö Using {len(context_texts)} dynamically selected chunks for generation")
                else:
                    print(f"‚ö†Ô∏è  Failed to select dynamic chunks, using original chunks")
                    context_texts = [chunk for chunk in package_chunks]
            else:
                print(f"üìö Using original chunks for generation (no dynamic selection available)")
                context_texts = [chunk for chunk in package_chunks]

            self.hypothesis_timer.start()
            time_left = self.hypothesis_timer.get_remaining_time()
            
            # Add timeout protection for hypothesis generation
            try:
                hypothesis_list = self.hypothesis_generator.generate(context_texts, n=1)
                if not hypothesis_list:
                    print("‚ùå Failed to generate hypothesis. Skipping...")
                    continue
            except Exception as e:
                logger.error(f"‚ùå Error generating hypothesis: {e}")
                print(f"‚ùå Error generating hypothesis: {e}")
                continue
                
            hypothesis = hypothesis_list[0]
            
            # Validate hypothesis format before proceeding
            from src.ai.hypothesis_tools import validate_hypothesis_format
            is_valid_format, format_reason = validate_hypothesis_format(hypothesis)
            if not is_valid_format:
                print(f"‚ùå Hypothesis rejected due to format: {format_reason}")
                print(f"‚è≠Ô∏è  Skipping critique for format-rejected hypothesis")
                # Record format rejection without critique
                format_rejected_record = {
                    'HypothesisNumber': hyp_idx+1,
                    'Iteration': 0,
                    'Status': 'FORMAT_REJECTED',
                    'Hypothesis': hypothesis,
                    'Novelty': None,
                    'Accuracy': None,
                    'Critique': f'Hypothesis rejected: {format_reason}',
                    'TimeLeft': int(time_left),
                    '': '',
                    'Verifier Novelty Score (0-100)': '',
                    'Verifier Accuracy Score (0-100)': '',
                    'Verifier Verdict (accept, refuse)': ''
                }
                self.hypothesis_records.append(format_rejected_record)
                # Append to Excel
                try:
                    df = pd.DataFrame([format_rejected_record])
                    with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
                except Exception as e:
                    print(f"‚ùå Error appending to Excel: {e}")
                continue
            
            print(f"üìù Generated: {hypothesis}")
            
            # Check timer before critique
            if self.hypothesis_timer.check_expired():
                print(f"‚è∞ Time limit reached before critique")
                timeout_record = {
                    'HypothesisNumber': hyp_idx+1,
                    'Iteration': 0,
                    'Status': 'TIMEOUT',
                    'Hypothesis': hypothesis,
                    'Novelty': None,
                    'Accuracy': None,
                    'Critique': 'Process timed out before critique',
                    'TimeLeft': 0,
                    '': '',
                    'Verifier Novelty Score (0-100)': '',
                    'Verifier Accuracy Score (0-100)': '',
                    'Verifier Verdict (accept, refuse)': ''
                }
                self.hypothesis_records.append(timeout_record)
                # Append to Excel
                try:
                    df = pd.DataFrame([timeout_record])
                    with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
                except Exception as e:
                    print(f"‚ùå Error appending to Excel: {e}")
                continue

            # Critique hypothesis
            print(f"üîç Critiquing hypothesis...")
            try:
                critique_result = self.hypothesis_critic.critique(hypothesis, context_texts, prompt="package query", lab_goals=get_lab_goals(), meta_hypothesis="package query")
            except Exception as e:
                print(f"‚ùå Error during critique: {e}")
                error_record = {
                    'HypothesisNumber': hyp_idx+1,
                    'Iteration': 0,
                    'Status': 'CRITIQUE_ERROR',
                    'Hypothesis': hypothesis,
                    'Novelty': None,
                    'Accuracy': None,
                    'Critique': f'Critique error: {str(e)}',
                    'TimeLeft': int(self.hypothesis_timer.get_remaining_time()),
                    '': '',
                    'Verifier Novelty Score (0-100)': '',
                    'Verifier Accuracy Score (0-100)': '',
                    'Verifier Verdict (accept, refuse)': ''
                }
                self.hypothesis_records.append(error_record)
                # Append to Excel
                try:
                    df = pd.DataFrame([error_record])
                    with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
                except Exception as e:
                    print(f"‚ùå Error appending to Excel: {e}")
                continue
            
            # Check timer after critique
            if self.hypothesis_timer.check_expired():
                print(f"‚è∞ Time limit reached after critique")
                timeout_record = {
                    'HypothesisNumber': hyp_idx+1,
                    'Iteration': 0,
                    'Status': 'TIMEOUT',
                    'Hypothesis': hypothesis,
                    'Novelty': critique_result.get('novelty', None),
                    'Accuracy': critique_result.get('accuracy', None),
                    'Critique': critique_result.get('critique', '') + ' [Process timed out]',
                    'TimeLeft': 0,
                    '': '',
                    'Verifier Novelty Score (0-100)': '',
                    'Verifier Accuracy Score (0-100)': '',
                    'Verifier Verdict (accept, refuse)': ''
                }
                self.hypothesis_records.append(timeout_record)
                # Append to Excel
                try:
                    df = pd.DataFrame([timeout_record])
                    with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
                except Exception as e:
                    print(f"‚ùå Error appending to Excel: {e}")
                continue

            # Extract scores and determine verdict
            novelty = critique_result.get('novelty', 0)
            accuracy = critique_result.get('accuracy', 0)
            relevancy = critique_result.get('relevancy', 0)
            
            # Handle None values by converting to 0
            if novelty is None:
                novelty = 0
            if accuracy is None:
                accuracy = 0
            if relevancy is None:
                relevancy = 0
                
            verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold, novelty_threshold, relevancy_threshold)
            remaining_time = self.hypothesis_timer.get_remaining_time()

            print(f"üìä Scores: Accuracy={accuracy}, Novelty={novelty}, Relevancy={relevancy}")
            print(f"‚öñÔ∏è  Automated Verdict: {verdict}")
            print(f"‚è±Ô∏è  Time remaining: {remaining_time:.1f}s")

            # Record hypothesis
            record = {
                'HypothesisNumber': hyp_idx+1,
                'Iteration': 0,
                'Status': verdict,
                'Hypothesis': hypothesis,
                'Novelty': novelty,
                'Accuracy': accuracy,
                'Critique': critique_result.get('critique', ''),
                'TimeLeft': int(remaining_time),
                '': '',
                'Verifier Novelty Score (0-100)': '',
                'Verifier Accuracy Score (0-100)': '',
                'Verifier Verdict (accept, refuse)': ''
            }
            self.hypothesis_records.append(record)

            # Append to Excel
            try:
                df = pd.DataFrame([record])
                with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
            except Exception as e:
                print(f"‚ùå Error appending to Excel: {e}")

            # Check if accepted
            if verdict == 'ACCEPTED':
                print(f"‚úÖ ACCEPTED! ({hyp_idx+1}/{n})")
            else:
                print(f"‚ùå REJECTED (below thresholds)")

        # Final summary
        end_time = time.time()
        total_time = end_time - start_time
        print(f"\n" + "=" * 80)
        print(f"üèÅ HYPOTHESIS GENERATION COMPLETE")
        print(f"=" * 80)
        print(f"Total hypotheses generated: {n}")
        print(f"Total records: {len(self.hypothesis_records)}")
        print(f"Total time: {total_time:.1f}s")
        print(f"Average time per hypothesis: {total_time/n:.1f}s")

        # Log performance
        self.log_performance("hypothesis_generation_complete", total_time, {
            "n_hypotheses": n,
            "total_records": len(self.hypothesis_records),
            "avg_time_per_hypothesis": total_time/n
        })

        print(f"‚úÖ Results saved to: {filename}")
        print(f"üìÅ File location: {os.path.dirname(os.path.abspath(filename))}")
        
        # Create comprehensive export with popup notification
        print(f"\nüì¶ Creating comprehensive export...")
        comprehensive_export_path = self.create_comprehensive_export()
        if comprehensive_export_path:
            print(f"üìã Comprehensive results exported to: {comprehensive_export_path}")
        else:
            print(f"‚ö†Ô∏è  Comprehensive export failed, but Excel export is complete")

        return self.hypothesis_records

    def is_lab_authored_paper(self, metadata):
        """Check if a paper is authored by the lab PI or lab members."""
        from src.ai.hypothesis_tools import get_lab_config

        config = get_lab_config()
        lab_name = config.get("lab_name", "Dr. Xiaojing Ma")

        # Get author information from metadata
        authors = metadata.get('author', metadata.get('authors', '')).lower()
        if not authors or authors == 'unknown authors':
            # If no authors found in metadata, return False
            return False

        # Clean author string - remove "et al" and other common patterns
        authors_clean = re.sub(r'\bet\s+al\.?\b', '', authors)  # Remove "et al" or "et al."
        authors_clean = re.sub(r'\band\s+others\b', '', authors_clean)  # Remove "and others"
        authors_clean = re.sub(r'\s*&\s*others\b', '', authors_clean)  # Remove "& others" (with optional space)
        authors_clean = re.sub(r'\s+', ' ', authors_clean).strip()  # Normalize whitespace
        authors_clean = re.sub(r'[,\s\.]+$', '', authors_clean)  # Remove trailing commas/spaces/periods

        # Check for lab PI name variations - more comprehensive matching
        pi_name_variations = [
            'xiaojing ma',
            'xiaojing',
            'ma xiaojing',
            'ma, xiaojing',
            'ma x',
            'x ma'
        ]

        # Check if any PI name variation is in the authors
        for pi_variation in pi_name_variations:
            if pi_variation in authors_clean:
                return True

        # Check for lab member names (if configured)
        lab_members = config.get('lab_members', [])
        if lab_members:
            for member in lab_members:
                member_lower = member.lower()
                if member_lower in authors_clean:
                    return True

        return False

    def retrieve_relevant_chunks_with_lab_papers(self, query, top_k=1500, lab_paper_ratio=None, preferred_authors=None):
        """
        Retrieve relevant chunks using the sophisticated lab paper search algorithm.
        This method now implements author prioritization, impact factor weighting, 
        and preprint status detection for enhanced lab paper identification.
        """
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available. Cannot perform sophisticated lab paper search.")
            return self.retrieve_relevant_chunks(query, top_k, lab_paper_ratio)

        print(f"üîç Sophisticated lab paper search for: {query}")
        print(f"üìä Target chunks: {top_k}, Lab paper ratio: {lab_paper_ratio}")

        # Get all results from ChromaDB
        try:
            results = self.chroma_manager.search(
                query=query,
                n_results=top_k * 3,  # Get more results to allow for filtering
                filter_dict=None
            )
        except Exception as e:
            print(f"‚ùå Error searching ChromaDB: {e}")
            return self.retrieve_relevant_chunks(query, top_k, lab_paper_ratio)

        if not results or not results.get('documents'):
            print("‚ùå No results found in ChromaDB")
            return []

        # Process results and apply sophisticated filtering
        processed_results = []
        lab_papers = []
        non_lab_papers = []

        for i, (doc, metadata, distance) in enumerate(zip(
            results['documents'], 
            results['metadatas'], 
            results['distances']
        )):
            # Add preprint metadata
            metadata = self.add_preprint_metadata(metadata)
            
            # Detect if this is a lab-authored paper
            is_lab_paper = self.is_lab_authored_paper(metadata)
            
            # Check if it's a preferred author paper
            is_preferred_author = self._is_preferred_author_paper(metadata, preferred_authors) if preferred_authors else False
            
            result_item = {
                'document': doc,
                'metadata': metadata,
                'distance': distance,
                'is_lab_paper': is_lab_paper,
                'is_preferred_author': is_preferred_author,
                'impact_factor': self._get_journal_impact_factor(metadata.get('journal', '')),
                'is_preprint': metadata.get('is_preprint', False)
            }
            
            if is_lab_paper:
                lab_papers.append(result_item)
            else:
                non_lab_papers.append(result_item)
            
            processed_results.append(result_item)

        print(f"üìä Found {len(lab_papers)} lab papers and {len(non_lab_papers)} non-lab papers")

        # Apply sophisticated selection strategy
        selected_results = self._apply_sophisticated_selection_strategy(
            lab_papers, non_lab_papers, top_k, lab_paper_ratio, preferred_authors
        )

        print(f"‚úÖ Selected {len(selected_results)} chunks using sophisticated lab paper algorithm")
        
        return selected_results

    def configure_lab_paper_ratio(self, ratio=None):
        """Configure the ratio of lab-authored papers to include in searches."""
        if ratio is None:
            print("Current lab paper ratio configuration:")
            print(f"  Lab paper ratio: {getattr(self, 'lab_paper_ratio', 'Not set (default: 0.2)')}")
            print(f"  Preferred authors: {getattr(self, 'preferred_authors', 'Not set')}")
            print(f"  Randomization strategy: {getattr(self, 'randomization_strategy', 'enhanced')}")
            return
        
        if not isinstance(ratio, (int, float)) or ratio < 0 or ratio > 1:
            print("‚ùå Lab paper ratio must be a number between 0 and 1")
            return
        
        self.lab_paper_ratio = ratio
        print(f"‚úÖ Lab paper ratio set to: {ratio}")
        
        # Provide guidance on ratio selection
        if ratio == 0:
            print("üìù Note: Lab papers will be excluded from searches")
        elif ratio <= 0.1:
            print("üìù Note: Very low lab paper inclusion - mostly external papers")
        elif ratio <= 0.3:
            print("üìù Note: Moderate lab paper inclusion - balanced approach")
        elif ratio <= 0.5:
            print("üìù Note: High lab paper inclusion - lab-focused approach")
        else:
            print("üìù Note: Very high lab paper inclusion - primarily lab papers")

    def generate_hypotheses_with_per_hypothesis_timer(self, n=5, max_rounds=10, filename=None):
        """
        Generate n hypotheses with optimized timer (1-minute timeout) and performance monitoring.
        After each iteration, print the result, time left, scores, and hypothesis text.
        Save every iteration in self.hypothesis_records and append to Excel after each iteration.
        Adds empty verifier columns separated by an empty column.
        """
        import pandas as pd
        from openpyxl import load_workbook
        from datetime import datetime
        if not self.current_package["chunks"]:
            print("‚ùå Package is empty. Add some chunks first.")
            return
        if not self.hypothesis_generator or not self.hypothesis_critic:
            print("‚ùå Hypothesis tools not initialized. Check if Gemini API key is available.")
            return
        novelty_threshold = 4
        accuracy_threshold = 4
        relevancy_threshold = 3
        print(f"\nüß† Generating {n} hypotheses (1-minute timer per hypothesis with performance monitoring)...")
        print(f"üì¶ Using {len(self.current_package['chunks'])} package chunks for critique")
        print(f"[INFO] Acceptance criteria: Novelty >= {novelty_threshold}, Accuracy >= {accuracy_threshold}, Relevancy >= {relevancy_threshold}")
        
        # Start performance monitoring
        start_time = time.time()
        self.log_performance("hypothesis_generation_start", 0, {"n_hypotheses": n, "max_rounds": max_rounds})
        package_chunks = self.current_package["chunks"]
        self.hypothesis_records = []  # Clear previous records
        # Prepare Excel file
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Ensure hypothesis_export directory exists
            export_dir = "hypothesis_export"
            os.makedirs(export_dir, exist_ok=True)
            filename = os.path.join(export_dir, f"hypothesis_export_{timestamp}.xlsx")
        # Write header row at the start, with empty separator and verifier columns
        columns = [
            'HypothesisNumber', 'Iteration', 'Status', 'Hypothesis', 'Novelty', 'Accuracy', 'Critique', 'TimeLeft',
            '',  # Empty separator column
            'Verifier Novelty Score (0-100)', 'Verifier Accuracy Score (0-100)', 'Verifier Verdict (accept, refuse)'
        ]
        pd.DataFrame(columns=columns).to_excel(filename, index=False)
        for hyp_idx in range(n):
            print(f"\nGenerating hypothesis {hyp_idx+1}/{n}\n")

            # ALWAYS select new chunks for EVERY hypothesis to ensure diversity
            if self.use_chromadb and self.chroma_manager and self.initial_add_quantity:
                print(f"üîÑ Selecting new chunks for hypothesis {hyp_idx+1}...")
                randomization_strategy = getattr(self, 'randomization_strategy', 'enhanced')
                dynamic_chunks = self.select_dynamic_chunks_for_generation("package query", randomization_strategy=randomization_strategy)
                if dynamic_chunks:
                    context_texts = dynamic_chunks
                    print(f"üìö Using {len(context_texts)} dynamically selected chunks for generation")
                else:
                    print(f"‚ö†Ô∏è  Failed to select dynamic chunks, using original chunks")
                    context_texts = [chunk for chunk in package_chunks]
            else:
                print(f"üìö Using original chunks for generation (no dynamic selection available)")
                context_texts = [chunk for chunk in package_chunks]

            self.hypothesis_timer.start()
            time_left = self.hypothesis_timer.get_remaining_time()
            
            # Add timeout protection for hypothesis generation
            try:
                hypothesis_list = self.hypothesis_generator.generate(context_texts, n=1)
                if not hypothesis_list:
                    print("‚ùå Failed to generate hypothesis. Skipping...")
                    continue
            except Exception as e:
                logger.error(f"‚ùå Error generating hypothesis: {e}")
                print(f"‚ùå Error generating hypothesis: {e}")
                continue
                
            hypothesis = hypothesis_list[0]
            
            # Validate hypothesis format before proceeding
            from src.ai.hypothesis_tools import validate_hypothesis_format
            is_valid_format, format_reason = validate_hypothesis_format(hypothesis)
            if not is_valid_format:
                print(f"‚ùå Hypothesis rejected due to format: {format_reason}")
                print(f"‚è≠Ô∏è  Skipping critique for format-rejected hypothesis")
                # Record format rejection without critique
                format_rejected_record = {
                    'HypothesisNumber': hyp_idx+1,
                    'Iteration': 0,
                    'Status': 'FORMAT_REJECTED',
                    'Hypothesis': hypothesis,
                    'Novelty': None,
                    'Accuracy': None,
                    'Critique': f'Hypothesis rejected: {format_reason}',
                    'TimeLeft': int(time_left),
                    '': '',
                    'Verifier Novelty Score (0-100)': '',
                    'Verifier Accuracy Score (0-100)': '',
                    'Verifier Verdict (accept, refuse)': ''
                }
                self.hypothesis_records.append(format_rejected_record)
                # Append to Excel
                df_row = pd.DataFrame([format_rejected_record])
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_row.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=len(self.hypothesis_records))
                continue
            
            accepted = False
            iteration = 0
            while not accepted and not self.hypothesis_timer.check_expired() and iteration < max_rounds:
                iteration += 1
                critique_result = self.hypothesis_critic.critique(hypothesis, context_texts, prompt=query, lab_goals=get_lab_goals(), meta_hypothesis=query)
                novelty = critique_result.get('novelty', 0)
                accuracy = critique_result.get('accuracy', 0)
                relevancy = critique_result.get('relevancy', 0)
                
                # Handle None values by converting to 0
                if novelty is None:
                    novelty = 0
                if accuracy is None:
                    accuracy = 0
                if relevancy is None:
                    relevancy = 0
                    
                time_left = self.hypothesis_timer.get_remaining_time()
                verdict = self.automated_verdict(accuracy, novelty, relevancy, accuracy_threshold=4, novelty_threshold=4, relevancy_threshold=3)
                print(f"Iteration {iteration}: {verdict} | Novelty: {novelty}/4 | Accuracy: {accuracy}/4 | Relevancy: {relevancy}/3 | Time left: {int(time_left)}s")
                print(f"Hypothesis: {hypothesis}")
                # Save this iteration to records
                record = {
                    'HypothesisNumber': hyp_idx+1,
                    'Iteration': iteration,
                    'Status': verdict,
                    'Hypothesis': hypothesis,
                    'Novelty': novelty,
                    'Accuracy': accuracy,
                    'Critique': critique_result.get('critique', ''),
                    'TimeLeft': int(time_left),
                    '': '',  # Empty separator column
                    'Verifier Novelty Score (0-100)': '',
                    'Verifier Accuracy Score (0-100)': '',
                    'Verifier Verdict (accept, refuse)': ''
                }
                self.hypothesis_records.append(record)
                # Append to Excel after each iteration
                df_row = pd.DataFrame([record])
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    writer.book = load_workbook(filename)
                    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
                    startrow = writer.book['Hypotheses'].max_row
                    df_row.to_excel(writer, sheet_name='Hypotheses', index=False, header=False, startrow=startrow)
                if verdict == "ACCEPTED":
                    print(f"‚úÖ Hypothesis accepted! (verdict: {verdict})")
                    accepted = True
                    break
                else:
                    # Print specific rejection reason
                    if accuracy < 4:
                        print(f"‚ùå Hypothesis rejected: Low accuracy score ({accuracy}/4)")
                    elif novelty < 4:
                        print(f"‚ùå Hypothesis rejected: Low novelty score ({novelty}/4)")
                    elif relevancy < 3:
                        print(f"‚ùå Hypothesis rejected: Low relevancy score ({relevancy}/3)")
                    else:
                        print(f"‚ùå Hypothesis rejected: Unknown reason (verdict: {verdict})")
                if self.hypothesis_timer.check_expired():
                    print(f"‚è∞ Time limit reached for hypothesis {hyp_idx+1}. Moving to next hypothesis.")
                    break
                # Regenerate hypothesis for next iteration
                new_hypothesis = self.hypothesis_generator.generate(context_texts, n=1)
                if new_hypothesis:
                    hypothesis = new_hypothesis[0]
        print(f"\n‚úÖ All iterations and results have been saved to: {filename}")
        print(f"üìÅ File saved in: {os.path.dirname(os.path.abspath(filename))}")
        return self.hypothesis_records

    def is_lab_authored_paper(self, metadata):
        """Check if a paper is authored by the lab PI or lab members."""
        from src.ai.hypothesis_tools import get_lab_config

        config = get_lab_config()
        lab_name = config.get("lab_name", "Dr. Xiaojing Ma")

        # Get author information from metadata
        authors = metadata.get('author', metadata.get('authors', '')).lower()
        if not authors or authors == 'unknown authors':
            # If no authors found in metadata, return False
            return False

        # Clean author string - remove "et al" and other common patterns
        authors_clean = re.sub(r'\bet\s+al\.?\b', '', authors)  # Remove "et al" or "et al."
        authors_clean = re.sub(r'\band\s+others\b', '', authors_clean)  # Remove "and others"
        authors_clean = re.sub(r'\s*&\s*others\b', '', authors_clean)  # Remove "& others" (with optional space)
        authors_clean = re.sub(r'\s+', ' ', authors_clean).strip()  # Normalize whitespace
        authors_clean = re.sub(r'[,\s\.]+$', '', authors_clean)  # Remove trailing commas/spaces/periods

        # Check for lab PI name variations - more comprehensive matching
        lab_pi_variations = [
            lab_name.lower(),
            lab_name.replace("Dr. ", "").lower(),
            lab_name.replace("Dr. ", "").replace(" ", "").lower(),
            "xiaojing ma",
            "xiaojing",
            "ma x",
            "ma, x",
            "ma, xiaojing",
            "xiaojing ma,",
            "x. ma",
            "x ma",
            "ma, x.",
            "ma xiaojing",
            "xiaojing m",
            "x. m.",
            "x m"
        ]

        # Check if any lab PI variation is in authors (using word boundaries)
        for variation in lab_pi_variations:
            # Use word boundaries to avoid partial matches
            if re.search(r'\b' + re.escape(variation) + r'\b', authors_clean):
                return True

        # Also check for common lab member patterns (if any are known)
        # This could be expanded with actual lab member names
        lab_member_patterns = [
            r'\bma\s+x\b',  # X Ma pattern
            r'\bxiaojing\s+ma\b',  # Xiaojing Ma pattern
            r'\bma,\s*x\b',  # Ma, X pattern
        ]
        
        for pattern in lab_member_patterns:
            if re.search(pattern, authors_clean, re.IGNORECASE):
                return True

        return False

    def retrieve_relevant_chunks_with_lab_papers(self, query, top_k=1500, lab_paper_ratio=None, preferred_authors=None):
        """
        Retrieve relevant chunks using the sophisticated lab paper search algorithm.
        This method now implements author prioritization, impact factor weighting, 
        citation analysis, and temporal balance (50/50 old/new mix).

        Args:
            query: Search query
            top_k: Total number of chunks to retrieve
            lab_paper_ratio: Optional fixed ratio (0.0 to 1.0). If None, automatically determined.
            preferred_authors: List of preferred author names to prioritize
        """
        # Use the new sophisticated search algorithm
        return self.sophisticated_lab_paper_search(query, top_k, lab_paper_ratio, preferred_authors)

    def configure_lab_paper_ratio(self, ratio=None):
        """Configure the ratio of lab-authored papers to include in searches."""
        if ratio is None:
            current_ratio = getattr(self, 'lab_paper_ratio', None)
            if current_ratio is None:
                print(f"üî¨ Lab paper ratio: AUTO (automatically determined based on available lab papers)")
                print(f"   ‚Ä¢ If lab papers found: 10-50% (based on availability)")
                print(f"   ‚Ä¢ If no lab papers: 0%")
            else:
                print(f"üî¨ Current lab paper ratio: {current_ratio:.1%} (fixed)")

            try:
                choice = input("Enter 'auto' for automatic determination, or a number (0.0 to 1.0, e.g., 0.3 for 30%): ").strip().lower()
                if choice == 'auto':
                    self.lab_paper_ratio = None
                    print(f"‚úÖ Lab paper ratio set to AUTO (automatically determined)")
                else:
                    new_ratio = float(choice)
                    if 0.0 <= new_ratio <= 1.0:
                        self.lab_paper_ratio = new_ratio
                        print(f"‚úÖ Lab paper ratio updated to {new_ratio:.1%} (fixed)")
                    else:
                        print("‚ùå Invalid ratio. Must be between 0.0 and 1.0.")
            except ValueError:
                print("‚ùå Invalid input. Please enter 'auto' or a number between 0.0 and 1.0.")
        else:
            if ratio == 'auto':
                self.lab_paper_ratio = None
                print(f"‚úÖ Lab paper ratio set to AUTO (automatically determined)")
            elif 0.0 <= ratio <= 1.0:
                self.lab_paper_ratio = ratio
                print(f"‚úÖ Lab paper ratio set to {ratio:.1%} (fixed)")
            else:
                print("‚ùå Invalid ratio. Must be 'auto' or between 0.0 and 1.0.")

    def test_lab_paper_detection(self, query="UBR5 cancer"):
        """Test function to debug lab paper detection."""
        print(f"üî¨ Testing lab paper detection for query: '{query}'")
        
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available for testing.")
            return
        
        # Get query embedding
        query_embedding = self.get_google_embedding(query)
        if not query_embedding:
            print("‚ùå Failed to get query embedding.")
            return
        
        print(f"üîç Searching ChromaDB for lab papers...")
        
        # Search broadly to find potential lab papers
        search_results = self.chroma_manager.search_similar(
            query_embedding=query_embedding,
            n_results=1000  # Search broadly
        )
        
        lab_papers_found = []
        other_papers = []
        
        for result in search_results:
            metadata = result.get('metadata', {})
            authors = metadata.get('author', metadata.get('authors', ''))
            title = metadata.get('title', '')
            
            if self.is_lab_authored_paper(metadata):
                lab_papers_found.append({
                    'title': title,
                    'authors': authors,
                    'doi': metadata.get('doi', 'No DOI'),
                    'source': metadata.get('source_name', 'Unknown')
                })
            else:
                other_papers.append({
                    'title': title,
                    'authors': authors,
                    'doi': metadata.get('doi', 'No DOI'),
                    'source': metadata.get('source_name', 'Unknown')
                })
        
        print(f"\nüìä Lab Paper Detection Results:")
        print(f"   Total papers searched: {len(search_results)}")
        print(f"   Lab papers found: {len(lab_papers_found)}")
        print(f"   Other papers: {len(other_papers)}")
        
        if lab_papers_found:
            print(f"\nüî¨ Lab Papers Found:")
            for i, paper in enumerate(lab_papers_found[:10], 1):  # Show first 10
                print(f"   {i}. {paper['title']}")
                print(f"      Authors: {paper['authors']}")
                print(f"      DOI: {paper['doi']}")
                print(f"      Source: {paper['source']}")
                print()
        else:
            print(f"\n‚ùå No lab papers found. This is expected because:")
            print(f"   ‚Ä¢ Current database sources: PubMed, bioRxiv, medRxiv")
            print(f"   ‚Ä¢ Lab papers may not be in these general databases")
            print(f"   ‚Ä¢ Lab paper detection logic is working correctly")
            print(f"   ‚Ä¢ Author name variations are properly configured")
            
            # Show some example papers to help debug
            if other_papers:
                print(f"\nüìö Sample papers in database:")
                for i, paper in enumerate(other_papers[:5], 1):
                    print(f"   {i}. {paper['title']}")
                    print(f"      Authors: {paper['authors']}")
                    print(f"      Source: {paper['source']}")
                    print()
                
                print(f"üí° To include lab papers, consider:")
                print(f"   ‚Ä¢ Adding lab papers manually to the database")
                print(f"   ‚Ä¢ Checking if lab papers exist in PubMed/bioRxiv/medRxiv")
                print(f"   ‚Ä¢ Using different data sources that include lab publications")

    def configure_randomization_strategy(self, strategy=None):
        """
        Configure the chunk randomization strategy for hypothesis generation.
        
        Args:
            strategy: Randomization strategy ('enhanced', 'shuffle', 'diversity', 'time_based', 'none')
        """
        available_strategies = {
            'enhanced': 'Enhanced randomization combining relevance and diversity',
            'shuffle': 'Simple random shuffle of all chunks',
            'diversity': 'Prioritize diversity by publication date and citations',
            'time_based': 'Time-based randomization (changes every minute)',
            'none': 'No additional randomization (use default ordering)'
        }
        
        if strategy is None:
            print("\nüé≤ Chunk Randomization Strategy Configuration")
            print("=" * 50)
            print("Current strategy:", getattr(self, 'randomization_strategy', 'enhanced'))
            print("\nAvailable strategies:")
            for key, description in available_strategies.items():
                print(f"  ‚Ä¢ {key}: {description}")
            
            print(f"\nüí° The system already ensures each hypothesis uses different papers.")
            print(f"üí° This randomization adds additional variety to chunk selection within those papers.")
            
            strategy = input("\nEnter strategy name (or press Enter to keep current): ").strip().lower()
            if not strategy:
                return
        
        if strategy not in available_strategies:
            print(f"‚ùå Invalid strategy: {strategy}")
            print(f"Available strategies: {', '.join(available_strategies.keys())}")
            return
        
        self.randomization_strategy = strategy
        print(f"‚úÖ Randomization strategy set to: {strategy}")
        print(f"   üìù {available_strategies[strategy]}")

    def configure_preferred_authors(self, authors=None):
        """Configure preferred authors for the sophisticated search algorithm."""
        if authors is None:
            current_authors = getattr(self, 'preferred_authors', [])
            if current_authors:
                print(f"üë• Current preferred authors: {', '.join(current_authors)}")
            else:
                print(f"üë• No preferred authors configured")

            try:
                print("\nEnter preferred author names (comma-separated, or 'clear' to remove all):")
                choice = input("Authors: ").strip()
                if choice.lower() == 'clear':
                    self.preferred_authors = []
                    print(f"‚úÖ Cleared preferred authors")
                elif choice:
                    author_list = [author.strip() for author in choice.split(',') if author.strip()]
                    self.preferred_authors = author_list
                    print(f"‚úÖ Set preferred authors: {', '.join(author_list)}")
                else:
                    print(f"‚úÖ Keeping current preferred authors")
            except KeyboardInterrupt:
                print("\n‚úÖ Configuration cancelled.")
        else:
            if isinstance(authors, list):
                self.preferred_authors = authors
                print(f"‚úÖ Set preferred authors: {', '.join(authors)}")
            else:
                print(f"‚ùå Invalid authors format. Expected list of strings.")

    def sophisticated_lab_paper_search(self, query, top_k=1500, lab_paper_ratio=None, preferred_authors=None):
        """
        Sophisticated lab paper search with author prioritization, impact factor weighting, 
        citation analysis, temporal balance (50/50 old/new mix), and preprint deprioritization.
        
        Args:
            query: Search query
            top_k: Total number of chunks to retrieve
            lab_paper_ratio: Optional fixed ratio (0.0 to 1.0). If None, automatically determined.
            preferred_authors: List of preferred author names to prioritize
        """
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available.")
            return []

        # Check if ChromaDB has data
        if not self.is_chromadb_ready():
            print("‚ùå ChromaDB is empty. Please load data first using the master processor (option 4).")
            return []

        # Get query embedding
        query_embedding = self.get_google_embedding(query)
        if not query_embedding:
            print("‚ùå Failed to get query embedding.")
            return []

        print(f"üî¨ Starting sophisticated lab paper search...")
        print(f"üìä Target: {top_k} total chunks")
        
        # Use configured preferred authors if none provided
        if preferred_authors is None:
            preferred_authors = getattr(self, 'preferred_authors', [])
        
        if preferred_authors:
            print(f"üë• Using preferred authors: {', '.join(preferred_authors)}")
        
        # Step 1: Broad search to get all relevant papers
        print(f"üîç Performing broad search for relevant papers...")
        # Limit search to reasonable number to avoid hanging
        max_search_results = min(top_k * 2, 2000)  # Further reduced to top_k * 2, max 2000 for better performance
        broad_search_results = self.chroma_manager.search_similar(
            query_embedding=query_embedding,
            n_results=max_search_results
        )

        if not broad_search_results:
            print("‚ùå No papers found in ChromaDB.")
            return []

        # Step 2: Categorize papers and extract metadata
        print(f"üìã Categorizing papers and extracting metadata...")
        print(f"üìä Processing {len(broad_search_results)} papers...")
        lab_papers = []
        preferred_author_papers = []
        other_papers = []
        
        # Track citation statistics for analysis
        all_citations = []
        journal_impact_factors = {}
        
        # Add progress tracking
        total_papers = len(broad_search_results)
        processed_count = 0
        start_time = time.time()
        
        for i, result in enumerate(broad_search_results):
            processed_count += 1
            # Show progress every 100 papers
            if processed_count % 100 == 0 or processed_count == total_papers:
                elapsed_time = time.time() - start_time
                print(f"üìä Progress: {processed_count}/{total_papers} papers processed ({processed_count/total_papers*100:.1f}%) - {elapsed_time:.1f}s elapsed")
            
            # Safety check: if processing takes too long, break early
            if time.time() - start_time > 60:  # 60 second timeout
                print(f"‚ö†Ô∏è Processing timeout reached after 60 seconds. Processed {processed_count}/{total_papers} papers.")
                break
            metadata = result.get('metadata', {})
            
            # Extract citation count for analysis
            citation_count = metadata.get('citation_count', 'not found')
            if citation_count != 'not found':
                try:
                    citations = int(citation_count)
                    all_citations.append(citations)
                except (ValueError, TypeError):
                    pass
            
            # Track journal for impact factor analysis
            journal = metadata.get('journal', 'Unknown journal')
            if journal not in journal_impact_factors:
                journal_impact_factors[journal] = []
            if citation_count != 'not found':
                try:
                    journal_impact_factors[journal].append(int(citation_count))
                except (ValueError, TypeError):
                    pass
            
            # Categorize papers
            if self.is_lab_authored_paper(metadata):
                lab_papers.append(result)
            elif preferred_authors and self._is_preferred_author_paper(metadata, preferred_authors):
                preferred_author_papers.append(result)
            else:
                other_papers.append(result)

        print(f"üìä Found {len(lab_papers)} lab papers, {len(preferred_author_papers)} preferred author papers, {len(other_papers)} other papers")

        # Step 3: Calculate statistics for prioritization
        print(f"üìà Calculating citation and journal statistics...")
        
        # Calculate average citation count
        avg_citations = np.mean(all_citations) if all_citations else 0
        median_citations = np.median(all_citations) if all_citations else 0
        
        # Calculate journal impact factors (average citations per journal)
        journal_impact_scores = {}
        for journal, citations in journal_impact_factors.items():
            if citations:
                journal_impact_scores[journal] = np.mean(citations)
        
        print(f"üìä Average citations: {avg_citations:.1f}, Median: {median_citations:.1f}")
        print(f"üìä Journal impact scores calculated for {len(journal_impact_scores)} journals")

        # Step 4: Determine lab paper ratio
        if lab_paper_ratio is None:
            if len(lab_papers) > 0:
                available_lab_ratio = min(len(lab_papers) / top_k, 0.5)
                lab_paper_ratio = max(available_lab_ratio, 0.1)
                print(f"üî¨ Auto-determined lab paper ratio: {lab_paper_ratio:.1%}")
            else:
                lab_paper_ratio = 0.0
                print(f"üî¨ No lab-authored papers found")
        else:
            print(f"üî¨ Using configured lab paper ratio: {lab_paper_ratio:.1%}")

        # Step 5: Author paper selection (Step 1 of search execution)
        print(f"üë• Step 1: Selecting author-prioritized papers...")
        lab_chunks_needed = int(top_k * lab_paper_ratio)
        preferred_author_chunks_needed = min(len(preferred_author_papers), int(top_k * 0.2))  # Up to 20% for preferred authors
        remaining_chunks_needed = top_k - lab_chunks_needed - preferred_author_chunks_needed

        selected_papers = []
        used_papers = set()

        # Add lab papers first (highest priority)
        for paper in lab_papers[:lab_chunks_needed]:
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id and paper_id not in used_papers:
                selected_papers.append(paper)
                used_papers.add(paper_id)

        # Add preferred author papers
        for paper in preferred_author_papers[:preferred_author_chunks_needed]:
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id and paper_id not in used_papers:
                selected_papers.append(paper)
                used_papers.add(paper_id)

        print(f"‚úÖ Selected {len(selected_papers)} author-prioritized papers")

        # Step 6: Remaining paper selection with temporal balance and preprint deprioritization (Step 2 of search execution)
        print(f"üìö Step 2: Selecting remaining papers with temporal balance and preprint deprioritization...")
        
        # Calculate median age of ALL search results (not just selected papers)
        all_years = []
        for paper in broad_search_results:
            metadata = paper.get('metadata', {})
            publication_date = metadata.get('publication_date', '')
            if publication_date and len(publication_date) >= 4:
                try:
                    year = int(publication_date[:4])
                    all_years.append(year)
                except (ValueError, TypeError):
                    pass
        
        median_year = np.median(all_years) if all_years else datetime.now().year - 5
        current_year = datetime.now().year
        
        print(f"üìÖ Median year of ALL search results: {median_year:.0f}")
        print(f"üìÖ Temporal dividing line: {median_year:.0f} (papers before = old, after = new)")
        
        # Show temporal distribution statistics
        if all_years:
            min_year = min(all_years)
            max_year = max(all_years)
            year_counts = {}
            for year in all_years:
                year_counts[year] = year_counts.get(year, 0) + 1
            
            print(f"üìä Temporal distribution: {min_year} to {max_year}")
            print(f"üìä Total papers with dates: {len(all_years)}")
            print(f"üìä Papers ‚â§ {median_year:.0f}: {sum(1 for y in all_years if y <= median_year)}")
            print(f"üìä Papers > {median_year:.0f}: {sum(1 for y in all_years if y > median_year)}")

        # Categorize remaining papers by temporal period and preprint status
        old_papers = []
        new_papers = []
        
        for paper in other_papers:
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id in used_papers:
                continue
                
            metadata = paper.get('metadata', {})
            publication_date = metadata.get('publication_date', '')
            
            # Determine temporal category
            temporal_category = None
            if publication_date and len(publication_date) >= 4:
                try:
                    year = int(publication_date[:4])
                    if year <= median_year:
                        temporal_category = 'old'
                    else:
                        temporal_category = 'new'
                except (ValueError, TypeError):
                    temporal_category = 'unknown'
            else:
                temporal_category = 'unknown'
            
            # Determine preprint status
            is_preprint = metadata.get('is_preprint', False)
            preprint_score = metadata.get('preprint_score', 0.0)
            preprint_type = metadata.get('preprint_type', 'unknown')
            
            # Add preprint information to metadata if not present
            if 'is_preprint' not in metadata:
                # Try to detect preprint status from existing fields
                source = metadata.get('source', '').lower()
                journal = metadata.get('journal', '').lower()
                
                if (source in ['biorxiv', 'medrxiv', 'arxiv', 'chemrxiv'] or 
                    any(indicator in journal for indicator in ['arxiv', 'biorxiv', 'medrxiv', 'chemrxiv', 'preprint'])):
                    is_preprint = True
                    preprint_score = 0.8
                    preprint_type = source if source else 'preprint_server'
                
                # Update metadata
                metadata['is_preprint'] = is_preprint
                metadata['preprint_score'] = preprint_score
                metadata['preprint_type'] = preprint_type
            
            # Categorize by temporal period
            if temporal_category == 'old':
                old_papers.append(paper)
            elif temporal_category == 'new':
                new_papers.append(paper)
            else:
                # Unknown date papers go to old category as fallback
                old_papers.append(paper)

        print(f"üìä Categorized remaining papers: {len(old_papers)} old, {len(new_papers)} new")
        
        # Count preprint papers for reporting
        preprint_count = sum(1 for p in old_papers + new_papers 
                           if p.get('metadata', {}).get('is_preprint', False))
        print(f"üìä Found {preprint_count} preprint papers (will be deprioritized)")

        # Step 7: Score and rank remaining papers with preprint deprioritization
        print(f"üéØ Step 3: Scoring and ranking papers with preprint deprioritization...")
        
        def calculate_paper_score(paper, is_old_paper):
            """Calculate comprehensive paper score with preprint deprioritization."""
            metadata = paper.get('metadata', {})
            
            # Base score starts at 100
            score = 100.0
            
            # Citation-based scoring (0-50 points)
            citation_count = metadata.get('citation_count', 'not found')
            if citation_count != 'not found':
                try:
                    citations = int(citation_count)
                    if citations > 0:
                        # Logarithmic scaling for citations
                        citation_score = min(50, 10 * np.log10(citations + 1))
                        score += citation_score
                except (ValueError, TypeError):
                    pass
            
            # Journal impact scoring (0-30 points)
            journal = metadata.get('journal', 'Unknown journal')
            if journal in journal_impact_scores:
                journal_score = min(30, journal_impact_scores[journal] / 10)
                score += journal_score
            
            # Temporal balance scoring (0-20 points)
            if is_old_paper:
                # Old papers get bonus to maintain balance
                score += 20
            else:
                # New papers get smaller bonus
                score += 10
            
            # PREPRINT DEPRIORITIZATION (penalty of 0-40 points)
            is_preprint = metadata.get('is_preprint', False)
            preprint_score = metadata.get('preprint_score', 0.0)
            
            if is_preprint:
                # Calculate preprint penalty based on preprint score
                # Higher preprint score = higher penalty (more clearly a preprint)
                preprint_penalty = min(40, preprint_score * 50)  # 0.8 score = 40 penalty
                score -= preprint_penalty
                
                # Additional penalty for certain preprint types
                preprint_type = metadata.get('preprint_type', '')
                if preprint_type in ['biorxiv', 'medrxiv', 'chemrxiv']:
                    score -= 10  # Extra penalty for bio/med/chem preprints
                elif preprint_type == 'working_paper':
                    score -= 5   # Smaller penalty for working papers
            
            # Ensure score doesn't go below 0
            score = max(0, score)
            
            return score
        
        # Score old papers
        old_papers_scored = []
        for paper in old_papers:
            score = calculate_paper_score(paper, is_old_paper=True)
            old_papers_scored.append((paper, score))
        
        # Score new papers
        new_papers_scored = []
        for paper in new_papers:
            score = calculate_paper_score(paper, is_old_paper=False)
            new_papers_scored.append((paper, score))
        
        # Sort by score (highest first)
        old_papers_scored.sort(key=lambda x: x[1], reverse=True)
        new_papers_scored.sort(key=lambda x: x[1], reverse=True)
        
        # Show top scored papers for each category
        print(f"üìä Top 5 old papers scores: {[f'{p[1]:.1f}' for p in old_papers_scored[:5]]}")
        print(f"üìä Top 5 new papers scores: {[f'{p[1]:.1f}' for p in new_papers_scored[:5]]}")
        
        # Count preprint papers in top scores
        top_old_preprints = sum(1 for p, _ in old_papers_scored[:20] 
                               if p.get('metadata', {}).get('is_preprint', False))
        top_new_preprints = sum(1 for p, _ in new_papers_scored[:20] 
                               if p.get('metadata', {}).get('is_preprint', False))
        print(f"üìä Preprint papers in top 20 scores: {top_old_preprints} old, {top_new_preprints} new")

        # Step 8: Select remaining papers with balanced distribution
        print(f"üéØ Step 4: Selecting remaining papers with balanced distribution...")
        
        # Calculate target chunks for each temporal category
        old_chunks_needed = remaining_chunks_needed // 2
        new_chunks_needed = remaining_chunks_needed - old_chunks_needed
        
        print(f"üìä Target distribution: {old_chunks_needed} old papers, {new_chunks_needed} new papers")
        
        # Select old papers
        old_selected = 0
        for paper, score in old_papers_scored:
            if old_selected >= old_chunks_needed:
                break
                
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id and paper_id not in used_papers:
                selected_papers.append(paper)
                used_papers.add(paper_id)
                old_selected += 1
        
        # Select new papers
        new_selected = 0
        for paper, score in new_papers_scored:
            if new_selected >= new_chunks_needed:
                break
                
            paper_id = self._get_paper_identifier(paper.get('metadata', {}))
            if paper_id and paper_id not in used_papers:
                selected_papers.append(paper)
                used_papers.add(paper_id)
                new_selected += 1
        
        print(f"‚úÖ Selected {old_selected} old papers and {new_selected} new papers")
        
        # If we don't have enough papers, fill with remaining high-scoring papers
        if len(selected_papers) < top_k:
            remaining_needed = top_k - len(selected_papers)
            print(f"‚ö†Ô∏è  Need {remaining_needed} more papers, filling with remaining high-scoring papers...")
            
            # Combine remaining papers and sort by score
            remaining_papers = []
            for paper, score in old_papers_scored:
                if paper not in selected_papers:
                    remaining_papers.append((paper, score))
            for paper, score in new_papers_scored:
                if paper not in selected_papers:
                    remaining_papers.append((paper, score))
            
            # Sort by score and add until we reach target
            remaining_papers.sort(key=lambda x: x[1], reverse=True)
            for paper, score in remaining_papers:
                if len(selected_papers) >= top_k:
                    break
                    
                paper_id = self._get_paper_identifier(paper.get('metadata', {}))
                if paper_id and paper_id not in used_papers:
                    selected_papers.append(paper)
                    used_papers.add(paper_id)

        # Step 9: Final statistics and validation
        print(f"üìä Final selection statistics:")
        print(f"   Total papers selected: {len(selected_papers)}")
        print(f"   Lab papers: {sum(1 for p in selected_papers if self.is_lab_authored_paper(p.get('metadata', {})))}")
        print(f"   Preferred author papers: {sum(1 for p in selected_papers if preferred_authors and self._is_preferred_author_paper(p.get('metadata', {}), preferred_authors))}")
        
        # Count preprint papers in final selection
        final_preprint_count = sum(1 for p in selected_papers 
                                  if p.get('metadata', {}).get('is_preprint', False))
        print(f"   Preprint papers: {final_preprint_count} (deprioritized)")
        
        # Show preprint distribution by type
        preprint_types = {}
        for paper in selected_papers:
            metadata = paper.get('metadata', {})
            if metadata.get('is_preprint', False):
                preprint_type = metadata.get('preprint_type', 'unknown')
                preprint_types[preprint_type] = preprint_types.get(preprint_type, 0) + 1
        
        if preprint_types:
            print(f"   Preprint types: {', '.join([f'{k}: {v}' for k, v in preprint_types.items()])}")
        
        # Validate we have enough papers
        if len(selected_papers) < top_k:
            print(f"‚ö†Ô∏è  Warning: Only found {len(selected_papers)} papers (requested {top_k})")
        elif len(selected_papers) > top_k:
            print(f"üìä Note: Selected {len(selected_papers)} papers (requested {top_k})")
        
        # Return the selected papers
        return selected_papers

    def retrieve_relevant_chunks_with_lab_papers_deprioritized(self, query, top_k=1500, lab_paper_ratio=None, preferred_authors=None):
        """
        Retrieve relevant chunks with deprioritization of previously used papers.
        This is a modified version of sophisticated_lab_paper_search that applies deprioritization.
        """
        if not self.use_chromadb or not self.chroma_manager:
            print("‚ùå ChromaDB not available.")
            return []

        # Check if ChromaDB has data
        if not self.is_chromadb_ready():
            print("‚ùå ChromaDB is empty. Please load data first using the master processor (option 4).")
            return []

        # Get query embedding
        query_embedding = self.get_google_embedding(query)
        if not query_embedding:
            print("‚ùå Failed to get query embedding.")
            return []

        print(f"üî¨ Starting sophisticated lab paper search with deprioritization...")
        print(f"üìä Target: {top_k} total chunks")
        print(f"üéØ Deprioritization factor: {self.deprioritization_factor * 100}% reduction per use")
        
        # Use configured preferred authors if none provided
        if preferred_authors is None:
            preferred_authors = getattr(self, 'preferred_authors', [])
        
        if preferred_authors:
            print(f"üë• Using preferred authors: {', '.join(preferred_authors)}")
        
        # Step 1: Broad search to get all relevant papers
        print(f"üîç Performing broad search for relevant papers...")
        broad_search_results = self.chroma_manager.search_similar(
            query_embedding=query_embedding,
            n_results=top_k * 5  # Search broadly to find all relevant papers
        )

        if not broad_search_results:
            print("‚ùå No papers found in ChromaDB.")
            return []

        # Step 2: Apply deprioritization to all results
        print(f"üìä Applying deprioritization to {len(broad_search_results)} papers...")
        deprioritized_results = []
        
        for result in broad_search_results:
            metadata = result.get('metadata', {})
            paper_id = self._get_paper_identifier(metadata)
            
            # Calculate deprioritization factor
            usage_count = self.paper_usage_count.get(paper_id, 0)
            deprioritization_multiplier = (1 - self.deprioritization_factor) ** usage_count
            
            # Apply deprioritization to similarity score
            original_similarity = result.get('similarity', 1.0)
            adjusted_similarity = original_similarity * deprioritization_multiplier
            
            # Create new result with deprioritization info
            deprioritized_result = result.copy()
            deprioritized_result.update({
                'adjusted_similarity': adjusted_similarity,
                'usage_count': usage_count,
                'deprioritization_multiplier': deprioritization_multiplier,
                'paper_id': paper_id
            })
            
            deprioritized_results.append(deprioritized_result)

        # Step 3: Re-sort by adjusted similarity
        deprioritized_results.sort(key=lambda x: x['adjusted_similarity'], reverse=True)
        
        # Step 4: Categorize papers and extract metadata (same as original method)
        print(f"üìã Categorizing papers and extracting metadata...")
        lab_papers = []
        preferred_author_papers = []
        other_papers = []
        
        # Track citation statistics for analysis
        all_citations = []
        journal_impact_factors = {}
        
        for result in deprioritized_results:
            metadata = result.get('metadata', {})
            
            # Extract citation count for analysis
            citation_count = metadata.get('citation_count', 'not found')
            if citation_count != 'not found':
                try:
                    citations = int(citation_count)
                    all_citations.append(citations)
                except (ValueError, TypeError):
                    pass
            
            # Track journal for impact factor analysis
            journal = metadata.get('journal', 'Unknown journal')
            if journal not in journal_impact_factors:
                journal_impact_factors[journal] = []
            if citation_count != 'not found':
                try:
                    journal_impact_factors[journal].append(int(citation_count))
                except (ValueError, TypeError):
                    pass
            
            # Categorize papers
            if self.is_lab_authored_paper(metadata):
                lab_papers.append(result)
            elif preferred_authors and self._is_preferred_author_paper(metadata, preferred_authors):
                preferred_author_papers.append(result)
            else:
                other_papers.append(result)

        print(f"üìä Found {len(lab_papers)} lab papers, {len(preferred_author_papers)} preferred author papers, {len(other_papers)} other papers")

        # Step 5: Calculate statistics for prioritization (same as original)
        print(f"üìà Calculating citation and journal statistics...")
        
        # Calculate average citation count
        avg_citations = np.mean(all_citations) if all_citations else 0
        median_citations = np.median(all_citations) if all_citations else 0
        
        # Calculate journal impact factors (average citations per journal)
        journal_impact_scores = {}
        for journal, citations in journal_impact_factors.items():
            if citations:
                journal_impact_scores[journal] = np.mean(citations)
        
        print(f"üìä Citation statistics: avg={avg_citations:.1f}, median={median_citations:.1f}")
        print(f"üìä Journal impact scores calculated for {len(journal_impact_scores)} journals")

        # Step 6: Select papers with deprioritization applied (same logic as original)
        selected_papers = []
        
        # Determine lab paper ratio
        if lab_paper_ratio is None:
            # Auto-determine based on available lab papers
            total_available = len(lab_papers) + len(preferred_author_papers) + len(other_papers)
            lab_paper_ratio = min(0.3, len(lab_papers) / total_available) if total_available > 0 else 0.1
        
        lab_papers_to_add = int(top_k * lab_paper_ratio)
        preferred_author_papers_to_add = int(top_k * 0.1)  # 10% for preferred authors
        other_papers_to_add = top_k - lab_papers_to_add - preferred_author_papers_to_add
        
        print(f"üìä Selection targets: {lab_papers_to_add} lab papers, {preferred_author_papers_to_add} preferred author papers, {other_papers_to_add} other papers")

        # Add lab papers first (highest priority)
        for result in lab_papers[:lab_papers_to_add]:
            selected_papers.append(result['document'])
            # Track usage
            paper_id = result['paper_id']
            self.paper_usage_count[paper_id] = self.paper_usage_count.get(paper_id, 0) + 1

        # Add preferred author papers
        for result in preferred_author_papers[:preferred_author_papers_to_add]:
            selected_papers.append(result['document'])
            # Track usage
            paper_id = result['paper_id']
            self.paper_usage_count[paper_id] = self.paper_usage_count.get(paper_id, 0) + 1

        # Add other papers
        for result in other_papers[:other_papers_to_add]:
            selected_papers.append(result['document'])
            # Track usage
            paper_id = result['paper_id']
            self.paper_usage_count[paper_id] = self.paper_usage_count.get(paper_id, 0) + 1

        # Step 7: Report deprioritization statistics
        used_papers_in_selection = sum(1 for result in deprioritized_results[:top_k] if result['usage_count'] > 0)
        print(f"üìä Deprioritization results: {used_papers_in_selection}/{top_k} papers previously used")
        
        if self.paper_usage_count:
            total_unique_papers_used = len(self.paper_usage_count)
            print(f"üìä Total unique papers used across all hypotheses: {total_unique_papers_used}")

        # Validate we have enough papers
        if len(selected_papers) < top_k:
            print(f"‚ö†Ô∏è  Warning: Only found {len(selected_papers)} papers (requested {top_k})")
        elif len(selected_papers) > top_k:
            print(f"üìä Note: Selected {len(selected_papers)} papers (requested {top_k})")
        
        # Return the selected papers
        return selected_papers

    def _is_preferred_author_paper(self, metadata, preferred_authors):
        """Check if a paper is authored by preferred authors."""
        if not preferred_authors:
            return False
        
        authors = metadata.get('author', metadata.get('authors', ''))
        if not authors:
            return False
        
        # Convert to string if it's a list
        if isinstance(authors, list):
            authors = '; '.join(authors)
        
        authors_lower = str(authors).lower()
        
        for preferred_author in preferred_authors:
            if preferred_author.lower() in authors_lower:
                return True
        
        return False

    def detect_preprint_status(self, metadata):
        """
        Detect preprint status from existing metadata fields.
        
        Args:
            metadata: Paper metadata dictionary
            
        Returns:
            tuple: (is_preprint: bool, preprint_type: str, preprint_score: float)
        """
        preprint_score = 0.0
        preprint_type = "unknown"
        
        # Check source field first (most reliable)
        source = metadata.get('source', '').lower()
        if source in ["biorxiv", "medrxiv", "arxiv", "chemrxiv", "preprints"]:
            preprint_score += 0.8
            preprint_type = source
        
        # Check journal name for preprint indicators
        journal = metadata.get('journal', '').lower()
        if journal:
            preprint_indicators = [
                "arxiv", "biorxiv", "medrxiv", "chemrxiv", "preprint", 
                "working paper", "manuscript", "draft"
            ]
            for indicator in preprint_indicators:
                if indicator in journal:
                    preprint_score += 0.6
                    preprint_type = "preprint_server"
                    break
        
        # Check if it's from a preprint server origin
        origin = metadata.get('origin', '').lower()
        if origin in ["biorxiv", "medrxiv", "chemrxiv", "preprints"]:
            preprint_score += 0.9
            preprint_type = origin
        
        # Check version information
        version = metadata.get('version', '')
        if version and version != "1":
            preprint_score += 0.3
            preprint_type = "versioned_preprint"
        
        # Check status
        status = metadata.get('status', '').lower()
        if status in ["submitted", "pending", "working"]:
            preprint_score += 0.4
            preprint_type = "submitted_preprint"
        
        # Determine final classification
        is_preprint = preprint_score >= 0.5
        
        return is_preprint, preprint_type, preprint_score

    def add_preprint_metadata(self, metadata):
        """
        Add preprint metadata to paper metadata if not present.
        
        Args:
            metadata: Paper metadata dictionary
            
        Returns:
            dict: Updated metadata with preprint information
        """
        if 'is_preprint' not in metadata:
            is_preprint, preprint_type, preprint_score = self.detect_preprint_status(metadata)
            metadata['is_preprint'] = is_preprint
            metadata['preprint_type'] = preprint_type
            metadata['preprint_score'] = preprint_score
        
        return metadata

    def _get_journal_impact_factor(self, journal_name):
        """
        Get journal impact factor (simplified implementation).
        In a real system, this would query a journal impact factor database.
        """
        # Simplified impact factor mapping - in practice, this would be a database lookup
        impact_factors = {
            'nature': 49.962,
            'science': 56.9,
            'cell': 66.85,
            'nature medicine': 87.241,
            'nature biotechnology': 68.164,
            'nature genetics': 41.307,
            'nature cell biology': 28.213,
            'nature immunology': 31.25,
            'nature reviews immunology': 108.555,
            'immunity': 43.474,
            'journal of immunology': 5.422,
            'journal of experimental medicine': 17.579,
            'proceedings of the national academy of sciences': 12.779,
            'plos one': 3.752,
            'bioinformatics': 6.937,
            'nucleic acids research': 19.16,
            'genome research': 11.093,
            'genome biology': 17.906,
            'cell reports': 9.995,
            'molecular cell': 19.328,
            'developmental cell': 13.417,
            'current biology': 10.834,
            'elife': 8.713,
            'plos biology': 9.593,
            'plos genetics': 6.02,
            'plos computational biology': 4.7,
            'bmc genomics': 4.317,
            'bmc bioinformatics': 3.169,
            'bioinformatics': 6.937,
            'nucleic acids research': 19.16,
            'genome research': 11.093,
            'genome biology': 17.906,
            'cell reports': 9.995,
            'molecular cell': 19.328,
            'developmental cell': 13.417,
            'current biology': 10.834,
            'elife': 8.713,
            'plos biology': 9.593,
            'plos genetics': 6.02,
            'plos computational biology': 4.7,
            'bmc genomics': 4.317,
            'bmc bioinformatics': 3.169
        }
        
        journal_lower = journal_name.lower()
        for key, impact in impact_factors.items():
            if key in journal_lower:
                return impact
                
        return 1.0  # Default impact factor for unknown journals

    def show_export_completion_popup(self, export_path):
        """Show a popup notification when hypothesis generation is complete."""
        try:
            # Create a root window if it doesn't exist
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            
            # Show the popup message
            messagebox.showinfo(
                "Hypothesis Generation Complete! üéâ",
                f"Your hypothesis generation has finished successfully!\n\n"
                f"üìÅ Results saved to:\n{export_path}\n\n"
                f"üìã Export includes:\n"
                f"‚Ä¢ metadata.json - Session metadata\n"
                f"‚Ä¢ metahypothesis_data.csv - Meta-hypothesis data\n"
                f"‚Ä¢ hypothesis.csv - Generated hypotheses\n"
                f"‚Ä¢ sources/ - Source chunks and data\n\n"
                f"Click OK to continue."
            )
            
            # Destroy the root window
            root.destroy()
            
        except Exception as e:
            print(f"‚ö†Ô∏è Could not show popup notification: {e}")
            print(f"üìÅ Results saved to: {export_path}")

    def create_comprehensive_export(self, base_filename=None):
        """Create a comprehensive export with the required folder structure."""
        from datetime import datetime
        import shutil
        
        # Create timestamp-based folder name
        timestamp = datetime.now().strftime("%m%d%Y-%H%M")
        export_folder_name = f"Hypothesis_Export_{timestamp}"
        
        # Create the main export directory
        export_dir = "hypothesis_export"
        os.makedirs(export_dir, exist_ok=True)
        
        # Create the specific export folder
        full_export_path = os.path.join(export_dir, export_folder_name)
        os.makedirs(full_export_path, exist_ok=True)
        
        # Create sources subdirectory
        sources_dir = os.path.join(full_export_path, "sources")
        os.makedirs(sources_dir, exist_ok=True)
        
        try:
            # 1. Create metadata.json
            metadata = {
                "export_timestamp": datetime.now().isoformat(),
                "total_hypotheses": len(self.hypothesis_records),
                "export_version": "1.0",
                "system_info": {
                    "chromadb_enabled": self.use_chromadb,
                    "total_chunks_processed": len(self.hypothesis_records) if self.hypothesis_records else 0
                }
            }
            
            metadata_file = os.path.join(full_export_path, "metadata.json")
            with open(metadata_file, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=2)
            
            # 2. Create hypothesis.csv
            if self.hypothesis_records:
                hypothesis_df = pd.DataFrame(self.hypothesis_records)
                hypothesis_file = os.path.join(full_export_path, "hypothesis.csv")
                hypothesis_df.to_csv(hypothesis_file, index=False, encoding='utf-8')
            
            # 3. Create metahypothesis_data.csv (if we have meta-hypothesis data)
            # For now, create an empty file - this can be populated by meta-hypothesis generation
            meta_hypothesis_file = os.path.join(full_export_path, "metahypothesis_data.csv")
            pd.DataFrame(columns=['MetaHypothesis', 'Generated_Hypotheses', 'Timestamp']).to_csv(
                meta_hypothesis_file, index=False, encoding='utf-8'
            )
            
            # 4. Export source chunks to sources/ directory with hypothesis associations
            if hasattr(self, 'last_context_chunks') and self.last_context_chunks:
                chunks_file = os.path.join(sources_dir, "chunks.json")
                chunks_data = {
                    "total_chunks": len(self.last_context_chunks),
                    "chunks": []
                }
                
                for i, chunk in enumerate(self.last_context_chunks):
                    chunk_data = {
                        "chunk_id": i,
                        "content": chunk.get("document", "") if isinstance(chunk, dict) else str(chunk),
                        "metadata": chunk.get("metadata", {}) if isinstance(chunk, dict) else {}
                    }
                    chunks_data["chunks"].append(chunk_data)
                
                with open(chunks_file, 'w', encoding='utf-8') as f:
                    json.dump(chunks_data, f, indent=2, ensure_ascii=False)
            
            # 5. Store citation chunks for each hypothesis
            hypothesis_citation_mappings = []
            if self.hypothesis_records:
                for i, record in enumerate(self.hypothesis_records):
                    hypothesis_text = record.get('Hypothesis', '')
                    if hypothesis_text and hypothesis_text != 'Error in attempt':
                        # Use raw citations if available, otherwise parse from text
                        citations = record.get('Raw_Citations', [])
                        if not citations:
                            # Fallback to parsing from formatted text
                            citations_text = record.get('Citations', '')
                            if citations_text and citations_text != 'No citations available':
                                citations = self._parse_citations_from_text(citations_text)
                        
                        if citations:
                            # Store citation chunks for this hypothesis
                            mapping_filepath, citation_files = self.store_citation_chunks_for_hypothesis(
                                hypothesis_text, citations, i, full_export_path
                            )
                            hypothesis_citation_mappings.append({
                                "hypothesis_index": i,
                                "hypothesis_text": hypothesis_text,
                                "mapping_filepath": mapping_filepath,
                                "citation_files": citation_files,
                                "citation_count": len(citations)
                            })
            
            # 6. Create master citation mapping file
            if hypothesis_citation_mappings:
                master_mapping_file = os.path.join(sources_dir, "master_hypothesis_citation_mapping.json")
                master_mapping_data = {
                    "total_hypotheses": len(hypothesis_citation_mappings),
                    "created_at": datetime.now().isoformat(),
                    "hypothesis_mappings": hypothesis_citation_mappings
                }
                with open(master_mapping_file, 'w', encoding='utf-8') as f:
                    json.dump(master_mapping_data, f, indent=2, ensure_ascii=False)
            
            # 7. Create a summary file
            summary_file = os.path.join(full_export_path, "export_summary.txt")
            with open(summary_file, 'w', encoding='utf-8') as f:
                f.write(f"Hypothesis Generation Export Summary\n")
                f.write(f"=====================================\n\n")
                f.write(f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Hypotheses: {len(self.hypothesis_records)}\n")
                f.write(f"Hypotheses with Citations: {len(hypothesis_citation_mappings)}\n")
                f.write(f"Export Location: {full_export_path}\n\n")
                f.write(f"Files included:\n")
                f.write(f"- metadata.json: Session metadata\n")
                f.write(f"- hypothesis.csv: Generated hypotheses with scores\n")
                f.write(f"- metahypothesis_data.csv: Meta-hypothesis data\n")
                f.write(f"- sources/chunks.json: Source chunks used\n")
                f.write(f"- sources/master_hypothesis_citation_mapping.json: Master citation mapping\n")
                f.write(f"- sources/hypothesis_X/: Individual hypothesis citation folders\n")
                f.write(f"  - citation_N_*.json: Individual citation files with chunk content\n")
                f.write(f"  - hypothesis_citation_mapping.json: Hypothesis-specific mapping\n")
                f.write(f"- export_summary.txt: This summary file\n\n")
                f.write(f"Citation System Features:\n")
                f.write(f"- Each hypothesis has its own citation folder\n")
                f.write(f"- Citations include full chunk content and metadata\n")
                f.write(f"- Citation cache keys are preserved for easy mapping\n")
                f.write(f"- Structured JSON format for programmatic access\n")
                f.write(f"- Use 'python scripts/analyze_citations.py' to analyze citation data\n")
            
            print(f"‚úÖ Comprehensive export created successfully!")
            print(f"üìÅ Export location: {full_export_path}")
            
            # Show popup notification
            self.show_export_completion_popup(full_export_path)
            
            return full_export_path
            
        except Exception as e:
            print(f"‚ùå Error creating comprehensive export: {e}")
            return None

def main():
    """Main function to run the enhanced RAG query system."""
    print("=== Enhanced RAG System Startup ===")
    print("üöÄ Starting Enhanced RAG System with ChromaDB...")

    # Initialize with ChromaDB enabled and fast startup
    rag_system = EnhancedRAGQuery(use_chromadb=True, load_data_at_startup=False)



    # Run interactive search
    rag_system.interactive_search()

if __name__ == "__main__":
    main()
